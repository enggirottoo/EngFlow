import os
import tempfile
import unittest

import openpyxl

from automocoes.planilha.planilha import (
    atualizar_planilha_local,
    construir_payload_planilha,
    detectar_pastas_onedrive,
)


class PlanilhaLocalTests(unittest.TestCase):
    def test_atualizar_planilha_local_insere_linha(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "test.xlsx")
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.append(["descricao", "status", "data_open", "linha", "usuario"])
            wb.save(workbook_path)

            payload = construir_payload_planilha("Descrição teste", linha="12", usuario="estagiaria")
            caminho = atualizar_planilha_local(payload, workbook_path=workbook_path)

            self.assertEqual(caminho, workbook_path)
            wb2 = openpyxl.load_workbook(workbook_path)
            self.assertEqual(wb2.active.max_row, 2)
            linha_valores = [cell.value for cell in wb2.active[2]]
            self.assertEqual(linha_valores, ["Descrição teste", "ON GOING", payload["data_open"], "12", "estagiaria"])

    def test_detectar_pastas_onedrive(self):
        env_backup = {key: os.environ.get(key) for key in ("OneDrive", "OneDriveCommercial", "OneDriveConsumer")}
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                os.environ["OneDrive"] = tmpdir
                pastas = detectar_pastas_onedrive()
                self.assertIn(tmpdir, pastas)
        finally:
            for key, value in env_backup.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value


if __name__ == "__main__":
    unittest.main()
