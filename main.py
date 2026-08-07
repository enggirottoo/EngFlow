
import sys
import os


def _garantir_chromium_instalado():
    """Verifica se o Chromium do Playwright está instalado; se não, instala automaticamente."""
    browsers_path = os.path.join(os.environ.get("LOCALAPPDATA", ""), "ms-playwright")
    # Verificar se alguma pasta chromium-* existe
    if os.path.isdir(browsers_path):
        for item in os.listdir(browsers_path):
            if item.startswith("chromium-"):
                chromium_dir = os.path.join(browsers_path, item)
                if os.path.isdir(chromium_dir) and os.listdir(chromium_dir):
                    return  # Chromium já instalado
    
    print("=" * 60)
    print("  PRIMEIRA EXECUÇÃO — Instalando navegador Chromium...")
    print("  (isso acontece apenas uma vez, aguarde ~1 minuto)")
    print("=" * 60)
    
    try:
        from playwright._impl._driver import compute_driver_executable
        driver_executable, cli_js = compute_driver_executable()
        import subprocess
        result = subprocess.run(
            [driver_executable, cli_js, "install", "chromium"],
            capture_output=False,
        )
        if result.returncode == 0:
            print("\n[OK] Chromium instalado com sucesso!\n")
        else:
            print(f"\n[AVISO] Instalação retornou código {result.returncode}")
    except Exception as exc:
        print(f"\n[AVISO] Falha ao instalar Chromium automaticamente: {exc}")
        print("Tente instalar manualmente: playwright install chromium")

def _dispatch_automacao() -> bool:
    """Verifica se o EXE foi chamado em modo automação. Retorna True se sim."""
    if "--automacao" not in sys.argv:
        return False

    args = sys.argv[1:]
    idx = args.index("--automacao")
    if idx + 1 >= len(args):
        return False

    nome = args[idx + 1]

    # Parsear argumentos
    arg_valor = None
    cred_user = os.environ.get("PHOENIX_CRED_USER")
    cred_pass = os.environ.get("PHOENIX_CRED_PASS")
    for i, a in enumerate(args):
        if a == "--arg" and i + 1 < len(args):
            arg_valor = args[i + 1]
        elif a == "--cred-user" and i + 1 < len(args):
            cred_user = cred_user or args[i + 1]
        elif a == "--cred-pass" and i + 1 < len(args):
            cred_pass = cred_pass or args[i + 1]

    # No Windows, se o EXE é windowed (console=False), criar console visível (a menos que seja automação invisível)
    if os.name == "nt" and "--invisible" not in sys.argv:
        try:
            import ctypes
            kernel32 = ctypes.windll.kernel32
            kernel32.AllocConsole()
            # Reconectar stdout/stderr ao novo console
            sys.stdout = open("CONOUT$", "w", encoding="utf-8", errors="replace")
            sys.stderr = open("CONOUT$", "w", encoding="utf-8", errors="replace")
            sys.stdin = open("CONIN$", "r", encoding="utf-8", errors="replace")
            # Título do console
            kernel32.SetConsoleTitleW(f"Flex-tax classification — Automação {nome.upper()}")
        except Exception:
            pass

    # Reconstruir sys.argv como a automação espera
    new_argv = ["automacao"]
    if arg_valor:
        new_argv.append(arg_valor)
    if cred_user:
        new_argv.extend(["--cred-user", cred_user])
    if cred_pass:
        new_argv.extend(["--cred-pass", cred_pass])
    sys.argv = new_argv

    # Garantir que o navegador Chromium está instalado
    try:
        _garantir_chromium_instalado()
    except Exception as exc:
        print(f"[AVISO] Não foi possível verificar/instalar Chromium: {exc}")

    # Despachar para a automação correta
    try:
        if nome == "phoenix":
            from automocoes.phoenix.phoenix import abrir_home_phoenix, nova_solicitacao_phoenix, sincronizar_solicitacoes_phoenix
            if arg_valor == "home":
                abrir_home_phoenix()
            elif arg_valor in ("capturar", "importar"):
                sincronizar_solicitacoes_phoenix()
            else:
                nova_solicitacao_phoenix()

        elif nome == "pegasus":
            from automocoes.pegasus.pegasus import abrir_home_pegasus, nova_solicitacao_pegasus
            if arg_valor == "home":
                abrir_home_pegasus()
            else:
                nova_solicitacao_pegasus()

        elif nome == "cost":
            from automocoes.cost.cost import nova_solicitacao_cost
            nova_solicitacao_cost()

        elif nome == "atualizar_pn":
            from automocoes.phoenix.atualizar_pn import atualizar_pn_phoenix
            atualizar_pn_phoenix(arg_valor)

        else:
            print(f"[ERRO] Automação desconhecida: {nome}")
            sys.exit(1)

    except Exception as exc:
        print(f"\n[ERRO] Falha na automação '{nome}': {exc}")
        import traceback
        traceback.print_exc()
        input("\nPressione ENTER para fechar...")
        sys.exit(1)

    return True


if _dispatch_automacao():
    sys.exit(0)

# =====================================================================
# A PARTIR DAQUI: imports normais da GUI (só carrega se NÃO for automação)
# =====================================================================
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime
from typing import Any, Dict, List, Optional
import threading
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

# matplotlib é importado sob demanda em _renderizar_dashboard_conteudo
# para reduzir consumo de memória no startup (~50 MB a menos).

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.paths import get_base_dir, get_db_path, set_db_path
from core.runner import executar_script, logger
# buscar_pn_por_ticket importado sob demanda em _atualizar_pn_dashboard
# para evitar carregar playwright no startup da GUI
from services.logging_config import registrar_log_auditoria, configurar_logger
from services.backup import fazer_backup
from services.database import (
    inicializar_banco,
    validar_usuario,
    validar_usuario_detalhado,
    iniciar_pegasus,
    finalizar_pegasus,
    iniciar_custo,
    finalizar_custo,
    finalizar_phoenix,
    adicionar_usuario,
    editar_usuario,
    desativar_usuario,
    reativar_usuario,
    excluir_usuario,
    esvaziar_lixeira,
    listar_usuarios,
    obter_usuario_por_id,
    obter_credencial,
    salvar_credencial,
    obter_credenciais_usuario,
    usuario_possui_pin_credenciais,
    definir_pin_credenciais,
    resetar_pin_credenciais,
    credencial_configurada,
    contar_por_status,
    exportar_usuarios_json,
    importar_usuarios_json,
    obter_e_limpar_codigos_importados,
    obter_estatisticas_banco,
    obter_e_limpar_primeiro_admin_gerado,
    listar_notificacoes,
    contar_notificacoes_nao_lidas,
    marcar_notificacao_lida,
    marcar_todas_notificacoes_lidas,
    verificar_e_gerar_alertas_inatividade,
    registrar_tentativa_bloqueada,
)
from services.backup import fazer_backup, listar_backups, verificar_e_criar_backup_diario
from services.importador_planilha import importar_excel_historico
from services.storage import (
    atualizar_campos_registro,
    carregar_config,
    carregar_estado_app,
    carregar_historico,
    encontrar_por_linha,
    salvar_estado_app,
    cancelar_registro,
)

from ui import theme
from ui.theme import (
    ACCENT,
    ACCENT_HOVER,
    ACCENT_SOFT,
    BG,
    BG_CARD,
    BORDA,
    ENTRY_BG,
    FONT_BOTAO,
    FONT_CAPTION,
    FONT_CARD_TITULO,
    FONT_SUBTITULO,
    FONT_TITULO,
    FOOTER_BG,
    HEADER_BG,
    TEMA_CLARO,
    TEMA_ESCURO,
    TEXTO,
    TEXTO_MUTED,
    definir_tema,
    espacar,
    tema_atual,
)

# Inicializar o banco SQLite na startup
inicializar_banco()

FRAME_LOGIN = "frame_login"
FRAME_MENU = "frame_menu"
FRAME_DASHBOARD = "frame_dashboard"
FRAME_MENU_PHOENIX = "frame_menu_phoenix"
FRAME_MENU_PEGASUS = "frame_menu_pegasus"
FRAME_ATUALIZAR_PN = "frame_atualizar_pn"
FRAME_ADMIN = "frame_admin"
FRAME_CREDENCIAIS = "frame_credenciais"

APP_VERSION = "1.0.0"


def mapear_tela_para_nav(nome: str) -> Optional[str]:
    mapa = {
        FRAME_DASHBOARD: "dashboard",
        FRAME_MENU_PHOENIX: "phoenix",
        FRAME_MENU_PEGASUS: "pegasus",
        FRAME_ATUALIZAR_PN: "phoenix",
        FRAME_ADMIN: "admin",
        FRAME_CREDENCIAIS: "credenciais",
    }
    return mapa.get(nome)


def resumo_ultimo_registro(historico: List[Dict[str, Any]]) -> str:
    if not historico:
        return "Último registro: nenhum"

    item = historico[-1]
    descricao = str(item.get("description") or "Sem description").strip() or "Sem description"
    linha = item.get("linha", "—")
    status = str(item.get("status", "")).strip().upper()

    if status == "ON GOING":
        status_texto = "Em andamento"
    elif status == "CANCELADO":
        status_texto = "Cancelado"
    elif status:
        status_texto = "Finalizado"
    else:
        status_texto = "Sem status"

    return f"Último registro: {descricao} • Linha {linha} • {status_texto}"


class PhoenixTool:

    def _finalizar_phoenix(self, item):
        if not self._validar_permissao("editar"):
            return
        finalizar_phoenix(item["linha"])
        self._abrir_dashboard()

    def _iniciar_pegasus(self, item):
        if not self._validar_permissao("pegasus"):
            return
        iniciar_pegasus(item["linha"])
        self._abrir_dashboard()

    def _finalizar_pegasus(self, item):
        if not self._validar_permissao("pegasus"):
            return
        finalizar_pegasus(item["linha"])
        self._abrir_dashboard()

    def _iniciar_custo(self, item):
        if not self._validar_permissao("cost"):
            return
        iniciar_custo(item["linha"])
        self._abrir_dashboard()

    def _finalizar_custo(self, item):
        if not self._validar_permissao("cost"):
            return
        finalizar_custo(item["linha"])
        self._abrir_dashboard()

    def _excluir_registro(self, item):
        if not self._validar_permissao("excluir"):
            return

        confirmar = messagebox.askyesno(
            "Excluir Registro",
            (
                f"Deseja excluir a linha {item.get('linha')}?\n\n"
                f"Descrição:\n{item.get('description')}\n\n"
                "Esta ação ocultará o registro do dashboard."
            )
        )

        if not confirmar:
            return

        sucesso = cancelar_registro(
            item.get("uuid"),
            "Excluído manualmente pelo usuário",
            usuario_cancelamento=self.usuario,
        )

        if sucesso:
            messagebox.showinfo("Sucesso", "Registro excluído com sucesso.")
            if FRAME_DASHBOARD in self.frames:
                self._build_dashboard()
                self.mostrar(FRAME_DASHBOARD)
        else:
            messagebox.showerror("Erro", "Não foi possível excluir o registro.")

    def __init__(self, root):
        self.root = root
        self.root.overrideredirect(True)
        self._drag_x = 0
        self._drag_y = 0
        self.root.title("Flex-tax classification 1.0")
        self.root.resizable(True, True)
        self.root.minsize(560, 720)
        self.root.state("zoomed")
        self._modo_tela_cheia = False
        self.root.option_add("*Font", "Arial")
        self.root.option_add("*tearOff", False)
        self.root.attributes("-alpha", 0.98)
        self.root.overrideredirect(False)
        self.root.bind("<ButtonPress-1>", self._iniciar_arraste)
        self.root.bind("<B1-Motion>", self._arrastar_janela)

        self.usuario = None
        self.usuario_id = None
        self.permissao = None
        self._cred_pin = None
        self.history = []
        self.frames = {}
        self._dashboard_filter = "todos"
        self._dashboard_last_signature = None
        self._cards = {}
        self._active_card_key = None

        estado_inicial = carregar_estado_app()
        theme.definir_tema(estado_inicial.get("theme") or theme.TEMA_ESCURO)
        self._sincronizar_cores()

        self.root.configure(bg=BG)
        self._montar_estrutura_chrome()

        self._build_login()
        self._show_loading_state()
        self.root.after(350, self._finalizar_inicializacao)

    def _sincronizar_cores(self):
        global BG, BG_CARD, BORDA, ENTRY_BG, TEXTO, TEXTO_MUTED, ACCENT, ACCENT_HOVER, ACCENT_SOFT, HEADER_BG, FOOTER_BG
        BG = theme.BG
        BG_CARD = theme.BG_CARD
        BORDA = theme.BORDA
        ENTRY_BG = theme.ENTRY_BG
        TEXTO = theme.TEXTO
        TEXTO_MUTED = theme.TEXTO_MUTED
        ACCENT = theme.ACCENT
        ACCENT_HOVER = theme.ACCENT_HOVER
        ACCENT_SOFT = theme.ACCENT_SOFT
        HEADER_BG = theme.HEADER_BG
        FOOTER_BG = theme.FOOTER_BG

    def _validar_permissao(self, acao: str) -> bool:
        """Validação interna estrita de permissões antes de qualquer ação crítica."""
        if not self.permissao:
            messagebox.showerror("Acesso Negado", "Usuário não autenticado.")
            return False

        perm_upper = str(self.permissao).upper()

        if perm_upper == "ADMIN":
            return True

        if perm_upper == "ENGENHARIA":
            permitidos = {
                "dashboard", "nova_solicitacao", "phoenix", "pegasus", "cost",
                "editar", "atualizar_pn", "exportar", "credenciais"
            }
            if acao in permitidos:
                return True

        if perm_upper == "VISITANTE":
            permitidos = {"dashboard", "consulta", "exportar"}
            if acao in permitidos:
                return True

        messagebox.showerror("Acesso Negado", f"Seu perfil ({perm_upper}) não tem permissão para a ação: {acao}.")
        return False

    def _montar_estrutura_chrome(self):
        self._aplicar_estilo()

        self.header = tk.Frame(self.root, bg=HEADER_BG, height=52, bd=0, highlightthickness=0)
        self.header.pack(fill="x", side="top")
        self.header.pack_propagate(False)

        tk.Label(self.header, text="FLEX • CLASSIFICAÇÃO FISCAL", bg=HEADER_BG, fg=ACCENT, font=("Arial", 10, "bold")).pack(side="left", padx=20)
        texto_badge = f"USUÁRIO: {self.usuario.upper()} ({self.permissao})" if self.usuario else "USUÁRIO: NÃO LOGADO"
        self.user_badge = tk.Label(self.header, text=texto_badge, bg=HEADER_BG, fg=TEXTO_MUTED, font=("Arial", 8))
        self.user_badge.pack(side="left", padx=(12, 0))
        tk.Label(self.header, text="STATUS: ONLINE", bg=HEADER_BG, fg=TEXTO_MUTED, font=("Arial", 9)).pack(side="right", padx=20)

        help_btn = tk.Button(
            self.header,
            text="❓",
            bg=HEADER_BG,
            fg=TEXTO_MUTED,
            activebackground=HEADER_BG,
            activeforeground=ACCENT,
            font=("Arial", 12),
            bd=0,
            highlightthickness=0,
            cursor="hand2",
            command=self._abrir_ajuda,
        )
        help_btn.pack(side="right", padx=(0, 2))

        self.notif_btn = tk.Button(
            self.header,
            text="🔔",
            bg=HEADER_BG,
            fg=TEXTO_MUTED,
            activebackground=HEADER_BG,
            activeforeground=ACCENT,
            font=("Arial", 12),
            bd=0,
            highlightthickness=0,
            cursor="hand2",
            command=self._abrir_notificacoes,
        )
        self.notif_btn.pack(side="right", padx=(0, 4))
        self._notif_after_id = None

        self.footer = tk.Frame(self.root, bg=FOOTER_BG, height=32, bd=0, highlightthickness=0)
        self.footer.pack(fill="x", side="bottom")
        self.footer.pack_propagate(False)

        self.nav_left = tk.Frame(self.footer, bg=FOOTER_BG)
        self.nav_left.pack(side="left", padx=20)
        self.nav_labels = {}

        nav_right = tk.Frame(self.footer, bg=FOOTER_BG)
        nav_right.pack(side="right", padx=20)
        self.footer_status = tk.Label(nav_right, text="PRONTO", bg=FOOTER_BG, fg=TEXTO_MUTED, font=("Arial", 8))
        self.footer_status.pack(side="right")
        tk.Label(nav_right, text="  •  ", bg=FOOTER_BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="right")
        tk.Label(nav_right, text="FLEX-TAX CLASSIFICATION 1.0", bg=FOOTER_BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="right")

        self.signature_frame = tk.Frame(self.root, bg=BG)
        self.signature_frame.pack(side="bottom", pady=6)

        self.signature_text = tk.Label(
            self.signature_frame,
            text="Developed by ",
            bg=BG,
            fg=TEXTO_MUTED,
            font=("Arial", 8)
        )
        self.signature_text.pack(side="left")

        self.signature_name = tk.Label(
            self.signature_frame,
            text="Gabriel Girotto",
            bg=BG,
            fg=TEXTO,
            font=("Arial", 8, "bold")
        )
        self.signature_name.pack(side="left")

        self.area_rolagem = tk.Frame(self.root, bg=BG)
        self.area_rolagem.pack(fill="both", expand=True)

        self.main_canvas = tk.Canvas(
            self.area_rolagem,
            bg=BG,
            highlightthickness=0,
            bd=0
        )

        self.vscrollbar = tk.Scrollbar(
            self.area_rolagem,
            orient="vertical",
            command=self.main_canvas.yview
        )

        self.main_canvas.configure(
            yscrollcommand=self.vscrollbar.set
        )

        self.main_canvas.pack(
            side="left",
            fill="both",
            expand=True
        )

        self.container = tk.Frame(
            self.main_canvas,
            bg=BG
        )

        self.container.grid_columnconfigure(0, weight=1)

        self._container_window = self.main_canvas.create_window(
            (0, 0),
            window=self.container,
            anchor="nw"
        )

        def _atualizar_scrollregion(event=None):
            self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

        def _ajustar_largura_container(event):
            largura = event.width
            self.main_canvas.itemconfig(self._container_window, width=largura)

        self.container.bind("<Configure>", _atualizar_scrollregion)
        self.main_canvas.bind("<Configure>", _ajustar_largura_container)

        def _scroll_vertical(event):
            self.main_canvas.yview_scroll(int(-2 * (event.delta / 120)), "units")

        def _scroll_horizontal(event):
            self.main_canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_scroll(event=None):
            self.main_canvas.bind_all("<MouseWheel>", _scroll_vertical)
            self.main_canvas.bind_all("<Shift-MouseWheel>", _scroll_horizontal)

        def _unbind_scroll(event=None):
            self.main_canvas.unbind_all("<MouseWheel>")
            self.main_canvas.unbind_all("<Shift-MouseWheel>")

        self.main_canvas.bind("<Enter>", _bind_scroll)
        self.main_canvas.bind("<Leave>", _unbind_scroll)

    # =========================================================================
    # Notificações
    # =========================================================================

    def _atualizar_badge_notificacoes(self):
        """Atualiza o texto do sino com a quantidade de notificações não lidas e verifica inatividade."""
        if not hasattr(self, "notif_btn") or not self.notif_btn.winfo_exists():
            return

        try:
            verificar_e_gerar_alertas_inatividade()
        except Exception as exc:
            logger.warning("Erro no monitoramento de inatividade de produtos: %s", exc)

        qtde = contar_notificacoes_nao_lidas(self.usuario) if self.usuario else 0

        if qtde > 0:
            texto = f"🔔 ({qtde if qtde < 100 else '99+'})"
            self.notif_btn.config(text=texto, fg="#ff5555")
        else:
            self.notif_btn.config(text="🔔", fg=TEXTO_MUTED)

        self._notif_after_id = self.root.after(30000, self._atualizar_badge_notificacoes)

    def _abrir_notificacoes(self):
        """Abre um painel com as notificações do usuário logado."""
        if not self.usuario:
            return

        try:
            verificar_e_gerar_alertas_inatividade()
        except Exception:
            pass

        notificacoes = listar_notificacoes(self.usuario)

        janela = tk.Toplevel(self.root)
        janela.title("Notificações")
        janela.configure(bg=BG)
        janela.geometry("460x420")
        janela.transient(self.root)

        cabecalho = tk.Frame(janela, bg=BG)
        cabecalho.pack(fill="x", padx=12, pady=(12, 6))
        tk.Label(cabecalho, text="Notificações", bg=BG, fg=TEXTO, font=FONT_SUBTITULO).pack(side="left")

        def _marcar_todas():
            marcar_todas_notificacoes_lidas(self.usuario)
            self._atualizar_badge_notificacoes_imediato()
            janela.destroy()
            self._abrir_notificacoes()

        tk.Button(
            cabecalho, text="Marcar todas como lidas", bg=BG_CARD, fg=TEXTO, bd=0,
            font=FONT_CAPTION, cursor="hand2", command=_marcar_todas,
        ).pack(side="right")

        area = tk.Frame(janela, bg=BG)
        area.pack(fill="both", expand=True, padx=12, pady=6)

        canvas = tk.Canvas(area, bg=BG, highlightthickness=0, bd=0)
        scrollbar = tk.Scrollbar(area, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        lista = tk.Frame(canvas, bg=BG)
        canvas.create_window((0, 0), window=lista, anchor="nw")

        def _atualizar_scrollregion(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
        lista.bind("<Configure>", _atualizar_scrollregion)

        if not notificacoes:
            tk.Label(lista, text="Nenhuma notificação por aqui.", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(pady=20)
        else:
            for notif in notificacoes:
                self._montar_item_notificacao(lista, notif, janela)

    def _montar_item_notificacao(self, parent, notif, janela_pai):
        lida = bool(notif.get("lida"))
        desc = str(notif.get("descricao") or "")

        # Níveis de criticidade visual
        borda_cor = BORDA
        if "🔴" in desc or "Vermelho" in desc:
            borda_cor = "#e74c3c"
        elif "🟠" in desc or "Laranja" in desc:
            borda_cor = "#e67e22"
        elif "🟡" in desc or "Amarelo" in desc:
            borda_cor = "#f1c40f"

        card = tk.Frame(parent, bg=BG_CARD if not lida else BG, bd=1, relief="solid", highlightbackground=borda_cor)
        card.pack(fill="x", pady=4)

        ticket = notif.get("ticket") or ""
        produto = notif.get("produto") or ""
        cabecalho_txt = f"{produto}" + (f"  •  Ticket #{ticket}" if ticket else "")
        lbl_hdr = tk.Label(card, text=cabecalho_txt, bg=card["bg"], fg=TEXTO, font=("Arial", 9, "bold"), anchor="w", justify="left")
        lbl_hdr.pack(fill="x", padx=10, pady=(8, 0))

        autor = notif.get("usuario_origem") or "?"
        quando = notif.get("data_hora") or ""
        lbl_desc = tk.Label(
            card, text=desc, bg=card["bg"], fg=TEXTO, font=("Arial", 8), anchor="w", justify="left", wraplength=380
        )
        lbl_desc.pack(fill="x", padx=10, pady=(2, 0))

        lbl_sub = tk.Label(
            card, text=f"Origem: {autor} em {quando}", bg=card["bg"], fg=TEXTO_MUTED, font=("Arial", 7), anchor="w"
        )
        lbl_sub.pack(fill="x", padx=10, pady=(2, 8))

        def _ir_para_ticket(e=None, t=ticket, nid=notif.get("id")):
            if nid and not lida:
                marcar_notificacao_lida(nid)
                self._atualizar_badge_notificacoes_imediato()
            try:
                janela_pai.destroy()
            except Exception:
                pass

            if hasattr(self, "notebook") and self.notebook:
                try:
                    self.notebook.select(0)
                except Exception:
                    pass

            if t and hasattr(self, "campo_busca") and self.campo_busca.winfo_exists():
                try:
                    self.campo_busca.delete(0, "end")
                    self.campo_busca.insert(0, t)
                    self._dashboard_search_query = t.strip().lower()
                    if hasattr(self, "_carregar_dashboard"):
                        self._carregar_dashboard()
                except Exception as exc:
                    logger.warning("Erro ao filtrar ticket no Dashboard: %s", exc)

        for w in (card, lbl_hdr, lbl_desc, lbl_sub):
            w.config(cursor="hand2")
            w.bind("<Button-1>", _ir_para_ticket)

        if not lida:
            def _marcar(e=None, nid=notif.get("id")):
                marcar_notificacao_lida(nid)
                self._atualizar_badge_notificacoes_imediato()
                janela_pai.destroy()
                self._abrir_notificacoes()

            tk.Button(
                card, text="Marcar como lida", bg=BG_CARD, fg=ACCENT, bd=0,
                font=("Arial", 7, "underline"), cursor="hand2", command=_marcar,
            ).pack(anchor="e", padx=10, pady=(0, 6))

    def _atualizar_badge_notificacoes_imediato(self):
        """Atualiza o badge sem esperar o próximo ciclo do polling."""
        if hasattr(self, "notif_btn") and self.notif_btn.winfo_exists():
            qtde = contar_notificacoes_nao_lidas(self.usuario) if self.usuario else 0
            if qtde > 0:
                self.notif_btn.config(text=f"🔔 ({qtde if qtde < 100 else '99+'})", fg="#ff5555")
            else:
                self.notif_btn.config(text="🔔", fg=TEXTO_MUTED)

    def _alternar_tema(self):
        novo_tema = TEMA_CLARO if tema_atual() == TEMA_ESCURO else TEMA_ESCURO
        definir_tema(novo_tema)
        self._sincronizar_cores()

        self.root.config(bg=BG)
        if hasattr(self, "header") and self.header.winfo_exists():
            self.header.config(bg=HEADER_BG)
            for widget in self.header.winfo_children():
                if isinstance(widget, tk.Label):
                    if widget.cget("text").startswith("FLEX"):
                        widget.config(bg=HEADER_BG, fg=ACCENT)
                    else:
                        widget.config(bg=HEADER_BG, fg=TEXTO_MUTED)
        if hasattr(self, "footer") and self.footer.winfo_exists():
            self.footer.config(bg=FOOTER_BG)
            for widget in self.footer.winfo_children():
                if isinstance(widget, tk.Frame):
                    widget.config(bg=FOOTER_BG)
                    for child in widget.winfo_children():
                        if isinstance(child, tk.Label):
                            child.config(bg=FOOTER_BG, fg=TEXTO_MUTED)
        if hasattr(self, "nav_left") and self.nav_left.winfo_exists():
            self.nav_left.config(bg=FOOTER_BG)
        if hasattr(self, "signature_frame") and self.signature_frame.winfo_exists():
            self.signature_frame.config(bg=BG)
        if hasattr(self, "signature_text") and self.signature_text.winfo_exists():
            self.signature_text.config(bg=BG, fg=TEXTO_MUTED)
        if hasattr(self, "signature_name") and self.signature_name.winfo_exists():
            self.signature_name.config(bg=BG, fg=TEXTO)
        if hasattr(self, "area_rolagem") and self.area_rolagem.winfo_exists():
            self.area_rolagem.config(bg=BG)
        if hasattr(self, "main_canvas") and self.main_canvas.winfo_exists():
            self.main_canvas.config(bg=BG)
        if hasattr(self, "container") and self.container.winfo_exists():
            self.container.config(bg=BG)
        if hasattr(self, "btn_tema") and self.btn_tema.winfo_exists():
            self.btn_tema.config(
                text="☀️ Modo Claro" if novo_tema == TEMA_ESCURO else "🌙 Modo Escuro",
                bg=ACCENT_SOFT, fg=ACCENT, activebackground=ACCENT, activeforeground="#ffffff"
            )

        self._recriar_itens_nav()

        if getattr(self, "history", None):
            tela_atual = self.history[-1]
            if tela_atual == FRAME_DASHBOARD:
                self._build_dashboard()
                self.mostrar(FRAME_DASHBOARD, empilhar=False)
            elif tela_atual == FRAME_LOGIN:
                self._build_login()
                self.mostrar(FRAME_LOGIN, empilhar=False)
            elif tela_atual == FRAME_ADMIN:
                self._build_administracao()
                self.mostrar(FRAME_ADMIN, empilhar=False)
            elif tela_atual == FRAME_CREDENCIAIS:
                self._build_minhas_credenciais()
                self.mostrar(FRAME_CREDENCIAIS, empilhar=False)
            elif tela_atual == FRAME_MENU:
                self._build_menu()
                self.mostrar(FRAME_MENU, empilhar=False)

    def _recriar_itens_nav(self):
        """Atualiza a barra de navegação inferior de acordo com o perfil."""
        for widget in self.nav_left.winfo_children():
            widget.destroy()
        self.nav_labels = {}

        if not self.permissao:
            return

        perm = str(self.permissao).upper()
        self._criar_item_nav(self.nav_left, "DASHBOARD", "dashboard")

        if perm in ("ADMIN", "ENGENHARIA"):
            tk.Label(self.nav_left, text="  •  ", bg=FOOTER_BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="left")
            self._criar_item_nav(self.nav_left, "PHOENIX", "phoenix")
            tk.Label(self.nav_left, text="  •  ", bg=FOOTER_BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="left")
            self._criar_item_nav(self.nav_left, "PEGASUS", "pegasus")
            tk.Label(self.nav_left, text="  •  ", bg=FOOTER_BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="left")
            self._criar_item_nav(self.nav_left, "CREDENCIAIS", "credenciais")

        if perm == "ADMIN":
            tk.Label(self.nav_left, text="  •  ", bg=FOOTER_BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="left")
            self._criar_item_nav(self.nav_left, "ADMINISTRAÇÃO", "admin")

    def _iniciar_arraste(self, event):
        self._drag_offset = (event.x_root - self.root.winfo_x(), event.y_root - self.root.winfo_y())

    def _arrastar_janela(self, event):
        if getattr(self, "_drag_offset", None) is None:
            return
        x = event.x_root - self._drag_offset[0]
        y = event.y_root - self._drag_offset[1]
        self.root.geometry(f"+{x}+{y}")

    def _alternar_tela_cheia(self):
        if self.root.attributes("-fullscreen"):
            self.root.attributes("-fullscreen", False)
            self.root.geometry("1100x900")
            self._modo_tela_cheia = False
        else:
            self.root.attributes("-fullscreen", True)
            self._modo_tela_cheia = True

    def _criar_item_nav(self, parent, texto, chave):
        def _navegar():
            if chave == "dashboard":
                self._abrir_dashboard()
            elif chave == "phoenix":
                self._abrir_menu_phoenix()
            elif chave == "pegasus":
                self._abrir_menu_pegasus()
            elif chave == "credenciais":
                self._abrir_minhas_credenciais()
            elif chave == "admin":
                self._abrir_administracao()

        label = tk.Label(
            parent,
            text=texto,
            bg=FOOTER_BG,
            fg=TEXTO_MUTED,
            font=("Arial", 8, "bold"),
            cursor="hand2",
        )
        label.pack(side="left")
        label.bind("<Button-1>", lambda e: _navegar())
        self.nav_labels[chave] = label
        return label

    def _atualizar_nav(self, nome):
        ativo = mapear_tela_para_nav(nome)
        for chave, label in self.nav_labels.items():
            label.config(fg=ACCENT if chave == ativo else TEXTO_MUTED)
        if hasattr(self, "footer_status"):
            if self.usuario:
                self.footer_status.config(text=f"USUÁRIO: {self.usuario.upper()} ({self.permissao})  •  TELA: {nome}")
            else:
                self.footer_status.config(text=f"TELA: {nome}")

    def _aplicar_estilo(self):
        style = ttk.Style()
        if "clam" in style.theme_names():
            style.theme_use("clam")

        style.configure("TFrame", background=BG)
        style.configure("TLabel", background=BG, foreground=TEXTO)
        style.configure("Muted.TLabel", background=BG, foreground=TEXTO_MUTED)
        style.configure("TButton", background=BG_CARD, foreground=TEXTO, borderwidth=1, padding=(12, 10))
        style.map(
            "TButton",
            background=[("active", BORDA), ("pressed", BORDA)],
            foreground=[("active", TEXTO)]
        )
        style.configure("Accent.TButton", background=ACCENT_SOFT, foreground=TEXTO, borderwidth=1, padding=(12, 10))
        style.map(
            "Accent.TButton",
            background=[("active", "#263b57"), ("pressed", "#263b57")],
            foreground=[("active", TEXTO)]
        )
        style.configure("TEntry", fieldbackground=BG_CARD, foreground=TEXTO, borderwidth=1, padding=(8, 8))
        style.map("TEntry", fieldbackground=[("focus", "#1c2230")], foreground=[("focus", TEXTO)])

    def mostrar(self, nome, empilhar=True):
        if nome != FRAME_DASHBOARD:
            self._cancelar_timer_dashboard()
        for f in self.frames.values():
            f.pack_forget()
        if empilhar and (not self.history or self.history[-1] != nome):
            self.history.append(nome)
        if nome in self.frames:
            self.frames[nome].pack(fill="both", expand=True)
            self._animar_troca_tela(self.frames[nome])
            self._atualizar_nav(nome)
            self._atualizar_cards_ativos(nome)
            salvar_estado_app(nome)

    def voltar(self):
        if len(self.history) > 1:
            self.history.pop()
            anterior = self.history[-1]
            self.mostrar(anterior, empilhar=False)

    def _animar_troca_tela(self, frame):
        try:
            frame.configure(bg=BG)
            frame.update_idletasks()
        except Exception:
            pass
        self._animar_fade()

    def _animar_fade(self, alvo=0.98, passos=6, intervalo=18):
        try:
            self.root.attributes("-alpha", alvo)
        except Exception:
            pass

    def _show_loading_state(self):
        self._loading_frame = tk.Frame(self.container, bg=BG)
        self._loading_frame.pack(pady=(24, 0))

        self._loading_label = tk.Label(
            self._loading_frame,
            text="Inicializando ambiente operacional...",
            bg=BG,
            fg=TEXTO_MUTED,
            font=("Arial", 10),
        )
        self._loading_label.pack(anchor="w")

        self._loading_bar = tk.Frame(self._loading_frame, bg=ACCENT_SOFT, height=8)
        self._loading_bar.pack(fill="x", pady=(10, 0), ipady=2)
        self._loading_fill = tk.Frame(self._loading_bar, bg=ACCENT, width=30, height=8)
        self._loading_fill.pack(side="left")
        self._loading_fill.pack_propagate(False)
        self._loading_progress = 0
        self._animar_loading()

    def _animar_loading(self):
        loading_fill = getattr(self, "_loading_fill", None)
        if loading_fill is None or not loading_fill.winfo_exists():
            return
        self._loading_progress = (self._loading_progress + 2) % 101
        largura = 20 + (self._loading_progress % 80)
        loading_fill.configure(width=largura)
        if getattr(self, "_loading_frame", None) is not None and self._loading_frame.winfo_exists():
            self.root.after(50, self._animar_loading)

    def _finalizar_inicializacao(self):
        if getattr(self, "_loading_frame", None) is not None:
            self._loading_frame.destroy()
            self._loading_frame = None
            self._loading_fill = None

        self.mostrar(FRAME_LOGIN, empilhar=False)

    def linha_divisoria(self, parent):
        tk.Frame(parent, bg=BORDA, height=1).pack(fill="x", padx=40, pady=(0, 20))

    def cabecalho(self, parent, caption, titulo):
        tk.Label(
            parent, text=espacar(caption), bg=BG, fg=ACCENT, font=FONT_CAPTION
        ).pack(pady=(40, 6))

        tk.Label(
            parent, text=titulo.upper(), bg=BG, fg=TEXTO, font=FONT_TITULO
        ).pack(pady=(0, 12))

        self.linha_divisoria(parent)

    def botao_voltar(self, parent):
        lbl = tk.Label(
            parent, text=espacar("← Voltar"), bg=BG, fg=ACCENT,
            font=FONT_CAPTION, cursor="hand2"
        )
        lbl.pack(anchor="w", padx=40, pady=(20, 0))
        lbl.bind("<Enter>", lambda e: lbl.config(fg=TEXTO))
        lbl.bind("<Leave>", lambda e: lbl.config(fg=TEXTO_MUTED))
        lbl.bind("<Button-1>", lambda e: self.voltar())
        return lbl

    def botao_flat(self, parent, texto, comando, largura=None, primario=True):
        texto_exibicao = espacar(texto) if len(texto) <= 8 else texto.upper()
        bg_cor = ACCENT if primario else ACCENT_SOFT
        fg_cor = "#ffffff" if primario else ACCENT
        bg_hover = ACCENT_HOVER if primario else ACCENT

        btn = tk.Button(
            parent, text=texto_exibicao, command=comando,
            bg=bg_cor, fg=fg_cor, activebackground=bg_hover, activeforeground="#ffffff",
            relief="flat", bd=0, font=("Arial", 9, "bold"), cursor="hand2",
            width=largura, highlightthickness=0,
            padx=16, pady=8
        )
        def _on_enter(e):
            btn.config(bg=bg_hover, fg="#ffffff")
        def _on_leave(e):
            btn.config(bg=bg_cor, fg=fg_cor)
        btn.bind("<Enter>", _on_enter)
        btn.bind("<Leave>", _on_leave)
        return btn

    def _aplicar_estado_card(self, card, lbl_titulo, lbl_num, ativo):
        bg_card_hover = "#24242e" if tema_atual() == TEMA_ESCURO else "#f0f2f5"
        card.configure(bg=bg_card_hover if ativo else BG_CARD)
        lbl_titulo.configure(fg=ACCENT if ativo else TEXTO)
        lbl_num.configure(fg=ACCENT)

    def _atualizar_cards_ativos(self, nome):
        chave = mapear_tela_para_nav(nome)
        self._active_card_key = chave
        for card_key, widgets in self._cards.items():
            self._aplicar_estado_card(*widgets, ativo=(card_key == chave))

    def card_navegavel(self, parent, numero, titulo, comando, descricao=None, chave=None):
        card = tk.Frame(parent, bg=BG_CARD, cursor="hand2", highlightthickness=0, bd=0)
        card.pack(fill="x", pady=(0, 8))

        linha = tk.Frame(card, bg=BG_CARD, padx=16, pady=12)
        linha.pack(fill="x")

        lbl_num = tk.Label(
            linha, text=f".{numero:02d}", bg=BG_CARD, fg=ACCENT, font=("Arial", 11, "bold")
        )
        lbl_num.pack(side="left", padx=(0, 12))

        lbl_titulo = tk.Label(
            linha, text=titulo.upper(), bg=BG_CARD, fg=TEXTO, font=FONT_CARD_TITULO
        )
        lbl_titulo.pack(side="left")

        if descricao:
            tk.Label(
                linha,
                text=descricao,
                bg=BG_CARD,
                fg=TEXTO_MUTED,
                font=("Arial", 9),
            ).pack(side="right")

        if chave:
            self._cards[chave] = (card, lbl_titulo, lbl_num)

        bg_card_hover = "#24242e" if tema_atual() == TEMA_ESCURO else "#f4f6fa"

        def entrar(e=None):
            lbl_titulo.config(fg=ACCENT)
            lbl_num.config(fg=ACCENT)
            card.config(bg=bg_card_hover)
            linha.config(bg=bg_card_hover)

        def sair(e=None):
            ativo = (self._active_card_key == chave)
            self._aplicar_estado_card(card, lbl_titulo, lbl_num, ativo)
            bg_normal = bg_card_hover if ativo else BG_CARD
            card.config(bg=bg_normal)
            linha.config(bg=bg_normal)

        for widget in (card, linha, lbl_num, lbl_titulo):
            widget.bind("<Enter>", entrar)
            widget.bind("<Leave>", sair)
            widget.bind("<Button-1>", lambda e: comando())

        self._aplicar_estado_card(card, lbl_titulo, lbl_num, self._active_card_key == chave)
        return card

    def campo_entry(self, parent, mostrar=None):
        entry = tk.Entry(
            parent, bg=ENTRY_BG, fg=TEXTO, insertbackground=ACCENT,
            relief="flat", font=("Arial", 11), show=mostrar,
            highlightthickness=0, bd=0
        )
        return entry

    def build_menu_automacao(self, parent, titulo, itens):
        self.botao_voltar(parent)

        center_box = tk.Frame(parent, bg=BG)
        center_box.pack(anchor="n", expand=True, pady=16, padx=28)

        # Cabeçalho Minimalista
        painel = tk.Frame(center_box, bg=BG_CARD, highlightbackground=BORDA, highlightthickness=1)
        painel.pack(fill="x", pady=(0, 16))

        inner_p = tk.Frame(painel, bg=BG_CARD, padx=20, pady=16)
        inner_p.pack(fill="x")

        tk.Label(
            inner_p, text="AUTOMAÇÃO", bg=BG_CARD, fg=ACCENT, font=("Arial", 9, "bold")
        ).pack(anchor="w", pady=(0, 2))

        tk.Label(
            inner_p, text=titulo.upper(), bg=BG_CARD, fg=TEXTO, font=("Arial", 18, "bold")
        ).pack(anchor="w")

        # Opções limpas em formato card
        for index, item in enumerate(itens, start=1):
            if isinstance(item, tuple) and len(item) == 3:
                label, comando, descricao = item
            else:
                label, comando = item
                descricao = None

            card = tk.Frame(
                center_box, bg=BG_CARD, cursor="hand2",
                highlightbackground=BORDA, highlightthickness=1
            )
            card.pack(fill="x", pady=5)

            def _on_enter(e, t=card):
                t.configure(highlightbackground=ACCENT)

            def _on_leave(e, t=card):
                t.configure(highlightbackground=BORDA)

            card.bind("<Enter>", _on_enter)
            card.bind("<Leave>", _on_leave)

            inner = tk.Frame(card, bg=BG_CARD, padx=16, pady=12)
            inner.pack(fill="both", expand=True)

            lbl_num = tk.Label(
                inner, text=f".{index:02d}", bg=BG_CARD, fg=ACCENT,
                font=("Arial", 11, "bold")
            )
            lbl_num.pack(side="left", padx=(0, 12))

            text_area = tk.Frame(inner, bg=BG_CARD)
            text_area.pack(side="left", fill="both", expand=True)

            lbl_title = tk.Label(
                text_area, text=label.upper(), bg=BG_CARD, fg=TEXTO,
                font=("Arial", 10, "bold"), anchor="w"
            )
            lbl_title.pack(fill="x")

            lbl_desc = None
            if descricao:
                lbl_desc = tk.Label(
                    text_area, text=descricao, bg=BG_CARD, fg=TEXTO_MUTED,
                    font=("Arial", 8), anchor="w"
                )
                lbl_desc.pack(fill="x", pady=(1, 0))

            cmd = comando
            for w in [card, inner, text_area, lbl_num, lbl_title]:
                w.bind("<Button-1>", lambda e, c=cmd: c())
                w.bind("<Enter>", _on_enter)
                w.bind("<Leave>", _on_leave)
            if lbl_desc:
                lbl_desc.bind("<Button-1>", lambda e, c=cmd: c())
                lbl_desc.bind("<Enter>", _on_enter)
                lbl_desc.bind("<Leave>", _on_leave)

    # =========================================================================
    # LOGIN (Refatorado: Nome Completo + Código Engenharia, sem Senha)
    # =========================================================================

    def _build_login(self):
        if FRAME_LOGIN in self.frames:
            self.frames[FRAME_LOGIN].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_LOGIN] = frame

        # Wrapper centralizador que garante alinhamento horizontal e vertical
        wrapper = tk.Frame(frame, bg=BG)
        wrapper.pack(expand=True, fill="both", pady=40)

        painel = tk.Frame(
            wrapper, bg=BG_CARD,
            bd=0, highlightthickness=0
        )
        painel.pack(anchor="center", expand=True)

        inner = tk.Frame(painel, bg=BG_CARD, padx=44, pady=36)
        inner.pack()

        tk.Label(
            inner, text="FLEX • CLASSIFICAÇÃO FISCAL", bg=BG_CARD, fg=ACCENT, font=("Arial", 9, "bold")
        ).pack(pady=(0, 6))

        tk.Label(
            inner, text="LOGIN CORPORATIVO", bg=BG_CARD, fg=TEXTO, font=FONT_TITULO
        ).pack(pady=(0, 6))

        tk.Label(
            inner, text="Informe seu Nome Completo e Código de Engenharia para acessar.", bg=BG_CARD, fg=TEXTO_MUTED, font=FONT_SUBTITULO
        ).pack(pady=(0, 20))

        bloco = tk.Frame(inner, bg=BG_CARD)
        bloco.pack(fill="x")

        tk.Label(
            bloco, text=espacar("Nome Completo"), bg=BG_CARD, fg=TEXTO_MUTED, font=FONT_CAPTION
        ).pack(anchor="w", pady=(12, 6))

        self.campo_nome = self.campo_entry(bloco)
        self.campo_nome.pack(fill="x", ipady=8)

        tk.Label(
            bloco, text=espacar("Código Engenharia"), bg=BG_CARD, fg=TEXTO_MUTED, font=FONT_CAPTION
        ).pack(anchor="w", pady=(20, 6))

        self.campo_codigo = self.campo_entry(bloco, mostrar="*")
        self.campo_codigo.pack(fill="x", ipady=8)

        self.campo_nome.bind("<Return>", lambda e: self.entrar())
        self.campo_codigo.bind("<Return>", lambda e: self.entrar())

        # Botão Entrar Proeminente e Intuitivo
        btn_entrar = tk.Button(
            bloco,
            text="ENTRAR",
            command=self.entrar,
            bg=ACCENT,
            fg="#ffffff",
            activebackground=ACCENT_HOVER,
            activeforeground="#ffffff",
            bd=0,
            relief="flat",
            font=("Arial", 10, "bold"),
            cursor="hand2",
            pady=10
        )
        btn_entrar.pack(fill="x", pady=(24, 14))

        def _hover_entrar(e):
            btn_entrar.config(bg=ACCENT_HOVER)
        def _leave_entrar(e):
            btn_entrar.config(bg=ACCENT)

        btn_entrar.bind("<Enter>", _hover_entrar)
        btn_entrar.bind("<Leave>", _leave_entrar)

        btn_help = tk.Button(
            bloco,
            text="❓ Não tem cadastro? Clique aqui para saber como solicitar",
            command=self._dialog_ajuda_cadastro,
            bg=BG_CARD,
            fg=ACCENT,
            activebackground=BG_CARD,
            activeforeground=ACCENT,
            bd=0,
            font=("Arial", 9, "underline"),
            cursor="hand2"
        )
        btn_help.pack(anchor="center", pady=(0, 16))

    def _dialog_ajuda_cadastro(self):
        janela = tk.Toplevel(self.root)
        janela.title("Como Solicitar Cadastro")
        janela.configure(bg=BG)
        janela.attributes("-topmost", True)
        janela.resizable(False, False)
        janela.transient(self.root)

        corpo = tk.Frame(janela, bg=BG, padx=24, pady=20)
        corpo.pack(fill="both", expand=True)

        tk.Label(
            corpo, text="🔐 Solicitação de Cadastro e Autorização", bg=BG, fg=ACCENT, font=("Arial", 12, "bold")
        ).pack(anchor="w", pady=(0, 10))

        msg = (
            "Para solicitar acesso ao sistema Flex-tax classification:\n\n"
            "1. Envie um e-mail explicando o motivo da sua solicitação.\n"
            "2. Destinatário: gabriell.girotto@flex.com\n"
            "3. Informe no e-mail o seu Nome Completo.\n\n"
            "Após a análise e aprovação, seu cadastro será liberado pelo Administrador."
        )

        tk.Label(
            corpo, text=msg, bg=BG, fg=TEXTO, font=("Arial", 9), justify="left"
        ).pack(anchor="w", pady=(0, 16))

        self.botao_flat(corpo, "Entendido", janela.destroy).pack(fill="x")

        janela.update_idletasks()
        largura, altura = janela.winfo_width(), janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (largura // 2)
        y = (janela.winfo_screenheight() // 2) - (altura // 2)
        janela.geometry(f"+{x}+{y}")
        janela.grab_set()

    def _dialog_trocar_pin_primeiro_acesso(self, nome: str, user_id: int) -> bool:
        """Exibe um diálogo modal obrigatório para o usuário cadastrar seu próprio PIN pessoal no 1º login."""
        resultado = [False]

        janela = tk.Toplevel(self.root)
        janela.title("🔐 Definição Obrigatória de PIN Personalizado")
        janela.configure(bg=BG)
        janela.attributes("-topmost", True)
        janela.resizable(False, False)
        janela.transient(self.root)

        janela.update_idletasks()
        w, h = 480, 430
        x = (janela.winfo_screenwidth() // 2) - (w // 2)
        y = (janela.winfo_screenheight() // 2) - (h // 2)
        janela.geometry(f"{w}x{h}+{x}+{y}")

        hdr = tk.Frame(janela, bg=ACCENT, height=54)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        tk.Label(
            hdr, text="🔐  CRIE SEU PIN DE SEGURANÇA",
            bg=ACCENT, fg="#ffffff", font=("Arial", 12, "bold")
        ).pack(side="left", padx=16, pady=14)

        corpo = tk.Frame(janela, bg=BG, padx=24, pady=16)
        corpo.pack(fill="both", expand=True)

        tk.Label(
            corpo,
            text=f"Olá, {nome}!\nEsta é a sua primeira vez acessando (ou o Administrador forneceu um código temporário para você).\nPor questões de segurança, você deve definir seu novo PIN pessoal de 4+ dígitos agora.",
            bg=BG, fg=TEXTO, font=("Arial", 9), justify="left", wraplength=430
        ).pack(anchor="w", pady=(0, 10))

        alerta_frame = tk.Frame(corpo, bg="#3a2a0a" if tema_atual() == "dark" else "#fff3e0", bd=0, highlightthickness=1, highlightbackground=ACCENT)
        alerta_frame.pack(fill="x", pady=(0, 14))

        tk.Label(
            alerta_frame,
            text="⚠️ ATENÇÃO E AVISO IMPORTANTE:\nGuarde e salve seu novo PIN em local seguro!\nSe você esquecer o seu PIN, não será possível recuperar e você precisará solicitar um novo código ao Administrador.",
            bg=alerta_frame.cget("bg"), fg=ACCENT, font=("Arial", 8, "bold"), justify="left", wraplength=410
        ).pack(padx=10, pady=8)

        tk.Label(corpo, text="Novo PIN / Código (mínimo 4 dígitos):", bg=BG, fg=TEXTO_MUTED, font=("Arial", 8, "bold")).pack(anchor="w", pady=(4, 2))
        entry_pin1 = self.campo_entry(corpo, mostrar="*")
        entry_pin1.pack(fill="x", ipady=6, pady=(0, 10))

        tk.Label(corpo, text="Confirmar Novo PIN / Código:", bg=BG, fg=TEXTO_MUTED, font=("Arial", 8, "bold")).pack(anchor="w", pady=(0, 2))
        entry_pin2 = self.campo_entry(corpo, mostrar="*")
        entry_pin2.pack(fill="x", ipady=6, pady=(0, 14))

        entry_pin1.focus_set()

        def _confirmar_troca():
            pin1 = entry_pin1.get().strip()
            pin2 = entry_pin2.get().strip()

            if not pin1 or len(pin1) < 4:
                messagebox.showwarning("PIN Inválido", "O novo PIN deve conter no mínimo 4 caracteres.", parent=janela)
                entry_pin1.focus_set()
                return

            if pin1 != pin2:
                messagebox.showwarning("Confirmação Incorreta", "A confirmação do PIN não coincide com o novo PIN digitado.", parent=janela)
                entry_pin2.focus_set()
                return

            from services.database import alterar_codigo_usuario
            ok = alterar_codigo_usuario(user_id, pin1)
            if ok:
                messagebox.showinfo(
                    "PIN Salvo com Sucesso",
                    "✅ Seu novo PIN foi definido e salvo com sucesso!\n\n"
                    "Lembre-se de guardá-lo em local seguro. Use este novo PIN em seus próximos logins.",
                    parent=janela
                )
                resultado[0] = True
                janela.destroy()
            else:
                messagebox.showerror("Erro", "Não foi possível atualizar seu PIN no banco de dados.", parent=janela)

        def _cancelar():
            resultado[0] = False
            janela.destroy()

        btn_box = tk.Frame(corpo, bg=BG)
        btn_box.pack(fill="x", pady=(6, 0))

        btn_salvar = tk.Button(
            btn_box, text="SALVAR NOVO PIN", command=_confirmar_troca,
            bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HOVER, activeforeground="#ffffff",
            bd=0, relief="flat", font=("Arial", 9, "bold"), cursor="hand2", pady=8
        )
        btn_salvar.pack(side="left", fill="x", expand=True, padx=(0, 4))

        btn_canc = tk.Button(
            btn_box, text="Cancelar", command=_cancelar,
            bg=BG_CARD, fg=TEXTO, activebackground=BORDA, activeforeground=TEXTO,
            relief="solid", bd=1, font=("Arial", 9), cursor="hand2", pady=7
        )
        btn_canc.pack(side="right", padx=(4, 0))

        entry_pin1.bind("<Return>", lambda e: entry_pin2.focus_set())
        entry_pin2.bind("<Return>", lambda e: _confirmar_troca())

        janela.grab_set()
        self.root.wait_window(janela)
        return resultado[0]

    def entrar(self, automatic=False):
        nome = self.campo_nome.get().strip()
        codigo = self.campo_codigo.get().strip()

        if not nome:
            messagebox.showwarning("Aviso", "Digite o Nome Completo.")
            return
        if not codigo:
            messagebox.showwarning("Aviso", "Digite o Código Engenharia.")
            return

        sucesso, mensagem_motivo, resultado = validar_usuario_detalhado(nome, codigo)

        if not sucesso or not resultado:
            db_atual = get_db_path()
            # Segurança: registra a tentativa de login malsucedida para auditoria
            # (nunca o código/senha digitado, apenas o nome informado).
            try:
                registrar_tentativa_bloqueada("LOGIN_FALHOU", f"Nome informado: {nome}", usuario=nome)
            except Exception:
                pass
            messagebox.showerror(
                "Acesso Negado",
                f"{mensagem_motivo}\n\n"
                f"• Banco de Dados em Uso:\n{db_atual}\n\n"
                f"💡 Verifique se o cadastro foi salvo exatamente neste arquivo de banco de dados ou se a conta foi importada."
            )
            return

        permissao = resultado[0]
        user_id = resultado[1]
        requer_troca = resultado[2] if len(resultado) > 2 else False

        if requer_troca:
            trocou_ok = self._dialog_trocar_pin_primeiro_acesso(nome, user_id)
            if not trocou_ok:
                return

        self.usuario = nome
        self.usuario_id = user_id
        self.permissao = permissao
        self._cred_pin = None

        logger.info("Login efetuado (usuário=%s, id=%s, permissão=%s)", nome, user_id, permissao)
        registrar_log_auditoria("LOGIN", nome, f"ID: {user_id} | Perfil: {permissao}")

        if hasattr(self, "user_badge"):
            self.user_badge.config(text=f"USUÁRIO: {nome.upper()} ({permissao})")

        self._recriar_itens_nav()
        self._build_menu()
        self.history = []
        self.mostrar(FRAME_MENU)
        self._atualizar_badge_notificacoes()

    # =========================================================================
    # MENU PRINCIPAL (Específico por Perfil)
    # =========================================================================

    def _build_menu(self):
        if FRAME_MENU in self.frames:
            self.frames[FRAME_MENU].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_MENU] = frame

        # Container centralizado com largura estipulada (640px)
        center_box = tk.Frame(frame, bg=BG)
        center_box.pack(anchor="n", expand=True, pady=20, padx=28)

        # Header do Painel Principal
        painel = tk.Frame(center_box, bg=BG_CARD, highlightbackground=BORDA, highlightthickness=1)
        painel.pack(fill="x", pady=(0, 16))

        inner_p = tk.Frame(painel, bg=BG_CARD, padx=24, pady=18)
        inner_p.pack(fill="x")

        topo_p = tk.Frame(inner_p, bg=BG_CARD)
        topo_p.pack(fill="x")

        tk.Label(
            topo_p, text="FLEX • CLASSIFICAÇÃO FISCAL", bg=BG_CARD, fg=ACCENT, font=("Arial", 8, "bold")
        ).pack(side="left")

        tk.Label(
            topo_p, text=datetime.now().strftime("%d/%m/%Y  •  %H:%M"), bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 8)
        ).pack(side="right")

        tk.Label(
            inner_p, text=f"Bem-vindo, {self.usuario}", bg=BG_CARD, fg=TEXTO, font=("Arial", 16, "bold")
        ).pack(anchor="w", pady=(6, 2))

        tk.Label(
            inner_p, text=f"Perfil de acesso: {self.permissao}", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 9)
        ).pack(anchor="w")

        tk.Label(
            center_box, text="PAINEL DE NAVEGAÇÃO", bg=BG, fg=TEXTO_MUTED, font=("Arial", 8, "bold")
        ).pack(anchor="w", pady=(6, 10))

        perm = str(self.permissao).upper()

        menu_opcoes = [
            ("Dashboard", self._abrir_dashboard, "Visão geral operacional, estatísticas e exportações"),
            ("Phoenix", self._abrir_menu_phoenix, "Fluxo principal de solicitações P0032 e captura"),
            ("Pegasus", self._abrir_menu_pegasus, "Fluxo Pegasus e acompanhamento de etapas"),
            ("Cost Request", self._executar_cost_request, "Automação para solicitação de análise fiscal"),
            ("Cost by Template", self._abrir_cost_template, "Abre solicitação de Cost Request a partir de um template Excel"),
            ("Minhas Credenciais", self._abrir_minhas_credenciais, "Gerenciar logins e senhas com PIN seguro"),
            ("Administração", self._abrir_administracao, "Gerenciar usuários, permissões e backups"),
        ]

        idx = 1
        for label, comando, desc in menu_opcoes:
            # Filtrar por permissão
            if label in ("Phoenix", "Pegasus", "Cost Request", "Cost by Template", "Minhas Credenciais") and perm not in ("ADMIN", "ENGENHARIA"):
                continue
            if label == "Administração" and perm != "ADMIN":
                continue

            card = tk.Frame(
                center_box, bg=BG_CARD, cursor="hand2",
                highlightbackground=BORDA, highlightthickness=1
            )
            card.pack(fill="x", pady=5)

            def _on_enter(e, t=card):
                t.configure(highlightbackground=ACCENT)

            def _on_leave(e, t=card):
                t.configure(highlightbackground=BORDA)

            card.bind("<Enter>", _on_enter)
            card.bind("<Leave>", _on_leave)

            inner = tk.Frame(card, bg=BG_CARD, padx=16, pady=12)
            inner.pack(fill="both", expand=True)

            lbl_num = tk.Label(
                inner, text=f".{idx:02d}", bg=BG_CARD, fg=ACCENT,
                font=("Arial", 11, "bold")
            )
            lbl_num.pack(side="left", padx=(0, 12))

            text_area = tk.Frame(inner, bg=BG_CARD)
            text_area.pack(side="left", fill="both", expand=True)

            lbl_title = tk.Label(
                text_area, text=label.upper(), bg=BG_CARD, fg=TEXTO,
                font=("Arial", 10, "bold"), anchor="w"
            )
            lbl_title.pack(fill="x")

            lbl_desc = tk.Label(
                text_area, text=desc, bg=BG_CARD, fg=TEXTO_MUTED,
                font=("Arial", 8), anchor="w"
            )
            lbl_desc.pack(fill="x", pady=(1, 0))

            cmd = comando
            for w in [card, inner, text_area, lbl_num, lbl_title, lbl_desc]:
                w.bind("<Button-1>", lambda e, c=cmd: c())
                w.bind("<Enter>", _on_enter)
                w.bind("<Leave>", _on_leave)

            idx += 1

        tk.Label(
            frame, text=f"Flex-tax Classification v1.0 • Perfil: {perm}", bg=BG, fg=TEXTO_MUTED,
            font=FONT_CAPTION
        ).pack(side="bottom", pady=14)

    # =========================================================================
    # GERENCIAMENTO DE CREDENCIAIS INDIVIDUAIS PARA AUTOMAÇÕES
    # =========================================================================

    def _pedir_pin(self, titulo: str, mensagem: str, confirmar: bool = False) -> Optional[str]:
        """Modal para o usuário digitar (e opcionalmente confirmar) seu PIN pessoal de credenciais."""
        resultado = [None]
        janela = tk.Toplevel(self.root)
        janela.title(titulo)
        janela.configure(bg=BG)
        janela.attributes("-topmost", True)
        janela.resizable(False, False)
        janela.transient(self.root)

        corpo = tk.Frame(janela, bg=BG, padx=24, pady=20)
        corpo.pack(fill="both", expand=True)

        tk.Label(corpo, text=titulo, bg=BG, fg=TEXTO, font=("Arial", 13, "bold")).pack(anchor="w", pady=(0, 6))
        tk.Label(
            corpo, text=mensagem, bg=BG, fg=TEXTO_MUTED, font=("Arial", 9), justify="left", wraplength=340
        ).pack(anchor="w", pady=(0, 14))

        tk.Label(corpo, text="PIN (mínimo 4 caracteres):", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        campo_pin = self.campo_entry(corpo, mostrar="*")
        campo_pin.pack(fill="x", ipady=6, pady=(0, 10))

        campo_confirmar = None
        if confirmar:
            tk.Label(corpo, text="Confirme o PIN:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
            campo_confirmar = self.campo_entry(corpo, mostrar="*")
            campo_confirmar.pack(fill="x", ipady=6, pady=(0, 16))

        def _confirmar_click():
            pin = campo_pin.get().strip()
            if len(pin) < 4:
                messagebox.showwarning("Aviso", "O PIN deve ter no mínimo 4 caracteres.", parent=janela)
                return
            if confirmar and campo_confirmar.get().strip() != pin:
                messagebox.showwarning("Aviso", "Os PINs informados não coincidem.", parent=janela)
                return
            resultado[0] = pin
            janela.destroy()

        campo_pin.bind("<Return>", lambda e: _confirmar_click())
        self.botao_flat(corpo, "Confirmar", _confirmar_click).pack(fill="x")

        janela.update_idletasks()
        largura, altura = janela.winfo_width(), janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (largura // 2)
        y = (janela.winfo_screenheight() // 2) - (altura // 2)
        janela.geometry(f"+{x}+{y}")
        campo_pin.focus_set()
        janela.grab_set()
        self.root.wait_window(janela)

        return resultado[0]

    def _obter_pin_sessao(self) -> Optional[str]:
        """Garante que o usuário tenha um PIN pessoal de credenciais e o retorna para a sessão atual.

        O PIN nunca é gravado em disco — fica apenas em memória enquanto o app estiver aberto.
        """
        if getattr(self, "_cred_pin", None):
            return self._cred_pin

        if not usuario_possui_pin_credenciais(self.usuario_id):
            messagebox.showinfo(
                "Configurar PIN de Credenciais",
                "Antes de cadastrar suas credenciais de automação, crie um PIN pessoal.\n\n"
                "Esse PIN é conhecido apenas por você — nem administradores conseguem acessar "
                "suas credenciais sem ele. Se você esquecer o PIN, as credenciais salvas não "
                "poderão ser recuperadas e precisarão ser cadastradas novamente."
            )
            pin = self._pedir_pin(
                "Criar PIN de Credenciais",
                "Escolha um PIN pessoal para proteger suas credenciais de automação.",
                confirmar=True
            )
            if not pin:
                return None
            if not definir_pin_credenciais(self.usuario_id, pin):
                messagebox.showerror("Erro", "Não foi possível definir o PIN.")
                return None
            self._cred_pin = pin
            return pin

        pin = self._pedir_pin(
            "Digite seu PIN de Credenciais",
            "Informe o PIN pessoal para acessar suas credenciais de automação."
        )
        if pin:
            self._cred_pin = pin
        return pin

    def _obter_ou_pedir_credencial(self, tipo: str, forcar: bool = False) -> Optional[Dict[str, str]]:
        """Busca credencial no SQLite (protegida pelo PIN pessoal); se não existir ou
        se forcar=True, abre modal para solicitar/atualizar."""
        pin = self._obter_pin_sessao()
        if not pin:
            return None

        cred = obter_credencial(self.usuario_id, tipo, pin)
        if cred is None and credencial_configurada(self.usuario_id, tipo):
            self._cred_pin = None
            messagebox.showerror(
                "PIN incorreto",
                "Não foi possível acessar suas credenciais com o PIN informado."
            )
            return None
        if cred and cred.get("login") and cred.get("senha") and not forcar:
            return cred

        # Não encontrou credencial salva ou solicitou alteração -> Abrir modal
        resultado = [None]

        janela = tk.Toplevel(self.root)
        janela.title(f"Credenciais {tipo}")
        janela.configure(bg=BG)
        janela.attributes("-topmost", True)
        janela.resizable(False, False)
        janela.transient(self.root)

        corpo = tk.Frame(janela, bg=BG, padx=24, pady=20)
        corpo.pack(fill="both", expand=True)

        tk.Label(
            corpo, text=f"Credenciais {tipo}", bg=BG, fg=TEXTO, font=("Arial", 14, "bold")
        ).pack(anchor="w", pady=(0, 6))

        tk.Label(
            corpo,
            text=f"Informe suas credenciais individuais para a automação {tipo}.\nSerão salvas de forma segura no SQLite.",
            bg=BG, fg=TEXTO_MUTED, font=("Arial", 9), justify="left"
        ).pack(anchor="w", pady=(0, 16))

        tk.Label(corpo, text="JAG / Usuário:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        campo_jag = self.campo_entry(corpo)
        campo_jag.pack(fill="x", ipady=6, pady=(0, 10))

        tk.Label(corpo, text="Senha:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        campo_senha = self.campo_entry(corpo, mostrar="*")
        campo_senha.pack(fill="x", ipady=6, pady=(0, 16))

        # Preencher campos com valores existentes se disponíveis
        if cred:
            if cred.get("login"):
                campo_jag.insert(0, cred.get("login"))
            if cred.get("senha"):
                campo_senha.insert(0, cred.get("senha"))

        def salvar():
            jag = campo_jag.get().strip()
            senha = campo_senha.get().strip()
            if not jag or not senha:
                messagebox.showwarning("Aviso", "Preencha o JAG e a Senha.", parent=janela)
                return
            salvar_credencial(self.usuario_id, tipo, jag, senha, pin)
            resultado[0] = {"login": jag, "senha": senha}
            janela.destroy()

        self.botao_flat(corpo, "Salvar", salvar).pack(fill="x")

        janela.update_idletasks()
        largura, altura = janela.winfo_width(), janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (largura // 2)
        y = (janela.winfo_screenheight() // 2) - (altura // 2)
        janela.geometry(f"+{x}+{y}")
        campo_jag.focus_set()
        janela.grab_set()
        self.root.wait_window(janela)

        return resultado[0]

    def _executar_cost_request(self):
        if not self._validar_permissao("cost"):
            return
        cred = self._obter_ou_pedir_credencial("COST")
        if not cred:
            return
        executar_script("automocoes", "cost", "cost.py", cred_user=cred["login"], cred_pass=cred["senha"], tool_user=self.usuario)

    def _abrir_cost_template(self):
        if not self._validar_permissao("cost"):
            return
        self._abrir_modal_cost_template()

    def _carregar_historico(self):
        return carregar_historico()

    def _abrir_dashboard(self):
        self._build_dashboard()
        self.mostrar(FRAME_DASHBOARD)

    # =========================================================================
    # DASHBOARD & EXPORTAÇÕES EXCEL
    # =========================================================================

    def _build_dashboard(self):
        if FRAME_DASHBOARD in self.frames:
            self.frames[FRAME_DASHBOARD].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_DASHBOARD] = frame

        self.botao_voltar(frame)
        self.cabecalho(frame, "Visão geral", "Dashboard")

        painel = tk.Frame(frame, bg=BG_CARD, highlightbackground=BORDA, highlightthickness=0)
        painel.pack(fill="x", padx=28, pady=(0, 10))

        tk.Label(
            painel,
            text="Resumo operacional",
            bg=BG_CARD,
            fg=TEXTO,
            font=("Arial", 12, "bold"),
        ).pack(anchor="w", padx=16, pady=(16, 6))

        tk.Label(
            painel,
            text="● atualização em tempo real",
            bg=BG_CARD,
            fg="#67d28d",
            font=("Arial", 8, "bold"),
        ).pack(anchor="w", padx=16, pady=(0, 8))

        tk.Label(
            painel,
            text="Acompanhe solicitações, pendências e evolução do histórico em tempo real.",
            bg=BG_CARD,
            fg=TEXTO_MUTED,
            font=FONT_SUBTITULO,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(0, 16))

        topo_barra = tk.Frame(frame, bg=BG)
        topo_barra.pack(fill="x", padx=28, pady=(0, 10))

        # Barra de Pesquisa fixa à DIREITA (nunca foge da tela)
        busca_frame = tk.Frame(topo_barra, bg=BG)
        busca_frame.pack(side="right", padx=(10, 0))

        tk.Label(
            busca_frame, text="Buscar:", bg=BG, fg=TEXTO_MUTED, font=("Arial", 8, "bold")
        ).pack(side="left", padx=(0, 4))

        self.campo_busca = tk.Entry(
            busca_frame,
            bg=BG_CARD,
            fg=TEXTO,
            insertbackground=TEXTO,
            relief="solid",
            font=("Arial", 8),
            highlightbackground=BORDA,
            highlightthickness=1,
            bd=0,
            width=18,
        )
        self.campo_busca.pack(side="left", ipady=3)
        if hasattr(self, "_dashboard_search_query") and self._dashboard_search_query:
            self.campo_busca.insert(0, self._dashboard_search_query)
        self.campo_busca.bind("<KeyRelease>", lambda e: self._filtrar_busca_dashboard())

        # Botões de Ação uniformes e elegantes à ESQUERDA
        btn_base = dict(
            relief="flat", bd=0, font=("Arial", 8, "bold"),
            pady=4, cursor="hand2"
        )

        tk.Button(topo_barra, text="Atualizar", command=self._abrir_dashboard, bg=BG_CARD, fg=TEXTO, activebackground=BORDA, activeforeground=TEXTO, width=12, **btn_base).pack(side="left", padx=2)
        tk.Button(topo_barra, text="Capturar Phoenix", command=lambda: self._executar_phoenix(arg="capturar"), bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HOVER, activeforeground="#ffffff", width=14, **btn_base).pack(side="left", padx=2)
        tk.Button(topo_barra, text="Abrir Cost Request", command=self._executar_cost_manual, bg="#e67e22", fg="#ffffff", activebackground="#d35400", activeforeground="#ffffff", width=16, **btn_base).pack(side="left", padx=2)
        tk.Button(topo_barra, text="Open Request By Template", command=self._abrir_modal_cost_template, bg="#8e44ad", fg="#ffffff", activebackground="#732d91", activeforeground="#ffffff", width=22, **btn_base).pack(side="left", padx=2)
        tk.Button(topo_barra, text="Importar Planilha", command=self._importar_planilha_excel, bg="#27ae60", fg="#ffffff", activebackground="#219653", activeforeground="#ffffff", width=14, **btn_base).pack(side="left", padx=2)
        tk.Button(topo_barra, text="Exp. Dashboard", command=self._exportar_dashboard, bg=BG_CARD, fg=TEXTO, activebackground=BORDA, activeforeground=TEXTO, width=13, **btn_base).pack(side="left", padx=2)
        tk.Button(topo_barra, text="Exp. Registros", command=self._exportar_registros, bg=BG_CARD, fg=TEXTO, activebackground=BORDA, activeforeground=TEXTO, width=13, **btn_base).pack(side="left", padx=2)
        tk.Button(topo_barra, text="Esvaziar Lixeira", command=self._esvaziar_lixeira_confirm, bg=BG_CARD, fg=TEXTO, activebackground=BORDA, activeforeground=TEXTO, width=13, **btn_base).pack(side="left", padx=2)

        # Separador visual discreto
        tk.Frame(topo_barra, bg=BORDA, width=1, height=20).pack(side="left", padx=6)

        # Filtros de Status uniformes
        for texto, valor in (("Todos", "todos"), ("Em andamento", "on going"), ("Finalizadas", "finalizadas"), ("Lixeira", "lixeira")):
            is_active = getattr(self, "_dashboard_filter", "todos") == valor
            bg_f = ACCENT if is_active else BG_CARD
            fg_f = "#ffffff" if is_active else TEXTO
            btn = tk.Button(
                topo_barra,
                text=texto,
                command=lambda v=valor: self._aplicar_filtro_dashboard(v),
                bg=bg_f, fg=fg_f, activebackground=ACCENT_HOVER, activeforeground="#ffffff",
                width=12, **btn_base
            )
            btn.pack(side="left", padx=2)

        # Linha de filtro por intervalo de datas
        data_barra = tk.Frame(frame, bg=BG)
        data_barra.pack(fill="x", padx=28, pady=(0, 8))

        entry_kw = dict(
            bg=BG_CARD, fg=TEXTO, insertbackground=TEXTO,
            relief="solid", font=("Arial", 8),
            highlightbackground=BORDA, highlightthickness=1, bd=0, width=10
        )

        tk.Label(data_barra, text="📅 Filtrar por Data:", bg=BG, fg=TEXTO_MUTED, font=("Arial", 8, "bold")).pack(side="left", padx=(0, 6))

        tk.Label(data_barra, text="De:", bg=BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="left", padx=(0, 2))
        self._date_from = tk.Entry(data_barra, **entry_kw)
        self._date_from.pack(side="left", ipady=3, padx=(0, 6))
        if hasattr(self, "_dashboard_date_from") and self._dashboard_date_from:
            self._date_from.insert(0, self._dashboard_date_from)
        else:
            self._date_from.insert(0, "dd/mm/aaaa")
            self._date_from.configure(fg=TEXTO_MUTED)

        def _focus_in_from(e):
            if self._date_from.get() == "dd/mm/aaaa":
                self._date_from.delete(0, "end")
                self._date_from.configure(fg=TEXTO)

        def _focus_out_from(e):
            if not self._date_from.get().strip():
                self._date_from.insert(0, "dd/mm/aaaa")
                self._date_from.configure(fg=TEXTO_MUTED)

        self._date_from.bind("<FocusIn>", _focus_in_from)
        self._date_from.bind("<FocusOut>", _focus_out_from)

        tk.Label(data_barra, text="Até:", bg=BG, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="left", padx=(0, 2))
        self._date_to = tk.Entry(data_barra, **entry_kw)
        self._date_to.pack(side="left", ipady=3, padx=(0, 6))
        if hasattr(self, "_dashboard_date_to") and self._dashboard_date_to:
            self._date_to.insert(0, self._dashboard_date_to)
        else:
            self._date_to.insert(0, "dd/mm/aaaa")
            self._date_to.configure(fg=TEXTO_MUTED)

        def _focus_in_to(e):
            if self._date_to.get() == "dd/mm/aaaa":
                self._date_to.delete(0, "end")
                self._date_to.configure(fg=TEXTO)

        def _focus_out_to(e):
            if not self._date_to.get().strip():
                self._date_to.insert(0, "dd/mm/aaaa")
                self._date_to.configure(fg=TEXTO_MUTED)

        self._date_to.bind("<FocusIn>", _focus_in_to)
        self._date_to.bind("<FocusOut>", _focus_out_to)

        def _aplicar_filtro_data():
            de = self._date_from.get().strip()
            ate = self._date_to.get().strip()
            if de == "dd/mm/aaaa":
                de = ""
            if ate == "dd/mm/aaaa":
                ate = ""
            self._dashboard_date_from = de
            self._dashboard_date_to = ate
            if FRAME_DASHBOARD in self.frames:
                historico_r = self._carregar_historico()
                self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], historico_r)

        def _limpar_filtro_data():
            self._dashboard_date_from = ""
            self._dashboard_date_to = ""
            self._date_from.delete(0, "end")
            self._date_from.insert(0, "dd/mm/aaaa")
            self._date_from.configure(fg=TEXTO_MUTED)
            self._date_to.delete(0, "end")
            self._date_to.insert(0, "dd/mm/aaaa")
            self._date_to.configure(fg=TEXTO_MUTED)
            if FRAME_DASHBOARD in self.frames:
                historico_r = self._carregar_historico()
                self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], historico_r)

        tk.Button(data_barra, text="Filtrar", command=_aplicar_filtro_data, bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HOVER, activeforeground="#ffffff", relief="flat", bd=0, font=("Arial", 8, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=2)
        tk.Button(data_barra, text="Limpar", command=_limpar_filtro_data, bg=BG_CARD, fg=TEXTO, activebackground=BORDA, activeforeground=TEXTO, **btn_base).pack(side="left", padx=2)

        # Indicador de filtro ativo
        date_from_val = getattr(self, "_dashboard_date_from", "")
        date_to_val = getattr(self, "_dashboard_date_to", "")
        if date_from_val or date_to_val:
            texto_filtro = f"Filtro ativo: {date_from_val or '...'} → {date_to_val or '...'}"
            tk.Label(data_barra, text=texto_filtro, bg=BG, fg=ACCENT, font=("Arial", 8, "bold")).pack(side="left", padx=(10, 0))

        historico = self._carregar_historico()
        self._dashboard_last_signature = self._dashboard_signature(historico)
        self._renderizar_dashboard_conteudo(frame, historico)
        self._cancelar_timer_dashboard()
        self._dashboard_after_id = self.root.after(4000, self._check_dashboard_refresh)

    def _cancelar_timer_dashboard(self):
        timer_id = getattr(self, "_dashboard_after_id", None)
        if timer_id:
            try:
                self.root.after_cancel(timer_id)
            except Exception:
                pass
            self._dashboard_after_id = None

    def _filtrar_busca_dashboard(self):
        search_timer = getattr(self, "_search_debounce_id", None)
        if search_timer:
            try:
                self.root.after_cancel(search_timer)
            except Exception:
                pass
            self._search_debounce_id = None

        def _executar_busca():
            self._search_debounce_id = None
            if hasattr(self, "campo_busca"):
                self._dashboard_search_query = self.campo_busca.get().strip().lower()
            if FRAME_DASHBOARD in self.frames:
                historico = self._carregar_historico()
                self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], historico)

        self._search_debounce_id = self.root.after(250, _executar_busca)

    def _aplicar_filtro_dashboard(self, valor):
        self._dashboard_filter = valor
        if FRAME_DASHBOARD in self.frames:
            historico = self._carregar_historico()
            self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], historico)

    def _esvaziar_lixeira_confirm(self):
        if not self._validar_permissao("admin"):
            return
        res = messagebox.askyesno(
            "Esvaziar Lixeira",
            "Tem certeza que deseja EXCLUIR PERMANENTEMENTE todos os registros da lixeira?\n\nEsta ação não poderá ser desfeita.",
            parent=self.root
        )
        if res:
            qtd = esvaziar_lixeira()
            messagebox.showinfo("Lixeira Esvaziada", f"{qtd} registro(s) cancelado(s) foi/foram excluído(s) permanentemente do banco.")
            self._abrir_dashboard()

    def _desenhar_grafico_donut_nativo(self, parent, v_fin, v_and, v_can, total_metrica):
        container = tk.Frame(parent, bg=BG)
        container._dashboard_area = True
        container.pack(pady=(10, 15))

        tk.Label(
            container,
            text="VISÃO GERAL DAS SOLICITAÇÕES",
            bg=BG,
            fg=ACCENT,
            font=("Arial", 10, "bold")
        ).pack(pady=(0, 10))

        valores_pie = [v_fin, v_and, v_can]
        cores_pie = ["#67d28d", "#ffbf69", "#ff6b6b"]
        labels_pie = ["Finalizadas", "Em andamento", "Canceladas"]
        total_val = sum(valores_pie)

        cv_dim = 180
        cv = tk.Canvas(container, width=cv_dim, height=cv_dim, bg=BG, highlightthickness=0, bd=0)
        cv.pack()

        cx, cy, r_out, r_in = cv_dim // 2, cv_dim // 2, 78, 48

        if total_val == 0:
            cv.create_oval(
                cx - r_out, cy - r_out, cx + r_out, cy + r_out,
                fill="#333b4d", outline=BG, width=2
            )
        else:
            angle_start = 90
            for val, color in zip(valores_pie, cores_pie):
                if val <= 0:
                    continue
                extent = -360.0 * (val / total_val)
                if abs(extent) >= 360:
                    extent = -359.99

                cv.create_arc(
                    cx - r_out, cy - r_out, cx + r_out, cy + r_out,
                    start=angle_start, extent=extent, fill=color, outline=BG, width=2, style="pieslice"
                )
                angle_start += extent

        # Recorte do miolo (donut hole)
        cv.create_oval(
            cx - r_in, cy - r_in, cx + r_in, cy + r_in,
            fill=BG, outline=BG
        )

        # Texto central
        cv.create_text(cx, cy - 8, text=str(total_metrica), fill=TEXTO, font=("Arial", 16, "bold"))
        cv.create_text(cx, cy + 12, text="TOTAL", fill=TEXTO_MUTED, font=("Arial", 8, "bold"))

        # Legenda infra com indicadores coloridos
        if total_val > 0:
            legenda_frame = tk.Frame(container, bg=BG)
            legenda_frame.pack(pady=(8, 0))
            for val, color, label in zip(valores_pie, cores_pie, labels_pie):
                if val > 0:
                    pct = int(round((val / total_val) * 100))
                    lbl_txt = f"● {label}: {val} ({pct}%)"
                    tk.Label(legenda_frame, text=lbl_txt, bg=BG, fg=color, font=("Arial", 8, "bold")).pack(side="left", padx=6)

    def _renderizar_dashboard_conteudo(self, frame, historico):
        for widget in frame.winfo_children():
            if getattr(widget, "_dashboard_area", False):
                widget.destroy()

        counts = contar_por_status()
        total_metrica = counts["total"]
        em_andamento_metrica = counts["em_andamento"]
        finalizadas_metrica = counts["finalizadas"]
        canceladas_metrica = counts["canceladas"]

        stats = tk.Frame(frame, bg=BG)
        stats._dashboard_area = True
        stats.pack(fill="x", padx=28, pady=(0, 8))

        self._stat_card(stats, total_metrica, "Total de solicitações")
        self._stat_card(stats, em_andamento_metrica, "Em andamento")
        self._stat_card(stats, finalizadas_metrica, "Finalizadas")
        self._stat_card(stats, canceladas_metrica, "Canceladas")

        # Gráfico Donut Nativo (Super leve, 0ms latency, sem Matplotlib)
        self._desenhar_grafico_donut_nativo(frame, finalizadas_metrica, em_andamento_metrica, canceladas_metrica, total_metrica)



        container = tk.Frame(frame, bg=BG)
        container._dashboard_area = True
        container.pack(fill="both", padx=28, pady=(0, 16))

        filtro = (self._dashboard_filter or "todos").strip().lower()
        if filtro == "on going":
            lista = [i for i in historico if str(i.get("status", "")).upper() == "ON GOING" and not i.get("cancelado")]
        elif filtro == "finalizadas":
            lista = [i for i in historico if str(i.get("status", "")).upper() not in ("ON GOING", "CANCELADO") and not i.get("cancelado")]
        elif filtro in ("lixeira", "canceladas"):
            lista = [i for i in historico if i.get("cancelado") or str(i.get("status", "")).upper() == "CANCELADO"]
        else:  # "todos" — exibe apenas solicitações ativas (não canceladas)
            lista = [i for i in historico if not i.get("cancelado") and str(i.get("status", "")).upper() != "CANCELADO"]

        query = getattr(self, "_dashboard_search_query", "").strip().lower()
        if query:
            lista = [
                item for item in lista
                if query in str(item.get("ticket", "")).lower()
                or query in str(item.get("pn", "")).lower()
                or query in str(item.get("part_number", "")).lower()
                or query in str(item.get("mpn", "")).lower()
                or query in str(item.get("description", "")).lower()
                or query in str(item.get("solicitante", "")).lower()
                or query in str(item.get("requisitante", "")).lower()
                or query in str(item.get("produto", "")).lower()
            ]

        # Filtro por intervalo de datas (data_abertura no formato dd/mm/yyyy)
        date_from_str = getattr(self, "_dashboard_date_from", "").strip()
        date_to_str = getattr(self, "_dashboard_date_to", "").strip()
        if date_from_str or date_to_str:
            def _parse_data(txt):
                """Tenta converter dd/mm/yyyy ou dd/mm/yy em date."""
                txt = txt.strip()
                if not txt:
                    return None
                for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
                    try:
                        return datetime.strptime(txt, fmt).date()
                    except ValueError:
                        continue
                return None

            dt_from = _parse_data(date_from_str) if date_from_str else None
            dt_to = _parse_data(date_to_str) if date_to_str else None

            if dt_from or dt_to:
                filtrada = []
                for item in lista:
                    data_ab = str(item.get("data_abertura") or "").strip()
                    # data_abertura pode ter formato "dd/mm/yyyy HH:MM" — pegar só a parte da data
                    data_ab_part = data_ab.split(" ")[0] if data_ab else ""
                    dt_item = _parse_data(data_ab_part)
                    if dt_item is None:
                        continue  # Se não tem data válida, não inclui no filtro
                    if dt_from and dt_item < dt_from:
                        continue
                    if dt_to and dt_item > dt_to:
                        continue
                    filtrada.append(item)
                lista = filtrada
        if not lista:
            msg = "Nenhum registro encontrado para a busca informada." if query else "Nenhum registro encontrado para este filtro."
            tk.Label(container, text=msg, bg=BG, fg=TEXTO_MUTED, font=FONT_SUBTITULO).pack(anchor="w", pady=10)
            return

        card_limit = getattr(self, "_dashboard_card_limit", 50)
        exibir = lista[:card_limit]

        for item in exibir:
            self._linha_historico(container, item)

        if len(lista) > card_limit:
            pagi_frame = tk.Frame(container, bg=BG)
            pagi_frame.pack(fill="x", pady=(12, 6))

            info_txt = f"Mostrando {len(exibir)} de {len(lista)} solicitações"
            tk.Label(pagi_frame, text=info_txt, bg=BG, fg=TEXTO_MUTED, font=("Arial", 9, "bold")).pack(side="left", padx=(4, 12))

            def _carregar_mais():
                self._dashboard_card_limit = card_limit + 50
                if FRAME_DASHBOARD in self.frames:
                    h_rec = self._carregar_historico()
                    self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], h_rec)

            btn_mais = tk.Button(
                pagi_frame, text="📥 Carregar Mais (+50 Registros)", command=_carregar_mais,
                bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HOVER, activeforeground="#ffffff",
                relief="flat", bd=0, font=("Arial", 8, "bold"), padx=12, pady=4, cursor="hand2"
            )
            btn_mais.pack(side="left")

    def _dashboard_signature(self, historico):
        return tuple(
            (
                str(item.get("id", "")),
                str(item.get("linha", "")),
                str(item.get("status", "")),
                str(item.get("description", "")),
                str(item.get("pn") or item.get("part_number") or ""),
                str(item.get("ultima_alteracao", "")),
            )
            for item in historico
        )

    def _check_dashboard_refresh(self):
        if self.history and self.history[-1] != FRAME_DASHBOARD:
            self._cancelar_timer_dashboard()
            return
        if FRAME_DASHBOARD not in self.frames:
            self._cancelar_timer_dashboard()
            return

        db_path = get_db_path()
        try:
            current_mtime = os.path.getmtime(db_path) if os.path.isfile(db_path) else 0
        except Exception:
            current_mtime = 0

        last_mtime = getattr(self, "_dashboard_db_mtime", None)
        if last_mtime != current_mtime:
            self._dashboard_db_mtime = current_mtime
            historico = self._carregar_historico()
            signature = self._dashboard_signature(historico)
            if signature != getattr(self, "_dashboard_last_signature", None):
                self._dashboard_last_signature = signature
                self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], historico)

        self._cancelar_timer_dashboard()
        self._dashboard_after_id = self.root.after(10000, self._check_dashboard_refresh)

    def _stat_card(self, parent, valor, legenda):
        # Cores de destaque por tipo de card
        cores_legenda = {
            "total": ("#4fc3f7", "📋"),
            "em andamento": ("#ffbf69", "⏳"),
            "finalizadas": ("#67d28d", "✅"),
            "canceladas": ("#ff6b6b", "🗑️"),
        }
        legenda_lower = legenda.strip().lower()
        cor_accent, icone = cores_legenda.get(legenda_lower, (ACCENT, "📊"))

        card = tk.Frame(parent, bg=BG_CARD, highlightbackground=BORDA, highlightthickness=1)
        card.pack(side="left", expand=True, fill="x", padx=(0, 8))

        # Barra colorida no topo do card
        barra = tk.Frame(card, bg=cor_accent, height=4)
        barra.pack(fill="x")

        conteudo = tk.Frame(card, bg=BG_CARD)
        conteudo.pack(fill="x", padx=12, pady=(10, 10))

        # Ícone + Número lado a lado
        topo_card = tk.Frame(conteudo, bg=BG_CARD)
        topo_card.pack(fill="x")

        tk.Label(
            topo_card, text=icone, bg=BG_CARD, fg=cor_accent, font=("Arial", 16)
        ).pack(side="left", padx=(0, 8))

        tk.Label(
            topo_card, text=str(valor), bg=BG_CARD, fg=TEXTO, font=("Arial", 22, "bold")
        ).pack(side="left")

        tk.Label(
            conteudo, text=legenda.upper(), bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 8, "bold"),
            anchor="w"
        ).pack(fill="x", pady=(4, 0))

    def _linha_historico(self, parent, item):
        perm = str(self.permissao or "").upper()

        status_raw = str(item.get("status", "—")).strip().upper()
        cancelado = bool(item.get("cancelado")) or status_raw == "CANCELADO"
        custo_fechado = bool(str(item.get("custo_fechamento") or "").strip())
        status_on_going = not cancelado and not custo_fechado

        if cancelado:
            stripe_color = "#ff6b6b"
            badge_bg = "#3a1414"
            badge_fg = "#ff6b6b"
            badge_texto = "CANCELADA"
        elif status_on_going:
            stripe_color = "#ffbf69"
            badge_bg = "#3a2d14"
            badge_fg = "#ffbf69"
            badge_texto = "EM ANDAMENTO"
        else:
            stripe_color = "#67d28d"
            badge_bg = "#163321"
            badge_fg = "#67d28d"
            badge_texto = "FINALIZADA"

        card = tk.Frame(
            parent, bg=BG_CARD, cursor="hand2",
            highlightbackground=BORDA, highlightthickness=1, bd=0
        )
        card.pack(fill="x", padx=0, pady=5)

        # Barra lateral colorida de status
        stripe = tk.Frame(card, bg=stripe_color, width=5)
        stripe.pack(side="left", fill="y")

        corpo = tk.Frame(card, bg=BG_CARD, padx=14, pady=10)
        corpo.pack(side="left", fill="both", expand=True)

        topo = tk.Frame(corpo, bg=BG_CARD)
        topo.pack(fill="x", pady=(0, 4))

        btn_area = tk.Frame(topo, bg=BG_CARD)
        btn_area.pack(side="right", anchor="ne", padx=(10, 0))

        badge = tk.Label(
            btn_area, text=badge_texto, bg=badge_bg, fg=badge_fg,
            font=("Arial", 8, "bold"), padx=8, pady=3
        )
        badge.pack(side="right", padx=(6, 0))

        # Ações baseadas no perfil
        if perm in ("ADMIN", "ENGENHARIA") and not cancelado:
            btn_editar = tk.Button(
                btn_area, text="Editar",
                command=lambda item=item: self._abrir_editar_registro(item),
                bg=BG_CARD, fg=ACCENT, activebackground=BORDA, activeforeground=TEXTO,
                relief="solid", bd=1, font=("Arial", 8, "bold"), padx=8, pady=2, cursor="hand2"
            )
            btn_editar.pack(side="right", padx=(4, 0))

        if perm == "ADMIN" and not cancelado:
            btn_excluir = tk.Button(
                btn_area, text="Excluir",
                command=lambda item=item: self._excluir_registro(item),
                bg=BG_CARD, fg="#ff6b6b", activebackground=BORDA, activeforeground=TEXTO,
                relief="solid", bd=1, font=("Arial", 8, "bold"), padx=8, pady=2, cursor="hand2"
            )
            btn_excluir.pack(side="right", padx=(4, 0))

        if perm in ("ADMIN", "ENGENHARIA") and not cancelado:
            btn_atualizar_pn = tk.Button(
                btn_area, text="Atualizar Flex PN",
                bg=BG_CARD, fg=ACCENT, activebackground=BORDA, activeforeground=TEXTO,
                relief="solid", bd=1, font=("Arial", 8, "bold"), padx=8, pady=2, cursor="hand2"
            )
            btn_atualizar_pn.configure(
                command=lambda item=item, b=btn_atualizar_pn: self._atualizar_pn_dashboard(item, b)
            )
            btn_atualizar_pn.pack(side="right", padx=(4, 0))

            phoenix_fechado = bool(str(item.get("phoenix_fechamento") or "").strip())
            pegasus_iniciado = bool(str(item.get("pegasus_abertura") or "").strip())
            pegasus_finalizado = bool(str(item.get("pegasus_fechamento") or "").strip())
            custo_iniciado = bool(str(item.get("custo_abertura") or "").strip())
            custo_finalizado = bool(str(item.get("custo_fechamento") or "").strip())

            if not phoenix_fechado:
                tk.Button(
                    btn_area, text="Fechar Phoenix",
                    command=lambda item=item: self._finalizar_phoenix(item),
                    bg=BG_CARD, fg=ACCENT, relief="solid", bd=1, font=("Arial", 8, "bold"), padx=6, pady=2
                ).pack(side="left", padx=2)

            if phoenix_fechado and not pegasus_iniciado:
                tk.Button(
                    btn_area, text="Iniciar Pegasus",
                    command=lambda item=item: self._iniciar_pegasus(item),
                    bg=BG_CARD, fg=ACCENT, relief="solid", bd=1, font=("Arial", 8, "bold"), padx=6, pady=2
                ).pack(side="left", padx=2)

            if pegasus_iniciado and not pegasus_finalizado:
                tk.Button(
                    btn_area, text="Finalizar Pegasus",
                    command=lambda item=item: self._finalizar_pegasus(item),
                    bg=BG_CARD, fg=ACCENT, relief="solid", bd=1, font=("Arial", 8, "bold"), padx=6, pady=2
                ).pack(side="left", padx=2)

            if pegasus_finalizado and not custo_iniciado:
                tk.Button(
                    btn_area, text="Iniciar Custo",
                    command=lambda item=item: self._iniciar_custo(item),
                    bg=BG_CARD, fg=ACCENT, relief="solid", bd=1, font=("Arial", 8, "bold"), padx=6, pady=2
                ).pack(side="left", padx=2)

            if custo_iniciado and not custo_finalizado:
                tk.Button(
                    btn_area, text="Finalizar Custo",
                    command=lambda item=item: self._finalizar_custo(item),
                    bg=BG_CARD, fg=ACCENT, relief="solid", bd=1, font=("Arial", 8, "bold"), padx=6, pady=2
                ).pack(side="left", padx=2)

        from services.database import limpar_descricao
        ticket = str(item.get("ticket") or "").strip()
        raw_desc = str(item.get("description") or "Sem descrição").strip()
        desc_limpa = limpar_descricao(raw_desc)
        if not desc_limpa or "muito alta" in desc_limpa.lower() or "\t" in desc_limpa:
            desc_limpa = raw_desc if ("muito alta" not in raw_desc.lower() and "\t" not in raw_desc) else ""

        if not desc_limpa:
            desc_limpa = f"Solicitação Phoenix {ticket}".strip() if ticket else "Solicitação Phoenix"

        if ticket and desc_limpa and desc_limpa != ticket and not desc_limpa.endswith(ticket):
            titulo_texto = f"{ticket} - {desc_limpa}"
        else:
            titulo_texto = ticket or desc_limpa

        lbl_titulo = tk.Label(
            topo, text=titulo_texto, bg=BG_CARD, fg=TEXTO,
            font=("Arial", 10, "bold"), justify="left", anchor="w", wraplength=480
        )
        lbl_titulo.pack(side="left", fill="x", expand=True)

        div = tk.Frame(corpo, bg=BORDA, height=1)
        div.pack(fill="x", pady=(4, 8))

        info_frame = tk.Frame(corpo, bg=BG_CARD)
        info_frame.pack(fill="x")

        # Stepper Visual de Fases (Phoenix -> Pegasus -> Custo)
        phoenix_fechado = bool(str(item.get("phoenix_fechamento") or "").strip())
        pegasus_iniciado = bool(str(item.get("pegasus_abertura") or "").strip())
        pegasus_finalizado = bool(str(item.get("pegasus_fechamento") or "").strip())
        custo_iniciado = bool(str(item.get("custo_abertura") or "").strip())
        custo_finalizado = bool(str(item.get("custo_fechamento") or "").strip())

        stepper_frame = tk.Frame(info_frame, bg=BG_CARD)
        stepper_frame.pack(anchor="w", pady=(0, 6))

        fases = [
            ("Phoenix", phoenix_fechado, not phoenix_fechado and not cancelado),
            ("Pegasus", pegasus_finalizado, pegasus_iniciado and not pegasus_finalizado and not cancelado),
            ("Custo", custo_finalizado, custo_iniciado and not custo_finalizado and not cancelado),
        ]

        for idx_f, (nome_fase, concluiu, ativa) in enumerate(fases):
            if concluiu:
                st_bg, st_fg, st_txt = ("#163321", "#67d28d", f"✓ {nome_fase}")
            elif ativa:
                st_bg, st_fg, st_txt = ("#3a2d14", "#ffbf69", f"⚡ {nome_fase}")
            else:
                st_bg, st_fg, st_txt = (ENTRY_BG, TEXTO_MUTED, f"○ {nome_fase}")

            tk.Label(
                stepper_frame, text=st_txt, bg=st_bg, fg=st_fg,
                font=("Arial", 7, "bold"), padx=6, pady=2
            ).pack(side="left")

            if idx_f < len(fases) - 1:
                tk.Label(stepper_frame, text=" ➔ ", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 8)).pack(side="left")

        solicitante = str(item.get("solicitante") or "").strip() or "Não informado"
        classificador = str(item.get("classificador") or item.get("requisitante") or "").strip() or "Não informado"
        data_abert = item.get("data_abertura", "—")
        hora_abert = item.get("hora_abertura", "")
        data_str = f"{data_abert} {hora_abert}".strip()

        tk.Label(
            info_frame, text=f"Solicitante: {solicitante}   •   Classificador: {classificador}",
            bg=BG_CARD, fg=TEXTO, font=("Arial", 9, "bold")
        ).pack(anchor="w", pady=(0, 2))

        tk.Label(
            info_frame, text=f"Data de Abertura: {data_str}",
            bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 9)
        ).pack(anchor="w", pady=(0, 2))

        pn_val = str(item.get("pn") or item.get("part_number") or "").strip()
        pn_row = tk.Frame(info_frame, bg=BG_CARD)
        pn_row.pack(anchor="w", pady=(2, 2))

        tk.Label(pn_row, text="PN Capturado: ", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 9, "bold")).pack(side="left")
        tk.Label(pn_row, text=pn_val or "Não capturado", bg=BG_CARD, fg=ACCENT if pn_val else TEXTO_MUTED, font=("Arial", 9, "bold" if pn_val else "italic")).pack(side="left")

        criador = str(item.get("criado_por") or item.get("user") or "—").strip()
        criado_em = str(item.get("criado_em") or data_str).strip()
        ultima_alt = str(item.get("ultima_alteracao") or criado_em).strip()
        usr_alt = str(item.get("ultimo_usuario_alterou") or criador).strip()

        tk.Label(
            info_frame,
            text=f"Criado por: {criador} ({criado_em})   •   Última alteração: {ultima_alt} por {usr_alt}",
            bg=BG_CARD, fg="#8c8c8c", font=("Arial", 8)
        ).pack(anchor="w", pady=(2, 0))

        def mostrar_detalhes(e=None):
            hist_log = "\n".join(
                f"  • {h.get('data')} - {h.get('descricao')} ({h.get('usuario')})"
                for h in item.get("historico_alteracoes", [])
            ) or "  • Nenhum histórico registrado."

            messagebox.showinfo(
                "Detalhes do Registro",
                f"ID: {item.get('id', '—')}\n"
                f"Linha: {item.get('linha', '—')}\n"
                f"Ticket: {item.get('ticket') or '—'}\n\n"
                f"Description: {item.get('description', '—')}\n"
                f"Produto: {item.get('produto') or 'Não informado'}\n"
                f"MPN: {item.get('mpn') or 'Não informado'}\n\n"
                f"Status: {item.get('status', '—')}\n"
                f"PN Capturado: {pn_val or '—'}\n"
                f"Origem: {item.get('origem') or 'Não informado'}\n"
                f"Custo: {item.get('custo') or 'Não informado'}\n\n"
                f"Solicitante: {solicitante}\n"
                f"Classificador: {classificador}\n\n"
                f"--- AUDITORIA ---\n"
                f"Criado por: {criador} em {criado_em}\n"
                f"Última alteração: {ultima_alt} por {usr_alt}\n\n"
                f"--- HISTÓRICO DE ALTERAÇÕES ---\n{hist_log}"
            )

        card.bind("<Enter>", lambda e: card.configure(highlightbackground=ACCENT))
        card.bind("<Leave>", lambda e: card.configure(highlightbackground=BORDA))

        for widget in (corpo, topo, info_frame, lbl_titulo):
            widget.bind("<Button-1>", mostrar_detalhes)

    # Importar Planilha Excel de Solicitações Históricas
    def _importar_planilha_excel(self):
        if not self._validar_permissao("geral"):
            return

        caminho_arquivo = filedialog.askopenfilename(
            title="Importar Planilha Excel de Solicitações Históricas",
            filetypes=[("Planilhas Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )

        if not caminho_arquivo:
            return

        try:
            res = importar_excel_historico(caminho_arquivo, usuario_importou=self.usuario)
            messagebox.showinfo(
                "Importação Concluída",
                f"Planilha importada e alocada com sucesso!\n\n"
                f"• Total de produtos processados: {res['total']}\n"
                f"• Produtos Encerrados (Concluídos): {res['encerrados']}\n"
                f"• Produtos Em Andamento: {res['em_andamento']}\n"
                f"• Produtos Cancelados: {res['cancelados']}\n\n"
                "Os produtos foram devidamente separados e organizados no Dashboard."
            )
            self._abrir_dashboard()
        except Exception as exc:
            logger.error("Erro ao importar planilha Excel: %s", exc)
            messagebox.showerror(
                "Erro na Importação",
                f"Não foi possível processar a planilha selecionada:\n{exc}"
            )

    # Exportar Dashboard para Excel
    def _exportar_dashboard(self):
        if not self._validar_permissao("exportar"):
            return

        historico = self._carregar_historico()
        counts = contar_por_status()
        data_str = datetime.now().strftime("%d-%m-%Y")
        nome_arquivo = f"Phoenix_Dashboard_{data_str}.xlsx"

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=nome_arquivo,
            filetypes=[("Excel Files", "*.xlsx")],
            title="Exportar Dashboard"
        )
        if not caminho:
            return

        try:
            wb = openpyxl.Workbook()
            # Aba Resumo
            ws_resumo = wb.active
            ws_resumo.title = "Resumo"

            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

            ws_resumo.append(["Métrica", "Quantidade"])
            ws_resumo.append(["Total de Solicitações", counts["total"]])
            ws_resumo.append(["Em Andamento", counts["em_andamento"]])
            ws_resumo.append(["Finalizadas", counts["finalizadas"]])
            ws_resumo.append(["Canceladas", counts["canceladas"]])

            for cell in ws_resumo[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row in ws_resumo.iter_rows(min_row=2, max_col=2):
                row[0].font = Font(name="Arial", size=10, bold=True)
                row[1].font = Font(name="Arial", size=10)
                row[1].alignment = Alignment(horizontal="center")

            # Aba Indicadores
            ws_ind = wb.create_sheet(title="Indicadores")
            ws_ind.append(["Status", "Quantidade", "Percentual"])
            total = counts["total"] or 1
            ws_ind.append(["Em Andamento", counts["em_andamento"], f"{(counts['em_andamento']/total)*100:.1f}%"])
            ws_ind.append(["Finalizadas", counts["finalizadas"], f"{(counts['finalizadas']/total)*100:.1f}%"])
            ws_ind.append(["Canceladas", counts["canceladas"], f"{(counts['canceladas']/total)*100:.1f}%"])

            for cell in ws_ind[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto ajustar largura
            for sheet in (ws_resumo, ws_ind):
                for col in sheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(caminho)
            messagebox.showinfo("Sucesso", f"Dashboard exportado com sucesso:\n{caminho}")
        except Exception as exc:
            logger.exception("Erro ao exportar dashboard")
            messagebox.showerror("Erro", f"Falha ao exportar dashboard: {exc}")

    # Exportar Registros para Excel
    def _exportar_registros(self):
        if not self._validar_permissao("exportar"):
            return

        # Perguntar se deseja exportar todos ou filtrados
        opcao = messagebox.askyesnocancel(
            "Exportar Registros",
            "Deseja exportar TODOS os registros?\n\nClique 'Sim' para todos ou 'Não' para exportar apenas os registros do filtro/busca atual."
        )
        if opcao is None:
            return

        historico = self._carregar_historico()

        if not opcao:  # Filtrados
            filtro = (self._dashboard_filter or "todos").strip().lower()
            if filtro == "on going":
                historico = [i for i in historico if str(i.get("status", "")).upper() == "ON GOING" and not i.get("cancelado")]
            elif filtro == "finalizadas":
                historico = [i for i in historico if str(i.get("status", "")).upper() not in ("ON GOING", "CANCELADO") and not i.get("cancelado")]
            elif filtro == "canceladas":
                historico = [i for i in historico if i.get("cancelado") or str(i.get("status", "")).upper() == "CANCELADO"]

            query = getattr(self, "_dashboard_search_query", "").strip().lower()
            if query:
                historico = [
                    item for item in historico
                    if query in str(item.get("ticket", "")).lower()
                    or query in str(item.get("pn", "")).lower()
                    or query in str(item.get("part_number", "")).lower()
                    or query in str(item.get("mpn", "")).lower()
                    or query in str(item.get("description", "")).lower()
                    or query in str(item.get("solicitante", "")).lower()
                ]

        data_str = datetime.now().strftime("%d-%m-%Y")
        nome_arquivo = f"Phoenix_Registros_{data_str}.xlsx"

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=nome_arquivo,
            filetypes=[("Excel Files", "*.xlsx")],
            title="Exportar Registros"
        )
        if not caminho:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registros"

            colunas = [
                "Ticket", "Solicitante", "Classificador", "DESCRIPTION", "PRODUTO", "FLEX PN",
                "FXB Status", "FXB OPEN", "FXB CLOSE",
                "CF", "CF OPEN", "CF CLOSE",
                "CUSTO", "CUSTO OPEN", "CUSTO CLOSE"
            ]
            ws.append(colunas)

            # Estilo dos cabeçalhos combinando com as imagens de referência
            gold_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font_dark = Font(name="Arial", size=9, bold=True, color="000000")
            header_font_light = Font(name="Arial", size=9, bold=True, color="FFFFFF")

            for idx, cell in enumerate(ws[1], start=1):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if idx in (6, 7, 8, 9):  # FLEX PN, FXB Status, FXB OPEN, FXB CLOSE
                    cell.fill = blue_fill
                    cell.font = header_font_light
                else:
                    cell.fill = gold_fill
                    cell.font = header_font_dark

            green_row_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_row_font = Font(name="Arial", size=9, color="006100")
            normal_font = Font(name="Arial", size=9, color="000000")

            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            from services.database import limpar_descricao, _parse_data_hora_generica

            def fmt_dt(val):
                val_str = str(val or "").strip()
                if not val_str:
                    return ""
                dt = _parse_data_hora_generica(val_str)
                return dt.strftime("%d-%b-%y") if dt else val_str

            for item in historico:
                solic = str(item.get("solicitante") or "").strip()
                classif = str(item.get("classificador") or item.get("requisitante") or item.get("criado_por") or "").strip()

                ticket = str(item.get("ticket") or "").strip()
                raw_desc = str(item.get("description") or "").strip()
                desc = limpar_descricao(raw_desc) or raw_desc
                if "muito alta" in desc.lower() or "\t" in desc:
                    desc = f"Solicitação Phoenix {ticket}".strip() if ticket else "Solicitação Phoenix"

                prod = str(item.get("produto") or "").strip()
                pn = str(item.get("pn") or item.get("part_number") or "").strip()

                status_raw = str(item.get("status") or "").strip().upper()
                cancelado = bool(item.get("cancelado")) or status_raw == "CANCELADO"

                # Phoenix / FXB
                fxb_op_raw = item.get("data_abertura") or ""
                fxb_cl_raw = item.get("phoenix_fechamento") or ""
                fxb_op = fmt_dt(fxb_op_raw)
                fxb_cl = fmt_dt(fxb_cl_raw)
                fxb_st = "CANCELADO" if cancelado else ("OK" if fxb_cl_raw else ("ON GOING" if fxb_op_raw else "OPEN"))

                # Pegasus / CF
                cf_op_raw = item.get("pegasus_abertura") or ""
                cf_cl_raw = item.get("pegasus_fechamento") or ""
                cf_op = fmt_dt(cf_op_raw)
                cf_cl = fmt_dt(cf_cl_raw)
                cf_st = "CANCELADO" if cancelado else ("OK" if cf_cl_raw else ("ON GOING" if cf_op_raw else "OPEN"))

                # Custo
                cst_op_raw = item.get("custo_abertura") or ""
                cst_cl_raw = item.get("custo_fechamento") or ""
                cst_op = fmt_dt(cst_op_raw)
                cst_cl = fmt_dt(cst_cl_raw)
                cst_fin = bool(cst_cl_raw) or (bool(item.get("custo_finalizado")) and bool(cst_cl_raw))
                cst_st = "CANCELADO" if cancelado else ("OK" if cst_fin else ("ON GOING" if cst_op_raw else "OPEN"))

                row_vals = [
                    ticket, solic, classif, desc, prod, pn,
                    fxb_st, fxb_op, fxb_cl,
                    cf_st, cf_op, cf_cl,
                    cst_st, cst_op, cst_cl
                ]
                ws.append(row_vals)

                row_idx = ws.max_row
                is_ok = (cst_fin and not cancelado)
                for col_idx in range(1, len(colunas) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if is_ok:
                        cell.fill = green_row_fill
                        cell.font = green_row_font
                    else:
                        cell.font = normal_font

            # Auto ajuste de colunas
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

            # Filtro automático e congelamento de primeira linha
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            wb.save(caminho)
            messagebox.showinfo("Sucesso", f"Registros exportados com sucesso ({len(historico)} linhas):\n{caminho}")
        except Exception as exc:
            logger.exception("Erro ao exportar registros")
            messagebox.showerror("Erro", f"Falha ao exportar registros: {exc}")

    def _atualizar_pn_dashboard(self, item, botao=None):
        if not self._validar_permissao("atualizar_pn"):
            return
        ticket = str(item.get("ticket") or "").strip()
        if not ticket:
            messagebox.showwarning("Atualizar Flex PN", "Este registro não possui ticket.")
            return

        cred = self._obter_ou_pedir_credencial("Phoenix")
        if not cred or not cred.get("login") or not cred.get("senha"):
            return

        cred_user = cred["login"]
        cred_pass = cred["senha"]

        if botao is not None:
            try:
                if botao.winfo_exists():
                    botao.configure(state="disabled", text="⏳ Buscando PN...")
            except Exception:
                pass

        def _worker():
            from automocoes.phoenix.atualizar_pn import buscar_pn_por_ticket
            resultado = buscar_pn_por_ticket(ticket, cred_user=cred_user, cred_pass=cred_pass)
            self.root.after(0, lambda: self._finalizar_atualizacao_pn(resultado, botao))

        threading.Thread(target=_worker, daemon=True).start()

    def _finalizar_atualizacao_pn(self, resultado, botao=None):
        if botao is not None:
            try:
                if botao.winfo_exists():
                    botao.configure(state="normal", text="Atualizar Flex PN")
            except Exception:
                pass

        messagebox.showinfo("Atualizar Flex PN", resultado.get("mensagem") or "Busca concluída.")

        if FRAME_DASHBOARD in self.frames:
            historico = self._carregar_historico()
            self._dashboard_last_signature = self._dashboard_signature(historico)
            self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], historico)

    def _abrir_editar_registro(self, item):
        if not self._validar_permissao("editar"):
            return

        if str(item.get("status", "")).upper() == "ENCERRADO":
            messagebox.showwarning("Registro Encerrado", "Este registro está encerrado e não pode ser editado.")
            return

        janela = tk.Toplevel(self.root)
        janela.title("Editar informações do produto")
        janela.configure(bg=BG)
        janela.attributes("-topmost", True)
        janela.resizable(False, False)
        janela.transient(self.root)

        corpo = tk.Frame(janela, bg=BG, padx=16, pady=16)
        corpo.pack(fill="both", expand=True)

        tk.Label(
            corpo,
            text=str(item.get("description") or "—").upper(),
            bg=BG, fg=TEXTO, font=("Arial", 11, "bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 14))

        campos_def = [
            ("Produto", "produto"),
            ("Solicitante", "solicitante"),
            ("Requisitante", "requisitante"),
            ("PN (Part Number)", "pn"),
            ("MPN", "mpn"),
            ("Origem", "origem"),
            ("Custo", "custo"),
        ]

        entries = {}
        for r_idx, (label_txt, campo_key) in enumerate(campos_def, start=1):
            tk.Label(corpo, text=espacar(label_txt), bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).grid(row=r_idx, column=0, sticky="w", pady=(0, 6))
            e = self.campo_entry(corpo)
            e.grid(row=r_idx, column=1, sticky="ew", pady=(0, 6), padx=(10, 0), ipady=4)
            val = str(item.get(campo_key) or "")
            if not val and campo_key == "origem":
                val = "IMPORTADO"
            e.insert(0, val)
            entries[campo_key] = e

        var_custo_fin = tk.BooleanVar(value=bool(item.get("custo_finalizado")))
        tk.Label(corpo, text=espacar("Custo Finalizado"), bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).grid(row=8, column=0, sticky="w", pady=(0, 6))
        chk_custo = tk.Checkbutton(corpo, variable=var_custo_fin, bg=BG, fg=TEXTO, activebackground=BG, selectcolor=BG_CARD, relief="flat")
        chk_custo.grid(row=8, column=1, sticky="w", pady=(0, 6), padx=(10, 0))

        corpo.grid_columnconfigure(1, weight=1)

        botoes = tk.Frame(corpo, bg=BG)
        botoes.grid(row=9, column=0, columnspan=2, pady=(16, 0))

        def salvar():
            novos = {key: e.get().strip() for key, e in entries.items()}
            novos["part_number"] = novos.get("pn", "")
            novos["custo_finalizado"] = var_custo_fin.get()

            atualizar_campos_registro(item.get("linha"), novos, usuario_alteracao=self.usuario)
            item.update(novos)
            janela.destroy()
            if FRAME_DASHBOARD in self.frames:
                historico = self._carregar_historico()
                self._dashboard_last_signature = self._dashboard_signature(historico)
                self._renderizar_dashboard_conteudo(self.frames[FRAME_DASHBOARD], historico)

        self.botao_flat(botoes, "Salvar", salvar, largura=12).pack(side="left", padx=6)
        self.botao_flat(botoes, "Cancelar", janela.destroy, largura=12).pack(side="left", padx=6)

        janela.update_idletasks()
        largura, altura = janela.winfo_width(), janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (largura // 2)
        y = (janela.winfo_screenheight() // 2) - (altura // 2)
        janela.geometry(f"+{x}+{y}")
        janela.grab_set()

    # =========================================================================
    # MENUS PHOENIX & PEGASUS
    # =========================================================================

    def _abrir_menu_phoenix(self):
        if not self._validar_permissao("phoenix"):
            return
        self._build_menu_phoenix()
        self.mostrar(FRAME_MENU_PHOENIX)

    def _executar_phoenix(self, arg=None):
        if not self._validar_permissao("phoenix"):
            return
        cred = self._obter_ou_pedir_credencial("PHOENIX")
        if not cred:
            return
        inv = arg in ("capturar", "importar")
        executar_script("automocoes", "phoenix", "phoenix.py", arg=arg, cred_user=cred["login"], cred_pass=cred["senha"], tool_user=self.usuario, invisible=inv)

    def _build_menu_phoenix(self):
        if FRAME_MENU_PHOENIX in self.frames:
            self.frames[FRAME_MENU_PHOENIX].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_MENU_PHOENIX] = frame

        self.build_menu_automacao(
            frame,
            "Phoenix",
            [
                ("Home Phoenix", lambda: self._executar_phoenix(arg="home"), "Retorna ao fluxo inicial"),
                ("Nova solicitação", lambda: self._executar_phoenix(), "Inicia uma nova solicitação"),
                ("Atualizar Flex PN", self._abrir_atualizar_pn_phoenix, "Busca e atualiza o PN"),
                ("Capturar/Importar Solicitações", lambda: self._executar_phoenix(arg="capturar"), "Varre o Phoenix e captura tickets abertos não registrados"),
            ],
        )

    def _abrir_menu_pegasus(self):
        if not self._validar_permissao("pegasus"):
            return
        self._build_menu_pegasus()
        self.mostrar(FRAME_MENU_PEGASUS)

    def _executar_pegasus(self, arg=None):
        if not self._validar_permissao("pegasus"):
            return
        cred = self._obter_ou_pedir_credencial("PEGASUS")
        if not cred:
            return
        executar_script("automocoes", "pegasus", "pegasus.py", arg=arg, cred_user=cred["login"], cred_pass=cred["senha"], tool_user=self.usuario)

    def _build_menu_pegasus(self):
        if FRAME_MENU_PEGASUS in self.frames:
            self.frames[FRAME_MENU_PEGASUS].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_MENU_PEGASUS] = frame

        self.build_menu_automacao(
            frame,
            "Pegasus",
            [
                ("Home Pegasus", lambda: self._executar_pegasus(arg="home"), "Retorna ao fluxo inicial"),
                ("Nova solicitação", lambda: self._executar_pegasus(), "Inicia uma nova solicitação"),
            ],
        )

    def _abrir_atualizar_pn_phoenix(self):
        if not self._validar_permissao("atualizar_pn"):
            return
        self._build_atualizar_pn_phoenix()
        self.mostrar(FRAME_ATUALIZAR_PN)

    def _build_atualizar_pn_phoenix(self):
        if FRAME_ATUALIZAR_PN in self.frames:
            self.frames[FRAME_ATUALIZAR_PN].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_ATUALIZAR_PN] = frame

        self.botao_voltar(frame)
        self.cabecalho(frame, "Phoenix", "Buscar PN")

        bloco = tk.Frame(frame, bg=BG)
        bloco.pack(padx=60, fill="x")

        tk.Label(
            bloco, text=espacar("Linha da planilha"), bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION
        ).pack(anchor="w", pady=(10, 6))

        campo_linha = self.campo_entry(bloco)
        campo_linha.pack(fill="x", ipady=6)

        def buscar():
            linha = campo_linha.get().strip()
            item = encontrar_por_linha(linha)
            if item is None:
                messagebox.showwarning("Phoenix", "Linha não encontrada.")
                return
            messagebox.showinfo(
                "Phoenix",
                f"Linha: {item.get('linha', '—')}\n"
                f"Ticket: {item.get('ticket') or '—'}\n"
                f"Description: {item.get('description', '—')}\n"
                f"Status: {item.get('status', '—')}\n"
                f"Data: {item.get('data_abertura', '—')}"
            )

        def rodar_automacao():
            linha = campo_linha.get().strip()
            if not linha:
                messagebox.showwarning("Phoenix", "Digite a linha.")
                return
            item = encontrar_por_linha(linha)
            if item is None:
                messagebox.showwarning("Phoenix", "Linha não encontrada no histórico.")
                return

            cred = self._obter_ou_pedir_credencial("PHOENIX")
            if not cred:
                return
            executar_script("automocoes", "phoenix", "atualizar_pn.py", arg=linha, cred_user=cred["login"], cred_pass=cred["senha"], tool_user=self.usuario)
            messagebox.showinfo("Phoenix", "Automação iniciada em segundo plano.")

        self.botao_flat(bloco, "Buscar Registro", buscar).pack(fill="x", pady=(14, 6))
        self.botao_flat(bloco, "Rodar Automação (PN)", rodar_automacao).pack(fill="x", pady=(0, 0))

    # =========================================================================
    # TELA MINHAS CREDENCIAIS
    # =========================================================================

    def _abrir_minhas_credenciais(self):
        if not self._validar_permissao("credenciais"):
            return
        self._build_minhas_credenciais()
        self.mostrar(FRAME_CREDENCIAIS)

    def _build_minhas_credenciais(self):
        if FRAME_CREDENCIAIS in self.frames:
            self.frames[FRAME_CREDENCIAIS].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_CREDENCIAIS] = frame

        self.botao_voltar(frame)
        self.cabecalho(frame, "Segurança", "Minhas Credenciais")

        painel = tk.Frame(frame, bg=BG_CARD, highlightbackground=BORDA, highlightthickness=0)
        painel.pack(fill="x", padx=28, pady=(0, 20))

        tk.Label(
            painel, text="Gerenciador de Credenciais Individuais", bg=BG_CARD, fg=TEXTO, font=("Arial", 12, "bold")
        ).pack(anchor="w", padx=16, pady=(16, 6))

        tk.Label(
            painel,
            text="Cadastre suas credenciais para automações Phoenix, Pegasus e Cost Request.\nProtegidas por um PIN pessoal: nem administradores conseguem visualizá-las.",
            bg=BG_CARD, fg=TEXTO_MUTED, font=FONT_SUBTITULO, justify="left"
        ).pack(anchor="w", padx=16, pady=(0, 16))

        tipos = [("PHOENIX", "Automação Phoenix (Portal)"), ("PEGASUS", "Automação Pegasus"), ("COST", "Automação Cost Request")]

        for tipo, desc in tipos:
            configurada = credencial_configurada(self.usuario_id, tipo)

            card = tk.Frame(frame, bg=BG_CARD, highlightbackground=BORDA, highlightthickness=1)
            card.pack(fill="x", padx=28, pady=6)

            corpo = tk.Frame(card, bg=BG_CARD, padx=16, pady=12)
            corpo.pack(fill="x")

            lbl_t = tk.Label(corpo, text=f"CREDENCIAIS {tipo}", bg=BG_CARD, fg=ACCENT, font=("Arial", 10, "bold"))
            lbl_t.pack(anchor="w")

            lbl_d = tk.Label(corpo, text=desc, bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 8))
            lbl_d.pack(anchor="w", pady=(0, 8))

            status_str = f"Status: {'Cadastrada (protegida por PIN)' if configurada else 'Não cadastrada'}"
            tk.Label(corpo, text=status_str, bg=BG_CARD, fg=TEXTO, font=("Arial", 9)).pack(anchor="w", pady=(0, 8))

            def _editar_cred(t=tipo):
                self._obter_ou_pedir_credencial(t, forcar=True)
                self._build_minhas_credenciais()
                self.mostrar(FRAME_CREDENCIAIS)

            btn = self.botao_flat(corpo, "Alterar Credenciais", _editar_cred, largura=18)
            btn.pack(anchor="w")

        def _esqueci_pin():
            if not messagebox.askyesno(
                "Esqueci meu PIN",
                "Isso vai apagar TODAS as suas credenciais salvas (Phoenix, Pegasus, Cost) e "
                "permitir que você crie um novo PIN. Elas precisarão ser recadastradas. Continuar?"
            ):
                return
            resetar_pin_credenciais(self.usuario_id)
            self._cred_pin = None
            messagebox.showinfo("PIN redefinido", "Suas credenciais anteriores foram removidas. Cadastre-as novamente quando precisar.")
            self._build_minhas_credenciais()
            self.mostrar(FRAME_CREDENCIAIS)

        btn_reset = self.botao_flat(frame, "Esqueci meu PIN (apagar credenciais e redefinir)", _esqueci_pin, largura=40)
        btn_reset.pack(anchor="center", padx=28, pady=(10, 0))

    # =========================================================================
    # TELA ÁREA ADMINISTRATIVA (Apenas ADMIN)
    # =========================================================================

    def _abrir_administracao(self):
        if not self._validar_permissao("admin"):
            return
        self._build_administracao()
        self.mostrar(FRAME_ADMIN)

    def _build_administracao(self):
        if not self._validar_permissao("admin"):
            return

        if FRAME_ADMIN in self.frames:
            self.frames[FRAME_ADMIN].destroy()

        frame = tk.Frame(self.container, bg=BG)
        self.frames[FRAME_ADMIN] = frame

        # Container centralizado
        center_box = tk.Frame(frame, bg=BG)
        center_box.pack(anchor="n", expand=True, fill="both", padx=40, pady=20)

        self.botao_voltar(center_box)
        self.cabecalho(center_box, "Administração", "Gerenciamento de Usuários")

        stats_db = obter_estatisticas_banco()

        # Painel Visual de Métricas do Banco de Dados
        db_panel = tk.Frame(center_box, bg=BG)
        db_panel.pack(fill="x", pady=(0, 14))

        # Card 1: Status do Banco
        c1 = tk.Frame(db_panel, bg=BG_CARD, bd=0, highlightthickness=0)
        c1.pack(side="left", expand=True, fill="x", padx=(0, 6))
        c1_in = tk.Frame(c1, bg=BG_CARD, padx=14, pady=12)
        c1_in.pack(fill="x")
        tk.Label(c1_in, text=f"🗄️ {stats_db['nome_arquivo']}", bg=BG_CARD, fg=ACCENT, font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Label(c1_in, text=f"Tamanho: {stats_db['tamanho_mb']} MB   •   Modo: {stats_db['modo_journal']}", bg=BG_CARD, fg=TEXTO, font=("Arial", 8)).pack(anchor="w", pady=(2, 0))
        tk.Label(c1_in, text=f"Última alteração: {stats_db['data_modificacao']}", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 7)).pack(anchor="w")

        # Card 2: Registros e Usuários
        c2 = tk.Frame(db_panel, bg=BG_CARD, bd=0, highlightthickness=0)
        c2.pack(side="left", expand=True, fill="x", padx=4)
        c2_in = tk.Frame(c2, bg=BG_CARD, padx=14, pady=12)
        c2_in.pack(fill="x")
        tk.Label(c2_in, text=f"📊 Registros: {stats_db['total_solicitacoes']} solicitações", bg=BG_CARD, fg=ACCENT, font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Label(c2_in, text=f"Usuários Cadastrados: {stats_db['total_usuarios']}   •   Credenciais: {stats_db['total_credenciais']}", bg=BG_CARD, fg=TEXTO, font=("Arial", 8)).pack(anchor="w", pady=(2, 0))
        tk.Label(c2_in, text="Permissão de Escrita: Administrador", bg=BG_CARD, fg="#67d28d", font=("Arial", 7, "bold")).pack(anchor="w")

        # Card 3: Backups Diários
        c3 = tk.Frame(db_panel, bg=BG_CARD, bd=0, highlightthickness=0)
        c3.pack(side="left", expand=True, fill="x", padx=(6, 0))
        c3_in = tk.Frame(c3, bg=BG_CARD, padx=14, pady=12)
        c3_in.pack(fill="x")
        tk.Label(c3_in, text=f"💾 Backups Diários: {stats_db['total_backups']} salvos", bg=BG_CARD, fg=ACCENT, font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Label(c3_in, text="Backup diário automático gerado no fim do dia", bg=BG_CARD, fg=TEXTO, font=("Arial", 8)).pack(anchor="w", pady=(2, 0))
        tk.Label(c3_in, text=f"Pasta: /backups", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 7)).pack(anchor="w")

        top_bar = tk.Frame(center_box, bg=BG)
        top_bar.pack(fill="x", pady=(6, 16))

        self.botao_flat(top_bar, "+ Adicionar Usuário", self._dialog_adicionar_usuario).pack(side="left")
        self.botao_flat(top_bar, "Criar Backup Agora", self._executar_backup_manual, primario=False).pack(side="left", padx=(6, 0))
        self.botao_flat(top_bar, "Alterar Banco (Rede)", self._alterar_local_banco, primario=False).pack(side="left", padx=(6, 0))
        self.botao_flat(top_bar, "Exportar Usuários", self._exportar_usuarios, primario=False).pack(side="left", padx=(6, 0))
        self.botao_flat(top_bar, "Atualizar Lista", self._abrir_administracao, primario=False).pack(side="left", padx=(6, 0))

        usuarios = listar_usuarios()

        container_users = tk.Frame(center_box, bg=BG)
        container_users.pack(fill="both", pady=(0, 16))

        for u in usuarios:
            card = tk.Frame(container_users, bg=BG_CARD, bd=0, highlightthickness=0)
            card.pack(fill="x", pady=4)

            corpo = tk.Frame(card, bg=BG_CARD, padx=16, pady=12)
            corpo.pack(fill="x")

            col_left = tk.Frame(corpo, bg=BG_CARD)
            col_left.pack(side="left", fill="x", expand=True)

            status_str = "ATIVO" if u["ativo"] else "DESATIVADO"
            status_col = "#67d28d" if u["ativo"] else "#ff6b6b"

            tk.Label(
                col_left, text=u["nome"],
                bg=BG_CARD, fg=TEXTO, font=("Arial", 10, "bold")
            ).pack(anchor="w")

            tk.Label(
                col_left, text=f"Permissão: {u['permissao']}   •   Status: {status_str}",
                bg=BG_CARD, fg=status_col, font=("Arial", 9)
            ).pack(anchor="w", pady=(2, 0))

            col_right = tk.Frame(corpo, bg=BG_CARD)
            col_right.pack(side="right")

            u_id = u["id"]

            btn_edit = tk.Button(
                col_right, text="Editar",
                command=lambda uid=u_id: self._dialog_editar_usuario(uid),
                bg=BG_CARD, fg=ACCENT, relief="flat", bd=0, font=("Arial", 9, "bold"), padx=8, pady=4, cursor="hand2"
            )
            btn_edit.pack(side="left", padx=4)

            if u["ativo"]:
                btn_toggle = tk.Button(
                    col_right, text="Desativar",
                    command=lambda uid=u_id: self._toggle_status_usuario(uid, True),
                    bg=BG_CARD, fg="#ff6b6b", relief="flat", bd=0, font=("Arial", 9, "bold"), padx=8, pady=4, cursor="hand2"
                )
            else:
                btn_toggle = tk.Button(
                    col_right, text="Reativar",
                    command=lambda uid=u_id: self._toggle_status_usuario(uid, False),
                    bg=BG_CARD, fg="#67d28d", relief="flat", bd=0, font=("Arial", 9, "bold"), padx=8, pady=4, cursor="hand2"
                )
            btn_toggle.pack(side="left", padx=4)

            btn_del = tk.Button(
                col_right, text="Excluir",
                command=lambda uid=u_id, unome=u["nome"]: self._excluir_usuario_confirm(uid, unome),
                bg=BG_CARD, fg="#ff0505", relief="flat", bd=0, font=("Arial", 9, "bold"), padx=8, pady=4, cursor="hand2"
            )
            btn_del.pack(side="left", padx=4)

    def _excluir_usuario_confirm(self, user_id: int, nome: str):
        if not self._validar_permissao("admin"):
            return
        if user_id == self.usuario_id:
            messagebox.showwarning("Aviso", "Você não pode excluir seu próprio usuário atualmente conectado.")
            return
        res = messagebox.askyesno(
            "Confirmar Exclusão",
            f"Tem certeza que deseja excluir permanentemente o usuário '{nome}'?",
            parent=self.root
        )
        if res:
            if excluir_usuario(user_id):
                messagebox.showinfo("Sucesso", f"Usuário '{nome}' excluído com sucesso.")
                self._abrir_administracao()
            else:
                messagebox.showerror("Erro", f"Falha ao excluir o usuário '{nome}'.")

    def _executar_backup_manual(self):
        if not self._validar_permissao("admin"):
            return
        dest = fazer_backup("manual_interface")
        if dest:
            verificar_e_criar_backup_diario()
            messagebox.showinfo("Backup Criado", f"Backup corporativo gerado com sucesso em:\n{dest}")
            self._abrir_administracao()
        else:
            messagebox.showerror("Erro", "Não foi possível criar o backup manual.")

    def _alterar_local_banco(self):
        if not self._validar_permissao("admin"):
            return
        novo_caminho = filedialog.asksaveasfilename(
            title="Selecionar / Definir Local do Banco de Dados (.db)",
            initialfile="phoenix_tool.db",
            filetypes=[("Banco SQLite", "*.db"), ("Todos os Arquivos", "*.*")]
        )
        if novo_caminho:
            try:
                set_db_path(novo_caminho)
                inicializar_banco()
                messagebox.showinfo("Sucesso", f"Caminho do banco de dados alterado com sucesso para:\n{novo_caminho}")
                self._abrir_administracao()
            except Exception as exc:
                messagebox.showerror("Erro", f"Não foi possível alterar o banco de dados:\n{exc}")

    def _exportar_usuarios(self):
        if not self._validar_permissao("admin"):
            return
        caminho = filedialog.asksaveasfilename(
            title="Exportar Cadastro de Usuários",
            initialfile="usuarios_phoenix_tool.json",
            defaultextension=".json",
            filetypes=[("Arquivo JSON", "*.json")]
        )
        if caminho:
            try:
                qtd = exportar_usuarios_json(caminho)
                messagebox.showinfo(
                    "Sucesso",
                    f"{qtd} usuário(s) exportado(s) com sucesso para:\n{caminho}\n\n"
                    "Por segurança, os códigos de engenharia (login) NÃO são incluídos "
                    "neste arquivo. Ao importar em outro banco, novos códigos serão "
                    "gerados automaticamente para cada usuário novo."
                )
            except Exception as exc:
                messagebox.showerror("Erro", f"Erro ao exportar usuários:\n{exc}")

    def _importar_usuarios(self):
        if not self._validar_permissao("admin"):
            return
        caminho = filedialog.askopenfilename(
            title="Importar Cadastro de Usuários",
            filetypes=[("Arquivo JSON", "*.json")]
        )
        if caminho:
            try:
                imp, exis = importar_usuarios_json(caminho)
                codigos_gerados = obter_e_limpar_codigos_importados()
                msg = f"Importação concluída!\n\nImportados com sucesso: {imp}\nJá existentes/ignorados: {exis}"
                if codigos_gerados:
                    linhas_codigos = "\n".join(f"• {c['nome']}: {c['codigo']}" for c in codigos_gerados)
                    msg += (
                        "\n\nNovos códigos de engenharia gerados (anote e informe com segurança a "
                        f"cada usuário — não serão mostrados novamente):\n{linhas_codigos}"
                    )
                messagebox.showinfo("Sucesso", msg)
                self._abrir_administracao()
            except Exception as exc:
                messagebox.showerror("Erro", f"Erro ao importar usuários:\n{exc}")

    def _dialog_adicionar_usuario(self):
        if not self._validar_permissao("admin"):
            return

        janela = tk.Toplevel(self.root)
        janela.title("Adicionar Usuário")
        janela.configure(bg=BG)
        janela.attributes("-topmost", True)
        janela.resizable(False, False)

        corpo = tk.Frame(janela, bg=BG, padx=20, pady=20)
        corpo.pack(fill="both", expand=True)

        db_path_atual = get_db_path()

        tk.Label(corpo, text="Novo Usuário", bg=BG, fg=TEXTO, font=("Arial", 12, "bold")).pack(anchor="w", pady=(0, 2))
        tk.Label(corpo, text=f"Salvando no Banco: {os.path.basename(db_path_atual)}", bg=BG, fg=ACCENT, font=("Arial", 8, "italic")).pack(anchor="w", pady=(0, 10))

        tk.Label(corpo, text="Nome Completo:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        e_nome = self.campo_entry(corpo)
        e_nome.pack(fill="x", ipady=6, pady=(0, 8))

        tk.Label(corpo, text="Código Engenharia:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        e_cod = self.campo_entry(corpo)
        e_cod.pack(fill="x", ipady=6, pady=(0, 2))
        tk.Label(corpo, text="O código será protegido (não fica salvo em texto puro).", bg=BG, fg=TEXTO_MUTED, font=("Arial", 7, "italic")).pack(anchor="w", pady=(0, 8))

        tk.Label(corpo, text="Permissão:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        cb_perm = ttk.Combobox(corpo, values=["ADMIN", "ENGENHARIA", "VISITANTE"], state="readonly", font=("Arial", 10))
        cb_perm.set("ENGENHARIA")
        cb_perm.pack(fill="x", ipady=4, pady=(0, 16))

        def salvar():
            nome = e_nome.get().strip()
            cod = e_cod.get().strip()
            perm = cb_perm.get().strip()
            if not nome or not cod:
                messagebox.showwarning("Aviso", "Preencha Nome e Código.", parent=janela)
                return
            res = adicionar_usuario(nome, cod, perm)
            if res:
                messagebox.showinfo("Sucesso", "Usuário cadastrado com sucesso.", parent=janela)
                janela.destroy()
                self._abrir_administracao()
            else:
                messagebox.showerror("Erro", "Não foi possível cadastrar (nome pode já existir).", parent=janela)

        self.botao_flat(corpo, "Salvar", salvar).pack(fill="x")

        janela.update_idletasks()
        janela.geometry(f"+{(janela.winfo_screenwidth()//2)-150}+{(janela.winfo_screenheight()//2)-180}")
        janela.grab_set()

    def _dialog_editar_usuario(self, user_id: int):
        if not self._validar_permissao("admin"):
            return

        u = obter_usuario_por_id(user_id)
        if not u:
            return

        janela = tk.Toplevel(self.root)
        janela.title("Editar Usuário")
        janela.configure(bg=BG)
        janela.attributes("-topmost", True)
        janela.resizable(False, False)

        corpo = tk.Frame(janela, bg=BG, padx=20, pady=20)
        corpo.pack(fill="both", expand=True)

        tk.Label(corpo, text=f"Editar {u['nome']}", bg=BG, fg=TEXTO, font=("Arial", 12, "bold")).pack(anchor="w", pady=(0, 10))

        tk.Label(corpo, text="Nome Completo:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        e_nome = self.campo_entry(corpo)
        e_nome.pack(fill="x", ipady=6, pady=(0, 8))
        e_nome.insert(0, u["nome"])

        tk.Label(corpo, text="Código Engenharia:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        e_cod = self.campo_entry(corpo)
        e_cod.pack(fill="x", ipady=6, pady=(0, 2))
        tk.Label(
            corpo, text="Por segurança, o código atual não é exibido. Deixe em branco para mantê-lo,\nou digite um novo código para substituí-lo.",
            bg=BG, fg=TEXTO_MUTED, font=("Arial", 7, "italic"), justify="left"
        ).pack(anchor="w", pady=(0, 8))

        tk.Label(corpo, text="Permissão:", bg=BG, fg=TEXTO_MUTED, font=FONT_CAPTION).pack(anchor="w", pady=(4, 2))
        cb_perm = ttk.Combobox(corpo, values=["ADMIN", "ENGENHARIA", "VISITANTE"], state="readonly", font=("Arial", 10))
        cb_perm.set(u["permissao"])
        cb_perm.pack(fill="x", ipady=4, pady=(0, 16))

        def salvar():
            nome = e_nome.get().strip()
            cod = e_cod.get().strip()
            perm = cb_perm.get().strip()
            if not nome:
                messagebox.showwarning("Aviso", "Preencha o Nome.", parent=janela)
                return
            ok = editar_usuario(user_id, nome, cod, perm)
            if ok:
                messagebox.showinfo("Sucesso", "Usuário atualizado com sucesso.", parent=janela)
                janela.destroy()
                self._abrir_administracao()
            else:
                messagebox.showerror("Erro", "Falha ao editar usuário.", parent=janela)

        self.botao_flat(corpo, "Salvar Alterações", salvar).pack(fill="x")

        janela.update_idletasks()
        janela.geometry(f"+{(janela.winfo_screenwidth()//2)-150}+{(janela.winfo_screenheight()//2)-180}")
        janela.grab_set()

    def _toggle_status_usuario(self, user_id: int, desativar: bool):
        if not self._validar_permissao("admin"):
            return

        if user_id == self.usuario_id:
            messagebox.showwarning("Aviso", "Você não pode desativar seu próprio usuário.")
            return

        u = obter_usuario_por_id(user_id)
        if not u:
            return

        acao_str = "desativar" if desativar else "reativar"
        if messagebox.askyesno("Confirmar", f"Deseja realmente {acao_str} o usuário '{u['nome']}'?"):
            if desativar:
                desativar_usuario(user_id)
            else:
                reativar_usuario(user_id)
            self._abrir_administracao()


    # =====================================================================
    # CENTRAL DE AJUDA / TUTORIAL INTERATIVO
    # =====================================================================
    def _abrir_ajuda(self):
        """Abre a Central de Ajuda com tutorial completo da ferramenta."""
        ajuda = tk.Toplevel(self.root)
        ajuda.title("Central de Ajuda — Flex-tax Classification")
        ajuda.geometry("920x700")
        ajuda.configure(bg=BG)
        ajuda.transient(self.root)
        ajuda.grab_set()

        # Tentar centralizar
        ajuda.update_idletasks()
        w = ajuda.winfo_width()
        h = ajuda.winfo_height()
        x = (ajuda.winfo_screenwidth() // 2) - (w // 2)
        y = (ajuda.winfo_screenheight() // 2) - (h // 2)
        ajuda.geometry(f"+{x}+{y}")

        # Header
        hdr = tk.Frame(ajuda, bg=ACCENT, height=60)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📖  CENTRAL DE AJUDA", bg=ACCENT, fg="#ffffff", font=("Arial", 16, "bold")).pack(side="left", padx=20, pady=12)
        tk.Label(hdr, text="Flex-tax Classification v1.0", bg=ACCENT, fg="#ffe0c0", font=("Arial", 9)).pack(side="right", padx=20)

        # Sidebar + Content area
        body = tk.Frame(ajuda, bg=BG)
        body.pack(fill="both", expand=True)

        sidebar = tk.Frame(body, bg=BG_CARD, width=200, bd=0, highlightthickness=0)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        content_area = tk.Frame(body, bg=BG)
        content_area.pack(side="left", fill="both", expand=True)

        # Scrollable content
        canvas = tk.Canvas(content_area, bg=BG, highlightthickness=0, bd=0)
        scrollbar = tk.Scrollbar(content_area, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        content_frame = tk.Frame(canvas, bg=BG)
        content_window = canvas.create_window((0, 0), window=content_frame, anchor="nw")

        def _atualizar_scroll(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _ajustar_largura(event):
            canvas.itemconfig(content_window, width=event.width)

        content_frame.bind("<Configure>", _atualizar_scroll)
        canvas.bind("<Configure>", _ajustar_largura)

        def _scroll(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _scroll)
        ajuda.protocol("WM_DELETE_WINDOW", lambda: (canvas.unbind_all("<MouseWheel>"), ajuda.destroy()))

        # ------------------------------------------------------------------
        # Helpers de renderização
        # ------------------------------------------------------------------
        def _limpar():
            for w in content_frame.winfo_children():
                w.destroy()
            canvas.yview_moveto(0)

        def _titulo_secao(texto, icone=""):
            f = tk.Frame(content_frame, bg=ACCENT, height=3)
            f.pack(fill="x", padx=24, pady=(20, 0))
            tk.Label(content_frame, text=f"{icone}  {texto}", bg=BG, fg=TEXTO, font=("Arial", 18, "bold")).pack(anchor="w", padx=28, pady=(12, 4))

        def _subtitulo(texto):
            tk.Label(content_frame, text=texto, bg=BG, fg=ACCENT, font=("Arial", 13, "bold")).pack(anchor="w", padx=28, pady=(14, 4))

        def _paragrafo(texto):
            lbl = tk.Label(content_frame, text=texto, bg=BG, fg=TEXTO, font=("Arial", 10), wraplength=620, justify="left")
            lbl.pack(anchor="w", padx=32, pady=(2, 6))

        def _bullet(texto, indent=40):
            tk.Label(content_frame, text=f"  •  {texto}", bg=BG, fg=TEXTO, font=("Arial", 10), wraplength=600, justify="left").pack(anchor="w", padx=indent, pady=1)

        def _dica(texto):
            f = tk.Frame(content_frame, bg="#1a3a1a" if tema_atual() == "dark" else "#e8f5e9", bd=0, highlightthickness=1, highlightbackground="#4caf50")
            f.pack(fill="x", padx=32, pady=(6, 8))
            tk.Label(f, text=f"💡 DICA: {texto}", bg=f.cget("bg"), fg="#4caf50" if tema_atual() == "dark" else "#2e7d32", font=("Arial", 9, "bold"), wraplength=580, justify="left").pack(anchor="w", padx=12, pady=8)

        def _aviso(texto):
            f = tk.Frame(content_frame, bg="#3a2a0a" if tema_atual() == "dark" else "#fff3e0", bd=0, highlightthickness=1, highlightbackground=ACCENT)
            f.pack(fill="x", padx=32, pady=(6, 8))
            tk.Label(f, text=f"⚠️ ATENÇÃO: {texto}", bg=f.cget("bg"), fg=ACCENT, font=("Arial", 9, "bold"), wraplength=580, justify="left").pack(anchor="w", padx=12, pady=8)

        def _tabela(headers, rows):
            """Desenha uma tabela formatada dentro do content_frame."""
            tbl = tk.Frame(content_frame, bg=BORDA, bd=1, relief="solid")
            tbl.pack(fill="x", padx=32, pady=(6, 10))

            # Cabeçalho
            hdr_frame = tk.Frame(tbl, bg=ACCENT)
            hdr_frame.pack(fill="x")
            for j, h in enumerate(headers):
                cell = tk.Label(hdr_frame, text=h, bg=ACCENT, fg="#ffffff", font=("Arial", 9, "bold"), padx=10, pady=6, anchor="w")
                cell.grid(row=0, column=j, sticky="nsew", padx=(0, 1))
                hdr_frame.grid_columnconfigure(j, weight=1)

            # Linhas
            for i, row in enumerate(rows):
                row_bg = BG_CARD if i % 2 == 0 else BG
                row_frame = tk.Frame(tbl, bg=row_bg)
                row_frame.pack(fill="x")
                for j, val in enumerate(row):
                    cell = tk.Label(row_frame, text=val, bg=row_bg, fg=TEXTO, font=("Arial", 9), padx=10, pady=5, anchor="w", wraplength=250)
                    cell.grid(row=0, column=j, sticky="nsew", padx=(0, 1))
                    row_frame.grid_columnconfigure(j, weight=1)

        def _diagrama_fluxo(etapas):
            """Desenha um diagrama de fluxo horizontal com setas."""
            f = tk.Frame(content_frame, bg=BG)
            f.pack(fill="x", padx=32, pady=(8, 12))
            for i, (icone, texto, cor) in enumerate(etapas):
                box = tk.Frame(f, bg=cor, bd=0, highlightthickness=1, highlightbackground=BORDA)
                box.pack(side="left", padx=2, pady=4)
                tk.Label(box, text=icone, bg=cor, fg="#ffffff", font=("Arial", 14)).pack(padx=8, pady=(6, 0))
                tk.Label(box, text=texto, bg=cor, fg="#ffffff", font=("Arial", 8, "bold"), wraplength=90).pack(padx=8, pady=(0, 6))
                if i < len(etapas) - 1:
                    tk.Label(f, text="→", bg=BG, fg=ACCENT, font=("Arial", 16, "bold")).pack(side="left", padx=2)

        def _separador():
            tk.Frame(content_frame, bg=BORDA, height=1).pack(fill="x", padx=28, pady=(16, 8))

        # ------------------------------------------------------------------
        # Conteúdo de cada seção
        # ------------------------------------------------------------------
        def _show_visao_geral():
            _limpar()
            _titulo_secao("Visão Geral do Sistema", "🏠")
            _paragrafo(
                "O Flex-tax Classification é uma ferramenta corporativa para gerenciar "
                "solicitações de classificação fiscal de componentes eletrônicos. "
                "Ele automatiza o fluxo de trabalho entre três sistemas: Phoenix, Pegasus e Cost."
            )

            _subtitulo("Fluxo Completo de uma Solicitação")
            _diagrama_fluxo([
                ("🦅", "PHOENIX\nAbrir Solicitação", "#1565c0"),
                ("📋", "PEGASUS\nRegistrar Item", "#6a1b9a"),
                ("💰", "CUSTO\nAnálise Fiscal", "#e65100"),
                ("✅", "FINALIZADO\nPN Atribuído", "#2e7d32"),
            ])

            _paragrafo(
                "Cada solicitação passa pelas etapas acima. A ferramenta registra "
                "automaticamente as datas de abertura e fechamento de cada fase, "
                "permitindo rastreabilidade total do processo."
            )

            _subtitulo("Estrutura da Ferramenta")
            _tabela(
                ["Módulo", "Função", "Acesso"],
                [
                    ["Dashboard", "Painel central com todos os registros, gráficos e exportações", "Todos"],
                    ["Phoenix", "Abrir novas solicitações e capturar tickets do sistema Phoenix", "Engenharia / Admin"],
                    ["Pegasus", "Registrar itens no sistema Pegasus", "Engenharia / Admin"],
                    ["Credenciais", "Salvar login/senha dos sistemas (protegido por PIN)", "Engenharia / Admin"],
                    ["Administração", "Gerenciar usuários, backups e configurações", "Admin"],
                ]
            )

            _dica("Use a barra de navegação inferior (DASHBOARD • PHOENIX • PEGASUS • CREDENCIAIS • ADMINISTRAÇÃO) para navegar rapidamente entre os módulos.")

        def _show_dashboard():
            _limpar()
            _titulo_secao("Dashboard", "📊")
            _paragrafo(
                "O Dashboard é a tela principal da ferramenta. Ele mostra todas as "
                "solicitações registradas, estatísticas em tempo real e permite "
                "realizar ações em cada registro."
            )

            _subtitulo("Cards de Estatísticas")
            _paragrafo(
                "No topo do dashboard, 4 cards mostram a contagem de solicitações:"
            )
            _tabela(
                ["Card", "O que mostra", "Cor"],
                [
                    ["Total", "Quantidade total de solicitações registradas", "Branco"],
                    ["Em Andamento", "Solicitações com status 'ON GOING'", "Azul"],
                    ["Finalizadas", "Solicitações concluídas com sucesso", "Verde"],
                    ["Canceladas", "Solicitações canceladas/excluídas", "Vermelho"],
                ]
            )

            _subtitulo("Gráfico de Pizza")
            _paragrafo(
                "O gráfico donut mostra a distribuição percentual das solicitações "
                "entre Em Andamento, Finalizadas e Canceladas. Ele é atualizado "
                "automaticamente ao abrir o dashboard."
            )

            _subtitulo("Barra de Ações")
            _tabela(
                ["Botão", "Ação"],
                [
                    ["Atualizar", "Recarrega o dashboard com os dados mais recentes"],
                    ["Capturar Phoenix", "Abre o Phoenix invisível e captura tickets que não foram registrados"],
                    ["Exp. Dashboard", "Exporta um resumo em Excel (.xlsx) com métricas e indicadores"],
                    ["Exp. Registros", "Exporta todos os registros para Excel com datas de abertura/fechamento por fase"],
                    ["Esvaziar Lixeira", "Remove permanentemente os registros cancelados"],
                ]
            )

            _subtitulo("Filtros e Busca")
            _paragrafo(
                "Use os filtros (Todos / Em Andamento / Finalizadas / Lixeira) para visualizar "
                "apenas registros de um determinado status. A barra de busca filtra por "
                "ticket, description, MPN, PN ou solicitante."
            )

            _subtitulo("Botões em Cada Registro")
            _tabela(
                ["Botão", "Quando aparece", "O que faz"],
                [
                    ["Finalizar Phoenix", "Quando o status é ON GOING", "Marca a fase Phoenix como fechada e registra a data"],
                    ["Iniciar Pegasus", "Após Phoenix fechado", "Registra a data de abertura do Pegasus"],
                    ["Finalizar Pegasus", "Pegasus iniciado", "Marca o Pegasus como fechado"],
                    ["Iniciar Custo", "Após Pegasus fechado", "Registra a abertura da análise de Custo"],
                    ["Finalizar Custo", "Custo iniciado", "Conclui a análise de custo"],
                    ["Atualizar Flex PN", "Sempre", "Busca o Part Number atribuído no Phoenix"],
                    ["Editar", "Sempre", "Abre um formulário para editar os campos do registro"],
                    ["Excluir", "Sempre", "Cancela/exclui o registro (vai para a lixeira)"],
                ]
            )

            _dica("Clique no ticket ou na description de qualquer registro para ver os detalhes completos.")
            _aviso("Registros excluídos vão para a lixeira. Use 'Esvaziar Lixeira' para removê-los permanentemente.")

        def _show_phoenix():
            _limpar()
            _titulo_secao("Módulo Phoenix", "🦅")
            _paragrafo(
                "O módulo Phoenix automatiza a criação de novas solicitações de classificação "
                "fiscal no sistema Phoenix da Flex."
            )

            _subtitulo("Fluxo: Nova Solicitação")
            _diagrama_fluxo([
                ("1️⃣", "Selecionar\nTemplate Excel", "#1565c0"),
                ("2️⃣", "Upload +\nClique na Lupa", "#0277bd"),
                ("3️⃣", "Preencher\nManualmente", "#00838f"),
                ("4️⃣", "Confirmar\nSolicitação", "#00695c"),
                ("5️⃣", "Ticket\nCapturado", "#2e7d32"),
            ])

            _paragrafo("Passo a passo detalhado:")
            _bullet("1. Clique em 'Nova Solicitação' no menu Phoenix")
            _bullet("2. O navegador Chromium abre automaticamente e faz login no Phoenix")
            _bullet("3. Selecione o Task P0032 (preenchido automaticamente)")
            _bullet("4. Escolha o arquivo template Excel (.xlsx) com os itens")
            _bullet("5. O template é carregado e a Lupa é clicada automaticamente")
            _bullet("6. A ferramenta captura a Description e o MPN do formulário")
            _bullet("7. Campos automáticos: Area, Priority, Company, Item Type, ROHS, Origin")
            _bullet("8. VOCÊ preenche manualmente: Assunto Principal e Manufacturer")
            _bullet("9. Clique em 'Confirmar Solicitação' no Phoenix")
            _bullet("10. O ticket é capturado automaticamente e salvo no Dashboard")

            _dica("O template Excel deve ter colunas 'Description' e 'MPN' na aba 'Items' ou 'Template'.")
            _aviso("NÃO feche o navegador antes de confirmar a solicitação! A ferramenta precisa detectar a mudança de URL para capturar o ticket.")

            _separador()

            _subtitulo("Capturar Phoenix (Sincronização)")
            _paragrafo(
                "O botão 'Capturar Phoenix' no Dashboard permite importar solicitações "
                "que foram abertas diretamente no Phoenix (sem a ferramenta) e que ainda "
                "não estão no Dashboard."
            )
            _bullet("Funciona 100% invisível (headless) — nenhuma janela abre")
            _bullet("Faz login automático no Phoenix")
            _bullet("Navega para 'Minhas Solicitações'")
            _bullet("Compara cada ticket com o banco local")
            _bullet("Insere automaticamente os tickets que não existem no Dashboard")

            _dica("Use esse botão sempre que abrir solicitações diretamente pelo Phoenix para manter o Dashboard sincronizado.")

        def _show_pegasus():
            _limpar()
            _titulo_secao("Módulo Pegasus", "📋")
            _paragrafo(
                "O módulo Pegasus permite registrar os itens no sistema Pegasus após "
                "a fase do Phoenix ser concluída."
            )

            _subtitulo("Fluxo do Pegasus")
            _diagrama_fluxo([
                ("✅", "Phoenix\nFechado", "#2e7d32"),
                ("▶️", "Iniciar\nPegasus", "#6a1b9a"),
                ("📋", "Preencher\nno Pegasus", "#7b1fa2"),
                ("⏹️", "Finalizar\nPegasus", "#4a148c"),
            ])

            _paragrafo("Como usar:")
            _bullet("1. No Dashboard, clique 'Iniciar Pegasus' no registro desejado")
            _bullet("2. A data de abertura do Pegasus é registrada automaticamente")
            _bullet("3. Após concluir o trabalho no Pegasus, clique 'Finalizar Pegasus'")
            _bullet("4. A data de fechamento é registrada e o registro avança para a fase de Custo")

            _dica("Você pode usar o módulo Pegasus do menu lateral para abrir o sistema Pegasus automaticamente com login preenchido.")

        def _show_custo():
            _limpar()
            _titulo_secao("Módulo Cost (Custo)", "💰")
            _paragrafo(
                "O módulo de Custo é a última fase do processo de classificação fiscal. "
                "Aqui é feita a análise fiscal e atribuição do Part Number final."
            )

            _subtitulo("Fluxo do Custo")
            _diagrama_fluxo([
                ("✅", "Pegasus\nFechado", "#4a148c"),
                ("▶️", "Iniciar\nCusto", "#e65100"),
                ("💰", "Análise\nFiscal", "#f57c00"),
                ("⏹️", "Finalizar\nCusto", "#bf360c"),
                ("🏁", "FINALIZADO", "#2e7d32"),
            ])

            _paragrafo("Como usar:")
            _bullet("1. No Dashboard, clique 'Iniciar Custo' após o Pegasus ser finalizado")
            _bullet("2. A data de abertura do Custo é registrada")
            _bullet("3. Realize a análise fiscal e atribua o Part Number")
            _bullet("4. Clique 'Finalizar Custo' para concluir a solicitação")
            _bullet("5. O registro muda para status FINALIZADO com a data de fechamento geral")

            _aviso("Após finalizar o custo, o registro é considerado concluído. Verifique se o Flex PN está correto antes de finalizar.")

        def _show_ciclo_vida():
            _limpar()
            _titulo_secao("Ciclo de Vida de um Registro", "🔄")
            _paragrafo(
                "Cada solicitação registrada no Dashboard passa por diversas fases. "
                "A ferramenta controla automaticamente as datas de cada transição."
            )

            _subtitulo("Fases e Datas")
            _tabela(
                ["Fase", "Campo Abertura", "Campo Fechamento", "Quando é preenchido"],
                [
                    ["Phoenix", "data_abertura + hora_abertura", "phoenix_fechamento", "Ao criar / Ao clicar 'Finalizar Phoenix'"],
                    ["Pegasus", "pegasus_abertura", "pegasus_fechamento", "Ao clicar 'Iniciar Pegasus' / 'Finalizar Pegasus'"],
                    ["Custo", "custo_abertura", "custo_fechamento", "Ao clicar 'Iniciar Custo' / 'Finalizar Custo'"],
                    ["Geral", "criado_em", "data_fechamento", "Ao criar / Ao finalizar todas as fases"],
                ]
            )

            _subtitulo("Status Possíveis")
            _tabela(
                ["Status", "Significado", "Cor no Dashboard"],
                [
                    ["ON GOING", "Solicitação em andamento (qualquer fase ativa)", "🔵 Azul"],
                    ["FINALIZADO", "Todas as fases concluídas com sucesso", "🟢 Verde"],
                    ["CANCELADO", "Registro excluído/cancelado pelo usuário", "🔴 Vermelho"],
                ]
            )

            _subtitulo("Diagrama Visual")
            _diagrama_fluxo([
                ("📝", "CRIAÇÃO", "#1565c0"),
                ("🦅", "PHOENIX", "#0277bd"),
                ("📋", "PEGASUS", "#6a1b9a"),
                ("💰", "CUSTO", "#e65100"),
                ("✅", "FINALIZADO", "#2e7d32"),
            ])

            _dica("Na exportação para Excel, cada fase tem suas próprias colunas de data de abertura e fechamento.")

        def _show_exportacao():
            _limpar()
            _titulo_secao("Exportação de Dados", "📤")
            _paragrafo(
                "A ferramenta oferece duas formas de exportação para planilhas Excel (.xlsx):"
            )

            _subtitulo("1. Exportar Dashboard")
            _paragrafo(
                "Gera um Excel com duas abas:"
            )
            _bullet("Aba 'Resumo': Total de Solicitações, Em Andamento, Finalizadas, Canceladas")
            _bullet("Aba 'Indicadores': Status com quantidade e percentual")

            _subtitulo("2. Exportar Registros")
            _paragrafo(
                "Gera um Excel completo com todos os registros e suas informações detalhadas:"
            )
            _tabela(
                ["Coluna", "Descrição"],
                [
                    ["Ticket", "Código da solicitação no Phoenix (ex: TKGP2026080521)"],
                    ["Description", "Descrição limpa do item"],
                    ["Produto", "Nome do produto"],
                    ["MPN", "Manufacturer Part Number"],
                    ["Flex PN", "Part Number atribuído pela Flex"],
                    ["Status", "EM ANDAMENTO / FINALIZADA / CANCELADA"],
                    ["Solicitante / Requisitante", "Quem solicitou a classificação"],
                    ["Data Abertura Phoenix", "Data + hora de criação da solicitação"],
                    ["Data Fechamento Phoenix", "Data em que a fase Phoenix foi finalizada"],
                    ["Data Abertura Pegasus", "Data de início da fase Pegasus"],
                    ["Data Fechamento Pegasus", "Data de encerramento do Pegasus"],
                    ["Data Abertura Custo", "Data de início da análise de custo"],
                    ["Data Fechamento Custo", "Data de conclusão do custo"],
                    ["Data Fechamento Geral", "Data final de conclusão de todo o processo"],
                    ["Criado Por / Criado Em", "Quem e quando criou o registro"],
                ]
            )

            _dica("Ao exportar registros, você pode escolher exportar TODOS ou apenas os registros do filtro/busca atual.")

        def _show_credenciais():
            _limpar()
            _titulo_secao("Credenciais", "🔑")
            _paragrafo(
                "O módulo de Credenciais permite salvar seus logins e senhas dos sistemas "
                "Phoenix, Pegasus e Cost de forma segura, protegidos por um PIN pessoal."
            )

            _subtitulo("Como funciona")
            _bullet("1. No primeiro acesso, defina um PIN de 4+ dígitos")
            _bullet("2. Cadastre login e senha para cada sistema (Phoenix, Pegasus, Cost)")
            _bullet("3. As credenciais são criptografadas com AES usando seu PIN")
            _bullet("4. Ao executar automações, basta digitar seu PIN para desbloquear")

            _subtitulo("Sistemas disponíveis")
            _tabela(
                ["Sistema", "Para quê serve"],
                [
                    ["Phoenix", "Login automático no Phoenix para novas solicitações"],
                    ["Pegasus", "Login automático no Pegasus"],
                    ["Cost", "Login automático no sistema de análise de custo"],
                ]
            )

            _aviso("Se você esquecer seu PIN, peça a um administrador para resetá-lo. As credenciais salvas serão perdidas e precisarão ser recadastradas.")
            _dica("Seu PIN nunca é salvo em texto puro — apenas o hash é armazenado. Nem mesmo administradores podem ver suas senhas.")

        def _show_admin():
            _limpar()
            _titulo_secao("Administração", "⚙️")
            _paragrafo(
                "O painel de Administração (acesso exclusivo para ADMIN) permite "
                "gerenciar usuários, backups e configurações do sistema."
            )

            _subtitulo("Gerenciamento de Usuários")
            _tabela(
                ["Ação", "Descrição"],
                [
                    ["Adicionar", "Cria um novo usuário com nome, código e nível de permissão"],
                    ["Editar", "Altera nome, código ou permissão de um usuário existente"],
                    ["Desativar", "Impede o login do usuário sem excluí-lo"],
                    ["Reativar", "Reativa um usuário desativado"],
                    ["Excluir", "Remove permanentemente o usuário do sistema"],
                ]
            )

            _subtitulo("Níveis de Permissão")
            _tabela(
                ["Nível", "Acesso"],
                [
                    ["ADMIN", "Acesso total: Dashboard, Phoenix, Pegasus, Credenciais, Administração, Exportação"],
                    ["ENGENHARIA", "Dashboard, Phoenix, Pegasus, Credenciais, Exportação (sem Administração)"],
                    ["USUARIO", "Apenas Dashboard, Consulta e Exportação (somente leitura)"],
                ]
            )

            _subtitulo("Backup do Banco de Dados")
            _bullet("Backups são criados automaticamente na inicialização")
            _bullet("Use 'Backup Manual' para criar um backup imediato")
            _bullet("Backups antigos são removidos automaticamente (mantém os 10 mais recentes)")
            _bullet("O banco pode ser relocado para outro caminho (rede, USB, etc.)")

            _dica("Exporte os usuários para JSON periodicamente como backup adicional. Isso permite importar em caso de perda do banco.")

        def _show_atalhos():
            _limpar()
            _titulo_secao("Atalhos e Dicas", "⚡")

            _subtitulo("Atalhos de Navegação")
            _tabela(
                ["Ação", "Como fazer"],
                [
                    ["Alternar tema (Claro/Escuro)", "Clique no ícone 🌙/☀️ no rodapé"],
                    ["Notificações", "Clique no ícone 🔔 no cabeçalho"],
                    ["Ajuda (esta tela)", "Clique no ícone ❓ no cabeçalho"],
                    ["Tela cheia", "Duplo clique na barra de título"],
                    ["Buscar no Dashboard", "Digite na barra de busca (canto superior direito do dashboard)"],
                ]
            )

            _subtitulo("Dicas de Produtividade")
            _bullet("Use 'Capturar Phoenix' para importar solicitações abertas diretamente no site")
            _bullet("Sempre salve suas credenciais para evitar digitar login/senha toda vez")
            _bullet("Exporte registros para Excel para compartilhar com sua equipe")
            _bullet("Use os filtros do dashboard para focar apenas nos registros relevantes")
            _bullet("O dashboard atualiza automaticamente — mas você pode clicar 'Atualizar' para forçar")

            _subtitulo("Solução de Problemas")
            _tabela(
                ["Problema", "Solução"],
                [
                    ["Navegador não abre", "Verifique se o Chromium do Playwright está instalado. Na primeira execução, ele é instalado automaticamente."],
                    ["Ticket não foi capturado", "Use 'Capturar Phoenix' no dashboard para sincronizar"],
                    ["Description mostra 'Muito Alta'", "A ferramenta limpa automaticamente essas descrições. Se persistir, edite o registro manualmente."],
                    ["Erro ao exportar", "Verifique se o arquivo Excel não está aberto em outro programa"],
                    ["Login falhou", "Recadastre suas credenciais no módulo Credenciais"],
                    ["Banco de dados corrompido", "Restaure um backup da pasta /backups"],
                ]
            )

        # ------------------------------------------------------------------
        # Montar sidebar com botões de seção
        # ------------------------------------------------------------------
        tk.Label(sidebar, text="SEÇÕES", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 8, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        secoes = [
            ("🏠  Visão Geral", _show_visao_geral),
            ("📊  Dashboard", _show_dashboard),
            ("🦅  Phoenix", _show_phoenix),
            ("📋  Pegasus", _show_pegasus),
            ("💰  Cost / Custo", _show_custo),
            ("🔄  Ciclo de Vida", _show_ciclo_vida),
            ("📤  Exportação", _show_exportacao),
            ("🔑  Credenciais", _show_credenciais),
            ("⚙️  Administração", _show_admin),
            ("⚡  Atalhos e Dicas", _show_atalhos),
        ]

        sidebar_btns = []

        def _selecionar(idx, fn):
            for i, b in enumerate(sidebar_btns):
                if i == idx:
                    b.configure(bg=ACCENT, fg="#ffffff")
                else:
                    b.configure(bg=BG_CARD, fg=TEXTO)
            fn()

        for i, (texto, fn) in enumerate(secoes):
            btn = tk.Button(
                sidebar,
                text=texto,
                bg=BG_CARD,
                fg=TEXTO,
                activebackground=ACCENT,
                activeforeground="#ffffff",
                font=("Arial", 9, "bold"),
                bd=0,
                highlightthickness=0,
                anchor="w",
                padx=16,
                pady=8,
                cursor="hand2",
                command=lambda idx=i, f=fn: _selecionar(idx, f),
            )
            btn.pack(fill="x")
            sidebar_btns.append(btn)

        # Footer da sidebar
        tk.Frame(sidebar, bg=BORDA, height=1).pack(fill="x", padx=12, pady=(12, 8), side="bottom")
        tk.Label(sidebar, text="v1.0 • Gabriel Girotto", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 7)).pack(side="bottom", pady=(0, 8))

        # Mostrar primeira seção por padrão
        _selecionar(0, _show_visao_geral)

    def _executar_cost_manual(self):
        """Executa a automação padrão do Cost Request (Abertura Manual)."""
        cred = self._obter_ou_pedir_credencial("Cost")
        if not cred or not cred.get("login") or not cred.get("senha"):
            return

        cred_user = cred["login"]
        cred_pass = cred["senha"]

        def _worker():
            try:
                from automocoes.cost.cost import nova_solicitacao_cost
                os.environ["PHOENIX_CRED_USER"] = cred_user
                os.environ["PHOENIX_CRED_PASS"] = cred_pass
                nova_solicitacao_cost()
            except Exception as exc:
                logger.error("Erro ao executar Cost Request manual: %s", exc)

        import threading
        threading.Thread(target=_worker, daemon=True).start()

    def _abrir_modal_cost_template(self):
        """Abre o modal corporativo para envio de Cost Request via Template Excel."""
        modal = tk.Toplevel(self.root)
        modal.title("Open Request By Template - Cost Request")
        modal.geometry("640x560")
        modal.configure(bg=BG_CARD)
        modal.transient(self.root)
        modal.grab_set()

        # Título do modal
        header_frame = tk.Frame(modal, bg=BG_CARD, padx=20, pady=16)
        header_frame.pack(fill="x")

        tk.Label(
            header_frame,
            text="Open Request By Template",
            bg=BG_CARD, fg=TEXTO, font=("Arial", 14, "bold")
        ).pack(anchor="w")

        tk.Label(
            header_frame,
            text="Selecione o Site/Company e anexe os arquivos de Template (.xlsx) e Cotação.",
            bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 9)
        ).pack(anchor="w", pady=(2, 0))

        tk.Frame(modal, bg=BORDA, height=1).pack(fill="x", padx=20)

        body = tk.Frame(modal, bg=BG_CARD, padx=20, pady=16)
        body.pack(fill="both", expand=True)

        # Combo Site
        tk.Label(body, text="Site *", bg=BG_CARD, fg=TEXTO, font=("Arial", 9, "bold")).pack(anchor="w", pady=(4, 2))
        combo_site = ttk.Combobox(body, values=["JAGUARIUNA", "MANAUS", "SOROCABA", "OUTRO"], state="readonly")
        combo_site.set("JAGUARIUNA")
        combo_site.pack(fill="x", pady=(0, 10))

        # Combo Company
        tk.Label(body, text="Company *", bg=BG_CARD, fg=TEXTO, font=("Arial", 9, "bold")).pack(anchor="w", pady=(4, 2))
        combo_company = ttk.Combobox(body, values=["361 - SITE JAGUARIUNA", "362 - SITE MANAUS", "OUTRO"], state="readonly")
        combo_company.set("361 - SITE JAGUARIUNA")
        combo_company.pack(fill="x", pady=(0, 10))

        # Arquivo Template (.xlsx)
        tk.Label(body, text="Arquivo Template (.xlsx) *", bg=BG_CARD, fg=TEXTO, font=("Arial", 9, "bold")).pack(anchor="w", pady=(4, 2))
        row_temp = tk.Frame(body, bg=BG_CARD)
        row_temp.pack(fill="x", pady=(0, 10))
        entry_template = tk.Entry(row_temp, bg=ENTRY_BG, fg=TEXTO, relief="solid", bd=1, font=("Arial", 9))
        entry_template.pack(side="left", fill="x", expand=True, padx=(0, 6))

        def _selecionar_template():
            caminho = filedialog.askopenfilename(
                title="Selecionar Template Cost Request (.xlsx)",
                filetypes=[("Planilha Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
            )
            if caminho:
                entry_template.delete(0, "end")
                entry_template.insert(0, caminho)

        tk.Button(
            row_temp, text="Buscar .xlsx", command=_selecionar_template,
            bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HOVER, activeforeground="#ffffff",
            relief="flat", bd=0, font=("Arial", 8, "bold"), padx=10, cursor="hand2"
        ).pack(side="right")

        # Arquivo Quotation (PDF)
        tk.Label(body, text="Arquivo de Cotação (PDF / E-mail)", bg=BG_CARD, fg=TEXTO, font=("Arial", 9, "bold")).pack(anchor="w", pady=(4, 2))
        row_quot = tk.Frame(body, bg=BG_CARD)
        row_quot.pack(fill="x", pady=(0, 16))
        entry_quotation = tk.Entry(row_quot, bg=ENTRY_BG, fg=TEXTO, relief="solid", bd=1, font=("Arial", 9))
        entry_quotation.pack(side="left", fill="x", expand=True, padx=(0, 6))

        def _selecionar_cotacao():
            caminho = filedialog.askopenfilename(
                title="Selecionar Arquivo de Cotação",
                filetypes=[("Documentos PDF / E-mail", "*.pdf;*.msg;*.eml"), ("Todos os arquivos", "*.*")]
            )
            if caminho:
                entry_quotation.delete(0, "end")
                entry_quotation.insert(0, caminho)

        tk.Button(
            row_quot, text="Buscar Cotação", command=_selecionar_cotacao,
            bg=BG_CARD, fg=TEXTO, activebackground=BORDA, activeforeground=TEXTO,
            relief="solid", bd=1, font=("Arial", 8, "bold"), padx=10, cursor="hand2"
        ).pack(side="right")

        # Barra de Progresso e Status
        lbl_status = tk.Label(body, text="", bg=BG_CARD, fg=TEXTO_MUTED, font=("Arial", 9))
        lbl_status.pack(anchor="w", pady=(0, 4))

        progress = ttk.Progressbar(body, mode="indeterminate")

        # Ações
        footer = tk.Frame(modal, bg=BG_CARD, padx=20, pady=16)
        footer.pack(fill="x", side="bottom")

        def _download_template():
            caminho_salvar = filedialog.asksaveasfilename(
                title="Salvar Template Oficial Cost Request",
                defaultextension=".xlsx",
                initialfile="CostRequest_Template_Oficial.xlsx",
                filetypes=[("Planilha Excel", "*.xlsx")]
            )
            if caminho_salvar:
                try:
                    from services.cost_template_service import gerar_template_oficial_excel
                    gerar_template_oficial_excel(caminho_salvar)
                    messagebox.showinfo("Download Concluído", f"Template gerado com sucesso em:\n{caminho_salvar}")
                except Exception as exc:
                    messagebox.showerror("Erro de Download", f"Não foi possível salvar o template:\n{exc}")

        btn_download = tk.Button(
            footer, text="Download Excel Template", command=_download_template,
            bg=BG_CARD, fg=ACCENT, activebackground=BORDA, activeforeground=TEXTO,
            relief="solid", bd=1, font=("Arial", 9, "bold"), padx=12, pady=6, cursor="hand2"
        )
        btn_download.pack(side="left")

        def _processar_e_enviar():
            site_val = combo_site.get().strip()
            company_val = combo_company.get().strip()
            temp_path = entry_template.get().strip()
            quot_path = entry_quotation.get().strip()

            if not site_val:
                messagebox.showwarning("Campo Obrigatório", "Por favor, selecione o Site.")
                return
            if not company_val:
                messagebox.showwarning("Campo Obrigatório", "Por favor, selecione a Company.")
                return
            if not temp_path:
                messagebox.showwarning("Campo Obrigatório", "Por favor, anexe o arquivo de Template (.xlsx).")
                return

            from services.cost_template_service import validar_template_cost_request
            valido, msg_val, registros_validos = validar_template_cost_request(temp_path)
            if not valido:
                messagebox.showerror("Validação do Template", msg_val)
                return

            lbl_status.config(text="Iniciando automação no portal...", fg=ACCENT)
            progress.pack(fill="x", pady=(0, 10))
            progress.start(10)
            btn_processar.config(state="disabled")
            btn_download.config(state="disabled")

            def _worker():
                try:
                    from automocoes.cost.cost_template import executar_abertura_by_template
                    from services.database import salvar_solicitacao_cost_template

                    cred_usr = getattr(self, "usuario_atual", {}).get("nome", "") or "JAGGGIRO"
                    cred_pass = ""

                    res = executar_abertura_by_template(
                        user=cred_usr,
                        password=cred_pass,
                        site=site_val,
                        company=company_val,
                        template_path=temp_path,
                        quotation_path=quot_path,
                        headless=False
                    )

                    def _on_finish():
                        progress.stop()
                        progress.pack_forget()
                        btn_processar.config(state="normal")
                        btn_download.config(state="normal")

                        if res.get("sucesso"):
                            tkt = res.get("ticket") or "TEMPLATE-AUT"
                            salvar_solicitacao_cost_template(
                                ticket=tkt,
                                solicitante="Solicitante Cost Request",
                                classificador=self.usuario_atual.get("nome", "Usuário") if hasattr(self, "usuario_atual") and isinstance(self.usuario_atual, dict) else "Usuário",
                                site=site_val,
                                company=company_val,
                                registros_items=registros_validos
                            )
                            messagebox.showinfo("Sucesso", f"Cost Request aberto com sucesso!\nTicket: {tkt}")
                            modal.destroy()
                            self._abrir_dashboard()
                        else:
                            lbl_status.config(text="Falha na automação.", fg="#ff6b6b")
                            messagebox.showerror("Erro na Automação", res.get("mensagem", "Falha desconhecida."))

                    self.root.after(0, _on_finish)
                except Exception as _exc:
                    def _on_err():
                        progress.stop()
                        progress.pack_forget()
                        btn_processar.config(state="normal")
                        btn_download.config(state="normal")
                        lbl_status.config(text="Erro no processamento.", fg="#ff6b6b")
                        messagebox.showerror("Erro Inesperado", f"Ocorreu um erro no processamento:\n{_exc}")
                    self.root.after(0, _on_err)

            import threading
            threading.Thread(target=_worker, daemon=True).start()

        btn_processar = tk.Button(
            footer, text="Processar & Enviar", command=_processar_e_enviar,
            bg="#27ae60", fg="#ffffff", activebackground="#219653", activeforeground="#ffffff",
            relief="flat", bd=0, font=("Arial", 9, "bold"), padx=14, pady=6, cursor="hand2"
        )
        btn_processar.pack(side="right")


def _tratar_excecao_global(exc_type, exc_value, exc_traceback):
    """Tratamento global de exceções não capturadas para evitar fechamento abrupoto do executável corporativo."""
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    logger.critical("Exceção não tratada capturada globalmente:", exc_info=(exc_type, exc_value, exc_traceback))
    try:
        messagebox.showerror(
            "Erro Inesperado no Sistema",
            f"Ocorreu um erro não esperado na aplicação:\n\n{exc_value}\n\n"
            "O detalhe do erro foi gravado no log corporativo em /logs."
        )
    except Exception:
        pass


if __name__ == "__main__":
    # Handler global para exceções do Python e do Tkinter
    sys.excepthook = _tratar_excecao_global

    root = tk.Tk()
    root.report_callback_exception = lambda et, ev, tb: _tratar_excecao_global(et, ev, tb)

   
    try:
        fazer_backup("startup_main")
    except Exception as _b_err:
        logger.warning("Não foi possível criar backup inicial: %s", _b_err)

    app = PhoenixTool(root)

    _primeiro_admin = obter_e_limpar_primeiro_admin_gerado()
    if _primeiro_admin:
        root.after(700, lambda: messagebox.showinfo(
            "Primeiro acesso ADMIN gerado",
            "Nenhum usuário existia neste banco de dados. Foi criada automaticamente "
            "uma conta ADMIN:\n\n"
            f"Nome Completo: {_primeiro_admin['nome']}\n"
            f"Código Engenharia: {_primeiro_admin['codigo']}\n\n"
            "Guarde este código agora em local seguro. Ele NÃO será exibido novamente."
        ))

    root.mainloop()
