
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from playwright.sync_api import Page, sync_playwright

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from services.logging_config import configurar_logger  # noqa: E402
from services.storage import carregar_login  # noqa: E402

logger = configurar_logger("planilha")

PLANILHA_URL = "https://exemplo365.sharepoint.com/:x:/r/sites/TestEngineeringExemplo/_layouts/15/Doc.aspx?sourcedoc=%7B00000000-0000-0000-0000-000000000000%7D&action=edit"
DEFAULT_PAYLOAD_PATH = os.path.join(BASE_DIR, "planilha_payload.json")
CONFIG_PATH = os.path.join(BASE_DIR, "..", "..", "config_tool.json")
ONE_DRIVE_ENV_KEYS = ("OneDrive", "OneDriveCommercial", "OneDriveConsumer")
WORKBOOK_SEARCH_PATTERNS = ("planilha", "phoenix", "exemplo", "sharepoint", "solicita")
WORKBOOK_SEARCH_EXTENSIONS = (".xlsx",)
WORKBOOK_SEARCH_MAX_DEPTH = 4


def construir_payload_planilha(descricao: str, linha: Optional[str] = None, usuario: Optional[str] = None) -> Dict[str, str]:
    data_open = datetime.now().strftime("%d/%m")
    return {
        "descricao": descricao or "",
        "status": "ON GOING",
        "data_open": data_open,
        "linha": str(linha or ""),
        "usuario": usuario or "",
    }


def salvar_payload_planilha(payload: Dict[str, Any], arquivo: str = DEFAULT_PAYLOAD_PATH) -> None:
    with open(arquivo, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def carregar_payload_planilha(arquivo: str = DEFAULT_PAYLOAD_PATH) -> Optional[Dict[str, Any]]:
    if not os.path.exists(arquivo):
        return None
    with open(arquivo, "r", encoding="utf-8") as f:
        return json.load(f)


def formatar_dados_planilha(payload: Dict[str, Any]) -> List[str]:
    return [
        str(payload.get("descricao") or ""),
        str(payload.get("status") or "ON GOING"),
        str(payload.get("data_open") or datetime.now().strftime("%d/%m")),
        str(payload.get("linha") or ""),
        str(payload.get("usuario") or ""),
    ]


def carregar_config() -> Dict[str, Any]:
    """Cache local extra (ex.: caminho do workbook do OneDrive), no mesmo
    arquivo usado por `services.storage` para usuário/senha."""
    if not os.path.exists(CONFIG_PATH):
        return {}
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}


def salvar_config(config: Dict[str, Any], arquivo: str = CONFIG_PATH) -> None:
    try:
        with open(arquivo, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except OSError:
        logger.warning("Não foi possível salvar configuração em %s", arquivo)


def detectar_pastas_onedrive():
    pastas = []
    for chave in ONE_DRIVE_ENV_KEYS:
        caminho = os.environ.get(chave)
        if caminho and os.path.isdir(caminho):
            pastas.append(caminho)
    return pastas


def _buscar_arquivo_planilha_local(roots=None, patterns=None, max_depth=WORKBOOK_SEARCH_MAX_DEPTH):
    roots = roots or detectar_pastas_onedrive()
    if not roots:
        return None

    patterns = patterns or WORKBOOK_SEARCH_PATTERNS
    for root in roots:
        for dirpath, dirnames, filenames in os.walk(root):
            rel = os.path.relpath(dirpath, root)
            depth = 0 if rel == "." else rel.count(os.sep)
            if depth > max_depth:
                dirnames[:] = []
                continue
            for filename in filenames:
                lower_name = filename.lower()
                if not lower_name.endswith(WORKBOOK_SEARCH_EXTENSIONS):
                    continue
                if any(pattern in lower_name for pattern in patterns):
                    return os.path.join(dirpath, filename)
    return None


def obter_caminho_planilha_local():
    config = carregar_config()
    caminho = config.get("workbook_path")
    if caminho and os.path.isfile(caminho):
        return caminho

    caminho = _buscar_arquivo_planilha_local()
    if caminho:
        config["workbook_path"] = caminho
        salvar_config(config)
        return caminho

    return None


def atualizar_planilha_local(payload, workbook_path=None, sheet_name=None):
    workbook_path = workbook_path or obter_caminho_planilha_local()
    if not workbook_path:
        raise FileNotFoundError("Nenhum arquivo local do Excel foi encontrado no OneDrive.")

    try:
        workbook = load_workbook(workbook_path)
    except InvalidFileException as error:
        raise ValueError(f"Arquivo não é um workbook válido: {workbook_path}") from error
    except Exception as error:
        raise IOError(f"Não foi possível abrir o workbook: {workbook_path}") from error

    sheet = workbook.active
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

    valores = formatar_dados_planilha(payload)
    linha_busca = str(payload.get("linha") or "").strip()
    atualizado = False

    if linha_busca:
        for row in sheet.iter_rows(min_row=1, max_col=len(valores)):
            if len(row) < 4:
                continue
            celula = row[3]
            if celula.value is None:
                continue
            if str(celula.value).strip() == linha_busca:
                for index, valor in enumerate(valores, start=1):
                    sheet.cell(row=celula.row, column=index, value=valor)
                atualizado = True
                break

    if not atualizado:
        sheet.append(valores)

    workbook.save(workbook_path)
    return workbook_path


def _tentar_login_page(page, config):
    try:
        if page.locator("#FlexUser").count() > 0:
            page.fill("#FlexUser", config.get("user", ""))
            page.fill("#Password", config.get("password", ""))
            for selector in [
                'xpath=//*[@id="formLogin"]/button',
                'xpath=//*[@id="formLogin"]/button/span',
            ]:
                if page.locator(selector).count() > 0:
                    page.click(selector)
                    break
            return True
    except Exception:
        return False
    return False


def _tentar_preencher_planilha(page, payload):
    valores = formatar_dados_planilha(payload)
    if not any(valores):
        return False

    try:
        page.wait_for_load_state("domcontentloaded", timeout=30000)
        page.wait_for_timeout(3000)
    except Exception:
        return False

    alvo = None
    candidatos = [
        'div[role="gridcell"]',
        'div[role="cell"]',
        'div[aria-label*="cell"]',
        '[contenteditable="true"]',
        'input',
    ]

    for selector in candidatos:
        try:
            locators = page.locator(selector)
            if locators.count() == 0:
                continue
            alvo = locators.first
            alvo.click()
            break
        except Exception:
            continue

    if alvo is None:
        return False

    try:
        page.keyboard.press("End")
        page.wait_for_timeout(500)
    except Exception:
        pass

    for valor in valores:
        try:
            page.keyboard.type(str(valor))
            page.keyboard.press("Tab")
            page.wait_for_timeout(400)
        except Exception:
            break

    try:
        texto = page.evaluate("() => document.body.innerText")
        return bool(texto and any(str(v) in texto for v in valores if str(v)))
    except Exception:
        return True


def enviar_para_planilha(
    descricao: str,
    linha: Optional[str] = None,
    usuario: Optional[str] = None,
    arquivo: str = DEFAULT_PAYLOAD_PATH,
    workbook_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> Optional[str]:
    payload = construir_payload_planilha(descricao, linha=linha, usuario=usuario)
    salvar_payload_planilha(payload, arquivo=arquivo)

    try:
        caminho_atualizado = atualizar_planilha_local(payload, workbook_path=workbook_path, sheet_name=sheet_name)
        logger.info("Planilha local atualizada em: %s", caminho_atualizado)
        return caminho_atualizado
    except FileNotFoundError:
        logger.info("Nenhum arquivo local do OneDrive encontrado. Abrindo planilha online...")
    except Exception as exc:
        logger.error("Falha ao atualizar planilha local: %s", exc)
        logger.info("Abrindo planilha online como fallback...")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            channel="msedge",
            headless=False,
            args=["--start-maximized"]
        )

        context = browser.new_context(viewport=None)
        page = context.new_page()

        logger.info("Abrindo planilha...")
        page.goto(PLANILHA_URL, wait_until="networkidle")
        logger.info("Planilha aberta.")

        config = carregar_login()
        _tentar_login_page(page, config)

        logger.info("=" * 50)
        logger.info("PAYLOAD: %s", json.dumps(payload, ensure_ascii=False, indent=2))
        logger.info("=" * 50)

        if _tentar_preencher_planilha(page, payload):
            logger.info("Tentativa de preenchimento iniciada na planilha.")
        else:
            logger.warning(
                "Não foi possível localizar uma célula editável na planilha; o payload foi "
                "salvo e a página foi aberta para revisão manual."
            )

        input("Pressione ENTER para fechar...")
        browser.close()


if __name__ == "__main__":
    descricao_cli = " ".join(sys.argv[1:]).strip() if len(sys.argv) > 1 else ""
    if not descricao_cli:
        logger.error("Uso: python planilha.py <descricao>")
        sys.exit(1)
    enviar_para_planilha(descricao_cli)