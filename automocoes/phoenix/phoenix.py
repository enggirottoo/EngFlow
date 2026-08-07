"""Automação Phoenix (Portal + Phoenix).

Duas entradas são usadas pelo Phoenix Tool (via `subprocess`, veja
`services/process_runner.py`):

    python phoenix.py home   -> abrir_home_phoenix()
    python phoenix.py        -> nova_solicitacao_phoenix()

O fluxo de login (`_preencher_login`) e a lógica de detectar se o clique em
um link/tile abriu uma nova aba (`_abrir_phoenix`) são compartilhados com
`pegasus.py`; qualquer ajuste de seletor de login deve ser espelhado lá
também caso o Portal mude.
"""

from __future__ import annotations

import os
import re
import sys
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Any, Dict, Optional, Tuple

from playwright.sync_api import Browser, BrowserContext, Page, Playwright
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from services.logging_config import configurar_logger  # noqa: E402
from services.storage import (  # noqa: E402
    criar_registro_descricao,
    existe_ticket,
)

# Grava também em arquivo (phoenix_tool.log): a automação roda num console
# separado (CREATE_NEW_CONSOLE) que pode ser fechado sem que o usuário veja
# o erro - sem log em arquivo, uma falha nessa janela some para sempre e o
# registro no histórico/dashboard nunca é diagnosticado.
logger = configurar_logger("phoenix")


def _parse_credenciais() -> Dict[str, str]:
    """Lê credenciais das variáveis de ambiente ou argumentos CLI."""
    user = os.environ.get("PHOENIX_CRED_USER", "")
    password = os.environ.get("PHOENIX_CRED_PASS", "")
    args = sys.argv[1:]
    for i, arg in enumerate(args):
        if arg == "--cred-user" and i + 1 < len(args):
            user = user or args[i + 1]
        elif arg == "--cred-pass" and i + 1 < len(args):
            password = password or args[i + 1]
    return {"user": user, "password": password}


def carregar_config() -> Dict[str, Any]:
    """Credenciais vêm dos argumentos de linha de comando passados pelo runner."""
    return _parse_credenciais()


def _esperar_elemento(page: Page, selector: str, timeout: int = 10000) -> bool:
    try:
        page.wait_for_selector(selector, timeout=timeout)
        return True
    except PlaywrightTimeoutError:
        return False


def _debug_screenshot(page: Page, nome: str) -> None:
    """Salva um print da tela para ajudar a diagnosticar problema de seletor/login."""
    try:
        pasta = os.path.join(BASE_DIR, "debug")
        os.makedirs(pasta, exist_ok=True)
        caminho = os.path.join(pasta, f"{nome}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        page.screenshot(path=caminho, full_page=True)
        logger.info("Screenshot salvo em: %s", caminho)
    except Exception as exc:
        logger.warning("Não foi possível salvar screenshot de debug: %s", exc)


def _diagnostico(page: Page, label: str) -> None:
    """Loga URL e título atuais da página para ajudar a diagnosticar sem precisar de screenshot."""
    try:
        logger.info("[%s] URL atual: %s", label, page.url)
        logger.info("[%s] Título: %s", label, page.title())
    except Exception as exc:
        logger.warning("[%s] Não consegui ler URL/título: %s", label, exc)


def _preencher_login(page: Page, config: Dict[str, Any], timeout: int = 15000) -> bool:
    if not _esperar_elemento(page, "#FlexUser", timeout=timeout):
        logger.info(
            "Campo de login (#FlexUser) não apareceu em %sms. "
            "Ou a página já está logada, ou o seletor mudou.",
            timeout,
        )
        _diagnostico(page, "sem-campo-login")
        _debug_screenshot(page, "login_campo_nao_encontrado")
        return False

    if not (config.get("user") or "").strip() or not (config.get("password") or "").strip():
        logger.warning(
            "Usuário/senha configurados estão vazios. Nada será digitado nos campos de login."
        )

    url_antes = page.url
    page.fill("#FlexUser", config.get("user", ""))
    page.fill("#Password", config.get("password", ""))

    clicou = False
    for selector in [
        'xpath=//*[@id="formLogin"]/button',
        'xpath=//*[@id="formLogin"]/button/span',
        'button[type="submit"]',
        'text=Entrar',
        'text=Login',
    ]:
        if _esperar_elemento(page, selector, timeout=3000):
            try:
                page.click(selector)
                clicou = True
                break
            except Exception:
                continue

    if not clicou:
        logger.info("Campos preenchidos, mas não achei o botão de login para clicar. Tentando ENTER.")
        page.keyboard.press("Enter")

    try:
        page.wait_for_url(lambda url: url != url_antes, timeout=8000)
    except PlaywrightTimeoutError:
        pass
    try:
        page.wait_for_load_state("domcontentloaded", timeout=5000)
    except Exception:
        pass

    try:
        ainda_com_login = page.locator("#FlexUser").count() > 0
    except Exception:
        ainda_com_login = False

    if ainda_com_login:
        logger.warning(
            "Depois de enviar o formulário, o campo de usuário ainda está na tela. "
            "O login provavelmente FALHOU (confira o usuário/senha salvos)."
        )
        _diagnostico(page, "login-falhou")
        _debug_screenshot(page, "login_falhou")
        return False

    logger.info("Login enviado e formulário de login não aparece mais na tela.")
    _diagnostico(page, "login-ok")
    return True


def _abrir_phoenix(
    playwright: Playwright, config: Dict[str, Any], headless: bool = False
) -> Tuple[Browser, BrowserContext, Page]:
    """Login no Portal + login no Phoenix. Retorna (browser, context, page) já logado no Phoenix.

    `headless=True` é usado por automações que rodam em segundo plano sem
    mostrar nenhuma janela ao usuário (ex.: atualizar_pn.buscar_pn_por_ticket)."""

    browser = playwright.chromium.launch(
        channel="msedge",
        headless=headless,
        args=["--start-maximized"] if not headless else [],
    )

    context = browser.new_context(viewport=None if not headless else {"width": 1366, "height": 768})
    page = context.new_page()

    # PORTAL
    page.goto("http://engenharia.br.flex.com/", wait_until="domcontentloaded")
    if _esperar_elemento(page, "#FlexUser", timeout=20000):
        _preencher_login(page, config, timeout=20000)
    else:
        logger.info("Portal: formulário de login não apareceu (provavelmente já autenticado).")
    _diagnostico(page, "portal")
    logger.info("Portal OK")

    # ABRIR PHOENIX
    url_antes_click = page.url
    seletor_disponivel: Optional[str] = None
    for selector in [
        'xpath=//*[@id="solution"]/div/div/div[2]/div/div[3]/a',
        'a[href*="Phoenix"]',
        'text=Phoenix',
        'text=PHOENIX',
    ]:
        if _esperar_elemento(page, selector, timeout=5000):
            seletor_disponivel = selector
            break

    abriu_phoenix = False
    nova_pagina = None
    if seletor_disponivel:
        try:
            with context.expect_page(timeout=6000) as info_nova_pagina:
                page.click(seletor_disponivel)
            nova_pagina = info_nova_pagina.value
            abriu_phoenix = True
        except PlaywrightTimeoutError:
            abriu_phoenix = True
        except Exception as exc:
            logger.error("Erro ao clicar no link do Phoenix: %s", exc)
    else:
        logger.warning(
            "Não encontrei nenhum link/botão para abrir o Phoenix na Home do Portal. "
            "O fluxo pode continuar preso na tela do Portal."
        )
        _debug_screenshot(page, "phoenix_link_nao_encontrado")

    if nova_pagina is not None:
        try:
            nova_pagina.wait_for_load_state("domcontentloaded")
        except Exception:
            pass
        page = nova_pagina
        logger.info("Phoenix abriu em uma nova aba.")
    else:
        try:
            page.wait_for_load_state("domcontentloaded")
        except Exception:
            pass
        if abriu_phoenix and page.url == url_antes_click:
            logger.warning(
                "Cliquei no link do Phoenix, mas a URL não mudou (%s). O clique "
                "provavelmente não teve efeito - verifique se o seletor ainda é o correto.",
                page.url,
            )
            _debug_screenshot(page, "phoenix_url_nao_mudou")

    _diagnostico(page, "phoenix-aberto")
    logger.info("Phoenix OK")

    # LOGIN PHOENIX
    try:
        if _esperar_elemento(page, "#FlexUser", timeout=10000):
            _preencher_login(page, config, timeout=10000)
        else:
            logger.info("Phoenix: campo de login não apareceu (já estava logado ou seletor mudou).")
            _diagnostico(page, "phoenix-sem-campo-login")
    except Exception as exc:
        logger.error("Erro ao tentar logar no Phoenix: %s", exc)
        _diagnostico(page, "phoenix-erro-login")

    return browser, context, page


# =====================================================
# FUNÇÃO 1: só abrir a home do Phoenix (botão "Home Phoenix")
# =====================================================

def abrir_home_phoenix() -> None:
    config = carregar_config()

    with sync_playwright() as p:
        browser, context, page = _abrir_phoenix(p, config)

        logger.info("Home Phoenix aberta.")
        input("Pressione ENTER para fechar...")

        browser.close()


# =====================================================
# FUNÇÃO 2: fluxo completo de solicitação P0032 (botão "Nova Solicitação")
# =====================================================

def nova_solicitacao_phoenix() -> None:
    config = carregar_config()

    # Produto/Solicitante/Requisitante não são mais pedidos aqui: o usuário
    # preenche esses campos depois, editando o registro já criado no
    # dashboard (botão "Editar"). Isso evita travar o início da automação
    # esperando um popup da ferramenta.
    with sync_playwright() as p:
        browser, context, page = _abrir_phoenix(p, config)
        try:
            _fluxo_nova_solicitacao(page, browser, config)
        except Exception:
            logger.exception(
                "Automação do Phoenix falhou antes de registrar no histórico/dashboard."
            )
            _debug_screenshot(page, "phoenix_falha_inesperada")
            try:
                messagebox.showerror(
                    "Erro na automação do Phoenix",
                    "A automação falhou e a solicitação NÃO foi registrada no dashboard.\n"
                    "Veja phoenix_tool.log para detalhes.",
                )
            except Exception:
                pass
        finally:
            try:
                browser.close()
            except Exception:
                pass


def ler_itens_template_excel(caminho_xlsx: str) -> list[Dict[str, str]]:
    """Lê a planilha template Excel selecionada e extrai descrições e MPNs de cada item."""
    itens = []
    if not caminho_xlsx or not os.path.exists(caminho_xlsx):
        return itens

    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
        sheet = wb.active

        header_desc_col = None
        header_mpn_col = None
        header_row = 1

        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row_vals = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
            for col_idx, val in enumerate(row_vals, start=1):
                val_str = str(val or "").strip().lower()
                if not header_desc_col and ("descrip" in val_str or "descriç" in val_str or "descric" in val_str or val_str == "desc"):
                    header_desc_col = col_idx
                if not header_mpn_col and ("mpn" in val_str or "part number" in val_str or "part_number" in val_str or val_str in ("pn", "p/n")):
                    header_mpn_col = col_idx

            if header_desc_col:
                header_row = row_idx
                break

        if not header_desc_col:
            header_desc_col = 1

        start_row = header_row + 1
        for r in range(start_row, sheet.max_row + 1):
            desc_val = str(sheet.cell(row=r, column=header_desc_col).value or "").strip() if header_desc_col else ""
            mpn_val = str(sheet.cell(row=r, column=header_mpn_col).value or "").strip() if header_mpn_col else ""

            if desc_val and desc_val.lower() not in ("description", "descrição", "descrição do item", "desc"):
                itens.append({"description": desc_val, "mpn": mpn_val})

        wb.close()
        logger.info("Itens lidos diretamente da planilha Excel (%s item(ns)): %s", len(itens), itens)
    except Exception as exc:
        logger.error("Erro ao ler planilha Excel diretamente: %s", exc)

    return itens


def _capturar_itens_do_formulario_phoenix(page: Page) -> list[Dict[str, str]]:
    """Captura a lista de descrições e MPNs renderizados no formulário do Phoenix logo após o clique na lupa (conforme a 2ª imagem)."""
    itens = []

    try:
        page.wait_for_selector('table tbody tr, input[name*="Description"], input[id*="Description"], [id$="__Description"]', timeout=15000)
    except PlaywrightTimeoutError:
        logger.warning("Timeout aguardando formulário de produtos após a Lupa.")

    # 1. Busca por atributos name e id típicos de formulários do Phoenix
    selectors_desc = [
        '[id$="__Description"]',
        'input[name*="Description"]',
        'input[id*="Description"]',
        'input[name*="description"]',
        'input[id*="description"]',
    ]
    selectors_mpn = [
        '[id$="__Mpn"]',
        'input[name*="Mpn"]',
        'input[id*="Mpn"]',
        'input[name*="mpn"]',
        'input[id*="mpn"]',
    ]

    for s_desc in selectors_desc:
        try:
            loc_desc = page.locator(s_desc)
            count = loc_desc.count()
            if count > 0:
                for idx in range(count):
                    val = loc_desc.nth(idx).input_value().strip()
                    if val:
                        mpn_val = ""
                        for s_mpn in selectors_mpn:
                            loc_mpn = page.locator(s_mpn)
                            if loc_mpn.count() > idx:
                                mpn_val = loc_mpn.nth(idx).input_value().strip()
                                if mpn_val:
                                    break
                        itens.append({"description": val, "mpn": mpn_val})
                if itens:
                    logger.info("Itens capturados do formulário via selector '%s' (%s item(ns)): %s", s_desc, len(itens), itens)
                    return itens
        except Exception:
            pass

    # 2. Varredura célula a célula por posição da tabela (conforme a 2ª imagem: Col 2=MPN, Col 3=Description)
    try:
        rows = page.locator("table tbody tr")
        row_count = rows.count()
        for i in range(row_count):
            r = rows.nth(i)
            tds = r.locator("td")
            td_count = tds.count()
            if td_count >= 3:
                desc_input = tds.nth(2).locator("input, textarea")
                mpn_input = tds.nth(1).locator("input, textarea")

                desc_val = desc_input.nth(0).input_value().strip() if desc_input.count() > 0 else tds.nth(2).inner_text().strip()
                mpn_val = mpn_input.nth(0).input_value().strip() if mpn_input.count() > 0 else tds.nth(1).inner_text().strip()

                if desc_val and desc_val.lower() not in ("description", "descrição", "desc"):
                    itens.append({"description": desc_val, "mpn": mpn_val})

        if itens:
            logger.info("Itens capturados do formulário via tabela da 2ª imagem (%s item(ns)): %s", len(itens), itens)
            return itens
    except Exception as exc:
        logger.error("Erro ao varrer tabela de produtos do formulário: %s", exc)

    return itens


def _fluxo_nova_solicitacao(
    page: Page,
    browser: Browser,
    config: Dict[str, Any],
) -> None:
    # P0032
    logger.info("Abrindo P0032...")
    page.fill("#TaskName", "32")
    page.click("text=P0032")
    page.click('xpath=//*[@id="formId"]/div/div[2]/button')
    page.wait_for_load_state("domcontentloaded")
    logger.info("P0032 OK")

    # ESCOLHER TEMPLATE
    logger.info("Abrindo janela para selecionar o template (.xlsx)...")
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.update()

    arquivo = filedialog.askopenfilename(
        parent=root,
        title="Selecione o Template",
        filetypes=[("Excel", "*.xlsx")]
    )

    root.destroy()

    if not arquivo:
        logger.info("Nenhum arquivo selecionado.")
        return

    logger.info("Arquivo selecionado: %s", arquivo)

    # Ler itens diretamente da planilha Excel selecionada
    itens_excel = ler_itens_template_excel(arquivo)

    # UPLOAD TEMPLATE
    page.set_input_files("#ExcelFile", arquivo)
    logger.info("Template carregado no Phoenix")

    # LUPA
    page.click('xpath=//*[@id="uploadForm"]/div/div[1]/div/div/div/button')
    logger.info("Lupa clicada. Capturando descrições geradas na janela do formulário (2ª imagem)...")

    # Capturar itens renderizados na janela do formulário (conforme 2ª imagem)
    itens_form = _capturar_itens_do_formulario_phoenix(page)

    if not itens_form and itens_excel:
        itens_form = itens_excel
        logger.info("Utilizando %s item(ns) lidos diretamente da planilha Excel.", len(itens_form))

    if not itens_form:
        itens_form = [{"description": "", "mpn": ""}]

    descricao = itens_form[0]["description"]

    # AREA (só existe para alguns usuários)
    try:
        if page.locator("#areaId").count() > 0:
            page.select_option("#areaId", value="8eb210b1-bf04-4fbf-848d-5f3aeac9a0b8")
            logger.info("Area OK - Engenharia de Teste")
            page.wait_for_timeout(1500)
    except Exception as exc:
        logger.error("Erro Area: %s", exc)

    # PRIORITY
    try:
        page.wait_for_selector("#Priority", timeout=15000)
        page.select_option("#Priority", value="URGENT")
        logger.info("Priority OK")
    except Exception as exc:
        logger.error("Erro Priority: %s", exc)

    # COMPANY
    try:
        page.select_option("#CompanyId", value="65bb402c-cb05-441f-9a4f-df9eb51c7ae5")
        logger.info("Company OK")
    except Exception as exc:
        logger.error("Erro Company: %s", exc)

    # ITEM TYPE
    try:
        item_types = page.locator('select[id$="__ItemType"]')

        for i in range(item_types.count()):
            item_types.nth(i).select_option(value="D")

        logger.info("Item Type OK")
    except Exception as exc:
        logger.error("Erro Item Type: %s", exc)


    # ROHS
    try:
        rohs = page.locator('select[id$="__Rohs"]')

        for i in range(rohs.count()):
            rohs.nth(i).select_option(value="na")

        logger.info("ROHS OK")
    except Exception as exc:
        logger.error("Erro ROHS: %s", exc)


    
    # ORIGIN
    try:
        origins = page.locator('select[id$="__Origin"]')

        for i in range(origins.count()):
            origins.nth(i).select_option(value="IMPORTADO")

        logger.info("Origin OK")
    except Exception as exc:
        logger.error("Erro Origin: %s", exc)


    logger.info("=" * 50)
    logger.info("PREENCHER MANUALMENTE: Assunto Principal, Manufacturer")
    logger.info("=" * 50)
    logger.info(
        "Quando terminar, clique no botão 'Confirmar Solicitação' dentro do Phoenix. "
        "A automação vai aguardar a solicitação ser criada e capturar o ticket automaticamente."
    )

    # Em vez de um popup pedindo confirmação manual da ferramenta, aguardamos
    # o envio REAL do formulário no Phoenix (botão "Confirmar Solicitação"),
    # detectado pela mudança de URL. O navegador continua aberto - nada é
    # fechado até o ticket ser capturado.
    url_formulario = page.url
    try:
        page.wait_for_url(lambda url: url != url_formulario, timeout=600000)
    except PlaywrightTimeoutError:
        logger.warning(
            "Não detectei o envio do formulário (URL não mudou a tempo). A solicitação "
            "NÃO foi registrada no histórico/dashboard - nenhum ticket foi capturado."
        )
        return

    try:
        page.wait_for_load_state("domcontentloaded", timeout=15000)
    except Exception:
        pass

    logger.info("Solicitação enviada no Phoenix. Aguardando criação e captura do ticket...")

    descricao_principal = itens_form[0]["description"]
    dados_ticket = None

    for tentativa in range(1, 4):
        page.wait_for_timeout(2000)
        dados_ticket = _capturar_dados_minhas_solicitacoes(page, descricao_principal)
        if dados_ticket and dados_ticket.get("ticket"):
            logger.info("Ticket capturado com sucesso na tentativa %s/3: %s", tentativa, dados_ticket.get("ticket"))
            break
        logger.info("Aguardando inserção no Phoenix... tentativa %s/3", tentativa)

    ticket_num = (dados_ticket.get("ticket") if dados_ticket else "") or f"TKGP_PENDENTE_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    total_produtos = len(itens_form)
    import getpass
    usuario_win = getpass.getuser().upper()
    user_final = (os.environ.get("PHOENIX_TOOL_USER") or config.get("user") or usuario_win).strip().upper()

    for seq, item_info in enumerate(itens_form, start=1):
        desc_item = item_info.get("description", "")
        mpn_item = item_info.get("mpn", "")
        registro = criar_registro_descricao(
            desc_item,
            user_final,
            ticket=ticket_num,
            status=(dados_ticket.get("status") if dados_ticket else "ON GOING") or "ON GOING",
            data_abertura=(dados_ticket.get("data_abertura") if dados_ticket else None),
            hora_abertura=(dados_ticket.get("hora_abertura") if dados_ticket else None),
            produto_seq=seq,
            total_produtos=total_produtos,
            mpn=mpn_item,
        )
        logger.info(
            "Histórico salvo com sucesso no banco local. Linha: %s | Ticket: %s | Produto Seq: %s/%s",
            registro["linha"], registro.get("ticket"), seq, total_produtos
        )

        


# =====================================================
# CAPTURA: Ticket/Status/Data/Hora em "Minhas Solicitações" logo após o
# envio do formulário (a solicitação só é registrada no histórico/dashboard
# se o ticket for encontrado aqui)
# =====================================================

# Formato observado do ticket: letras seguidas de dígitos (ex: TKGP2026072728).
# AJUSTAR se o padrão real divergir.
TICKET_REGEX = re.compile(r"([A-Z]{2,}\d{6,})")


def _ir_para_minhas_solicitacoes(page: Page) -> bool:
    """Navega até 'Minhas Solicitações' a partir da tela atual do Phoenix com fallback para URL direta e login automático se necessário."""
    for selector in [
        "text=Minhas Solicitações",
        "text=MINHAS SOLICITAÇÕES",
        'a[href*="MyTasks"]',
        'a:has-text("Minhas Solicitações")',
    ]:
        if _esperar_elemento(page, selector, timeout=4000):
            try:
                page.click(selector)
                page.wait_for_load_state("domcontentloaded")
                logger.info("Tela 'Minhas Solicitações' aberta via menu.")
                return True
            except Exception:
                pass

    logger.info("Navegando diretamente via URL para 'Minhas Solicitações'...")
    try:
        page.goto("http://engenharia.br.flex.com/Phoenix/Task/MyTasks", wait_until="domcontentloaded", timeout=15000)
        if _esperar_elemento(page, "#FlexUser", timeout=3000):
            logger.info("Redirecionado para a tela de login. Efetuando autenticação...")
            config = carregar_config()
            _preencher_login(page, config)
            page.goto("http://engenharia.br.flex.com/Phoenix/Task/MyTasks", wait_until="domcontentloaded", timeout=15000)
        logger.info("Tela 'Minhas Solicitações' aberta via URL direta.")
        return True
    except Exception as exc:
        logger.error("Erro ao navegar para 'Minhas Solicitações': %s", exc)
        return False


def _extrair_dados_linha(texto: str) -> Optional[Dict[str, str]]:
    """Extrai ticket/assunto/status/data/hora do texto renderizado."""

    match_ticket = TICKET_REGEX.search(texto)
    ticket = match_ticket.group(1) if match_ticket else ""

    if not ticket:
        return None

    partes = [p.strip() for p in texto.splitlines() if p.strip()]

    status = ""

    for candidato in partes:
        candidato_upper = candidato.upper()

        if candidato_upper in (
            "ON GOING",
            "OPEN",
            "CLOSED",
            "FINALIZADO",
            "PENDING",
            "CANCELLED",
            "APPROVED",
        ):
            status = candidato_upper
            break

    data_hora_match = re.search(r"(\d{2}/\d{2}/\d{4})\s+([01]?\d|2[0-3]):([0-5]\d)", texto)
    if data_hora_match:
        data_abertura = data_hora_match.group(1)
        hora_abertura = f"{data_hora_match.group(2).zfill(2)}:{data_hora_match.group(3)}"
    else:
        data_match = re.search(r"(\d{2}/\d{2}/\d{4})", texto)
        data_abertura = data_match.group(1) if data_match else ""

        hora_match = re.search(r"\b([01]?\d|2[0-3]):([0-5]\d)\b", texto)
        hora_abertura = f"{hora_match.group(1).zfill(2)}:{hora_match.group(2)}" if hora_match else ""

    from services.database import limpar_descricao

    assunto = ""

    for candidato in partes:
        if (
            candidato == ticket
            or candidato.upper() == status
            or re.fullmatch(r"\d{2}/\d{2}/\d{4}", candidato)
            or re.fullmatch(r"\d{2}:\d{2}", candidato)
        ):
            continue

        if len(candidato) > len(assunto):
            assunto = candidato

    assunto = limpar_descricao(assunto)

    return {
        "ticket": ticket,
        "assunto": assunto,
        "status": status,
        "data_abertura": data_abertura,
        "hora_abertura": hora_abertura,
    }


def _capturar_dados_minhas_solicitacoes(page: Page, descricao: str) -> Optional[Dict[str, str]]:
    """
    Abre 'Minhas Solicitações' e captura Ticket/Status/Data/Hora.
    Tenta correspondência exata, case-insensitive e parcial de descrição.
    Se nenhuma coincidir, faz fallback automático para o 1º registro (topo da tabela).
    """

    if not _ir_para_minhas_solicitacoes(page):
        _debug_screenshot(page, "minhas_solicitacoes_nao_encontrada")
        return None

    try:
        page.wait_for_selector("table tbody tr", timeout=20000)
    except PlaywrightTimeoutError:
        logger.warning("Tabela de 'Minhas Solicitações' não carregou a tempo.")
        _debug_screenshot(page, "minhas_solicitacoes_sem_tabela")
        return None

    linhas = page.locator("table tbody tr")
    total = linhas.count()

    if total == 0:
        logger.warning("Tabela de 'Minhas Solicitações' está vazia.")
        return None

    descricao_normalizada = (descricao or "").strip().lower()
    alvo = None

    if descricao_normalizada:
        for i in range(total):
            candidata = linhas.nth(i)
            try:
                texto_linha = candidata.inner_text()
                texto_lower = texto_linha.lower()
            except Exception:
                continue

            # Correspondência exata ou insensível a maiúsculas
            if descricao_normalizada in texto_lower:
                logger.info("Linha correspondente encontrada (por substring).")
                alvo = candidata
                break

            # Correspondência por palavras-chave principais
            palavras = [p for p in re.split(r"\W+", descricao_normalizada) if len(p) > 3]
            if palavras and any(p in texto_lower for p in palavras):
                logger.info("Linha correspondente encontrada (por palavra-chave).")
                alvo = candidata
                break

    if alvo is None:
        logger.warning(
            "Solicitação não encontrada por correspondência exata de descrição (%s). "
            "Fazendo fallback automático para a primeira solicitação do topo da tabela.",
            descricao,
        )
        alvo = linhas.nth(0)

    dados = _extrair_dados_linha(alvo.inner_text())

    if dados is None:
        logger.warning(
            "Não encontrei um ticket no formato esperado na linha capturada."
        )
        return None

    logger.info(
        "Capturado em 'Minhas Solicitações' -> ticket=%s status=%s data=%s hora=%s",
        dados["ticket"],
        dados["status"],
        dados["data_abertura"],
        dados["hora_abertura"],
    )

    return dados


def sincronizar_solicitacoes_phoenix(headless: bool = True) -> None:
    """
    Entra no Phoenix, navega até 'Minhas Solicitações' e captura todos os tickets abertos
    que ainda não foram registrados no banco de dados local (SQLite). Roda 100% invisível por padrão.
    """
    import getpass
    config = carregar_config()
    usuario_win = getpass.getuser().upper()
    tool_user = (os.environ.get("PHOENIX_TOOL_USER") or config.get("user") or usuario_win).strip().upper()

    logger.info("Iniciando sincronização de solicitações abertas do Phoenix (headless=%s)...", headless)

    with sync_playwright() as p:
        browser, context, page = _abrir_phoenix(p, config, headless=headless)
        try:
            if not _ir_para_minhas_solicitacoes(page):
                logger.error("Não foi possível acessar 'Minhas Solicitações' para sincronização.")
                return

            try:
                page.wait_for_selector("table tbody tr", timeout=20000)
            except PlaywrightTimeoutError:
                logger.warning("Tabela de 'Minhas Solicitações' não carregou a tempo.")
                return

            linhas = page.locator("table tbody tr")
            total = linhas.count()
            logger.info("Total de solicitações encontradas na tabela do Phoenix: %s", total)

            novos_capturados = 0
            for i in range(total):
                try:
                    candidata = linhas.nth(i)
                    texto_linha = candidata.inner_text()
                    dados = _extrair_dados_linha(texto_linha)
                    if not dados or not dados.get("ticket"):
                        continue

                    ticket = dados["ticket"]
                    if existe_ticket(ticket):
                        logger.info("Ticket %s já existe no banco local. Ignorando.", ticket)
                        continue

                    # Ticket novo encontrado! Inserir no banco local
                    desc = dados.get("assunto") or f"Solicitação Phoenix {ticket}"
                    reg = criar_registro_descricao(
                        desc,
                        tool_user,
                        ticket=ticket,
                        status=dados.get("status") or "ON GOING",
                        data_abertura=dados.get("data_abertura") or None,
                        hora_abertura=dados.get("hora_abertura") or None,
                    )
                    novos_capturados += 1
                    logger.info(
                        "NOVA SOLICITAÇÃO CAPTURADA: Linha %s | Ticket %s | Descrição: %s",
                        reg["linha"], ticket, desc
                    )
                except Exception as exc:
                    logger.error("Erro ao processar linha %s de Minhas Solicitações: %s", i, exc)

            logger.info("Sincronização do Phoenix concluída! %s nova(s) solicitação(ões) registrada(s).", novos_capturados)
            try:
                messagebox.showinfo(
                    "Sincronização Phoenix",
                    f"Sincronização concluída com sucesso!\n\n"
                    f"Foram encontradas e capturadas {novos_capturados} nova(s) solicitação(ões) do Phoenix para o dashboard."
                )
            except Exception:
                pass

        except Exception as exc:
            logger.exception("Erro durante a sincronização do Phoenix: %s", exc)
        finally:
            try:
                browser.close()
            except Exception:
                pass



def abrir_home_phoenix() -> None:
    config = carregar_config()

    with sync_playwright() as p:
        browser, context, page = _abrir_phoenix(p, config)

        logger.info("Home Phoenix aberta.")
        input("Pressione ENTER para fechar...")

        browser.close()


# =====================================================
# FUNÇÃO 2: fluxo completo de solicitação P0032 (botão "Nova Solicitação")
# =====================================================

def nova_solicitacao_phoenix() -> None:
    config = carregar_config()

    with sync_playwright() as p:
        browser, context, page = _abrir_phoenix(p, config)

        # P0032
        logger.info("Abrindo P0032...")
        page.fill("#TaskName", "32")
        page.click("text=P0032")
        page.click('xpath=//*[@id="formId"]/div/div[2]/button')
        page.wait_for_load_state("domcontentloaded")
        logger.info("P0032 OK")

        # ESCOLHER TEMPLATE
        logger.info("Abrindo janela para selecionar o template (.xlsx)...")
        root = tk.Tk()
        root.withdraw()
        # Sem isso a janela de seleção de arquivo pode abrir atrás do
        # navegador (que acabou de receber foco) e parecer que "não carrega".
        root.attributes("-topmost", True)
        root.update()

        arquivo = filedialog.askopenfilename(
            parent=root,
            title="Selecione o Template",
            filetypes=[("Excel", "*.xlsx")]
        )

        root.destroy()

        if not arquivo:
            logger.info("Nenhum arquivo selecionado.")
            browser.close()
            return

        logger.info("Arquivo selecionado: %s", arquivo)

        # UPLOAD TEMPLATE
        page.set_input_files("#ExcelFile", arquivo)
        logger.info("Template carregado")

        # LUPA
        page.click('xpath=//*[@id="uploadForm"]/div/div[1]/div/div/div/button')
        logger.info("Lupa clicada")

        try:
            descricao = page.locator('[id$="__Description"]').input_value()
            logger.info("DESCRIPTION: %s", descricao)
        except Exception as exc:
            logger.error("Erro ao capturar Description: %s", exc)
            descricao = ""

        # AREA (só existe para alguns usuários)
        try:
            if page.locator("#areaId").count() > 0:
                page.select_option("#areaId", label="ENGENHARIA DE TESTE (JAGUARIÚNA)")
                logger.info("Area OK - Engenharia de Teste")
                page.wait_for_timeout(1500)
        except Exception as exc:
            logger.error("Erro Area: %s", exc)

        # PRIORITY
        try:
            page.wait_for_selector("#Priority", timeout=15000)
            page.select_option("#Priority", value="URGENT")
            logger.info("Priority OK")
        except Exception as exc:
            logger.error("Erro Priority: %s", exc)

        # COMPANY
        try:
            page.select_option("#CompanyId", value="65bb402c-cb05-441f-9a4f-df9eb51c7ae5")
            logger.info("Company OK")
        except Exception as exc:
            logger.error("Erro Company: %s", exc)

        # ITEM TYPE
        try:
            page.locator('select[id$="__ItemType"]').select_option(value="D")
            logger.info("Item Type OK")
        except Exception as exc:
            logger.error("Erro Item Type: %s", exc)

        # ROHS
        try:
            page.locator('select[id$="__Rohs"]').select_option(value="na")
            logger.info("ROHS OK")
        except Exception as exc:
            logger.error("Erro ROHS: %s", exc)

        # ORIGIN
        try:
            page.locator('select[id$="__Origin"]').select_option(value="IMPORTADO")
            logger.info("Origin OK")
        except Exception as exc:
            logger.error("Erro Origin: %s", exc)

        logger.info("=" * 50)
        logger.info("PREENCHER MANUALMENTE: Assunto Principal, Manufacturer")
        logger.info("=" * 50)
        logger.info("Confirme a solicitação no Phoenix e pressione ENTER para registrar no histórico.")

        input("Pressione ENTER para registrar e finalizar...")

        registro = criar_registro_descricao(descricao or "", os.environ.get("PHOENIX_TOOL_USER") or config.get("user", ""))
        logger.info("Histórico salvo. Linha criada: %s", registro["linha"])

        browser.close()


if __name__ == "__main__":
    # Filtrar argumentos de automação (ignorar --cred-user e --cred-pass)
    action_args = [a for a in sys.argv[1:] if a not in ("--cred-user", "--cred-pass") and not any(
        sys.argv[i] in ("--cred-user", "--cred-pass") and sys.argv[i + 1] == a
        for i in range(len(sys.argv) - 1)
    )]
    if action_args and action_args[0] == "home":
        abrir_home_phoenix()
    elif action_args and action_args[0] in ("capturar", "importar"):
        sincronizar_solicitacoes_phoenix()
    else:
        nova_solicitacao_phoenix()
